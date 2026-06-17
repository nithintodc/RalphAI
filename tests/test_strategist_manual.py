"""Strategist manual mode: FINANCIAL zip and legacy register upload."""

from pathlib import Path

import pytest

from agents.strategist.agent import run_manual_from_financial, run_manual_from_register
from agents.strategist.register_reco import (
    BOTTOM_ADS_SLOT_COUNT,
    build_recommendations_from_register,
)

ROOT = Path(__file__).resolve().parents[1]
SAMPLE_ZIP = ROOT / "sample_data/new-sample-data/financial_2026-03-01_2026-05-31_ZfBpu_2026-06-10T06-14-22Z.zip"


def test_strategist_manual_from_register_csv(tmp_path, monkeypatch):
    monkeypatch.setenv("TODC_DATA_DIR", str(tmp_path))
    slots = tmp_path / "slots.csv"
    slots.write_text(
        "Slot,Mon,Tue,Wed,Thu,Fri,Sat,Sun\n"
        "Lunch,1,2,3,4,5,6,7\n"
        "Dinner,8,9,10,11,12,13,14\n"
        "Breakfast,15,16,17,18,19,20,21\n",
        encoding="utf-8",
    )
    reg = tmp_path / "register.csv"
    reg.write_text(
        "Merchant Store ID,Day,Day part,Sales,Payouts,Orders,AOV\n"
        "100,Monday,Lunch,100,80,10,10\n"
        "100,Monday,Dinner,200,100,5,40\n",
        encoding="utf-8",
    )

    reporting_root = tmp_path / "reporting"
    reporting_root.mkdir()
    (reporting_root / "slots.csv").write_text(slots.read_text(encoding="utf-8"), encoding="utf-8")
    monkeypatch.setattr(
        "agents.strategist.agent.REPORTING_ROOT",
        reporting_root,
    )
    monkeypatch.setattr(
        "agents.strategist.agent.OUTPUT_ROOT",
        tmp_path / "Strategist",
    )

    out = run_manual_from_register("op_test", register_report_path=reg, business_name="Test Op")
    assert out["status"] == "success"
    assert out["operator_id"] == "op_test"
    assert len(out.get("recommended_campaigns") or []) >= 1
    assert out.get("campaign_mappings")
    plan_path = tmp_path / "operators" / "op_test" / "reports" / "marketing_plan.json"
    assert plan_path.is_file()
    assert Path(out["campaigns_xlsx"]).is_file()
    assert "downloads" in out["campaigns_xlsx"]
    assert out["campaigns_xlsx"].endswith(".xlsx")
    assert "combined_analysis_" in Path(out["campaigns_xlsx"]).name

    import openpyxl
    from shared.campaign_workbook_format import ADS_CAMPAIGN_MAPPINGS_SHEET, CAMPAIGN_MAPPINGS_SHEET

    wb = openpyxl.load_workbook(out["campaigns_xlsx"], read_only=True)
    assert CAMPAIGN_MAPPINGS_SHEET in wb.sheetnames
    assert ADS_CAMPAIGN_MAPPINGS_SHEET in wb.sheetnames
    wb.close()


def test_build_recommendations_from_register_smoke(tmp_path):
    slots = tmp_path / "slots.csv"
    slots.write_text("Slot,Mon\nLunch,1\n", encoding="utf-8")
    reg = tmp_path / "register.csv"
    reg.write_text(
        "Merchant Store ID,Day,Day part,Sales,Payouts,Orders,AOV\n"
        "1,Monday,Lunch,100,80,10,10\n",
        encoding="utf-8",
    )
    built = build_recommendations_from_register(reg, slots_csv=slots)
    assert built["slot_recommendations"]


@pytest.mark.skipif(not SAMPLE_ZIP.is_file(), reason="sample financial zip missing")
def test_strategist_manual_from_financial_zip(tmp_path, monkeypatch):
    monkeypatch.setenv("TODC_DATA_DIR", str(tmp_path))
    reporting_root = ROOT / "reporting_browser_use"
    monkeypatch.setattr("agents.strategist.agent.REPORTING_ROOT", reporting_root)
    monkeypatch.setattr("agents.strategist.agent.OUTPUT_ROOT", tmp_path / "Strategist")

    out = run_manual_from_financial(
        "op_financial",
        financial_zip_path=SAMPLE_ZIP,
        business_name="Financial Test",
    )
    assert out["status"] == "success"
    assert out["input_type"] == "financial"
    combined = Path(out["combined_analysis"])
    assert combined.is_file()
    assert "downloads" in str(combined)
    assert combined.name.startswith("combined_analysis_")
    assert Path(out["slot_info_csv"]).is_file()
    ads_rows = [
        r for r in (out.get("slot_recommendations") or [])
        if str(r.get("action") or "").lower() in ("promo+ads", "ads")
    ]
    by_store: dict[str, int] = {}
    for row in ads_rows:
        sid = str(row.get("store_id") or "")
        by_store[sid] = by_store.get(sid, 0) + 1
    for store_id, count in by_store.items():
        assert count <= BOTTOM_ADS_SLOT_COUNT, f"store {store_id} has {count} ads slots"
