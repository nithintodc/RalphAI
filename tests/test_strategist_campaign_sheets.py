"""Tests for Strategist campaign workbook discovery and parsing."""

from __future__ import annotations

from pathlib import Path

import pytest

from shared.campaign_workbook_format import CAMPAIGN_MAPPINGS_SHEET, combined_analysis_filename
from shared.strategist_campaign_sheets import (
    find_latest_strategist_run_dir,
    load_ads_rows,
    load_offers_combos,
    load_offers_combos_from_path,
    safe_dirname,
    update_campaign_workbook_status,
    update_slot_info_campaign_status,
    write_strategist_campaign_statuses,
)


def test_safe_dirname_strips_invalid_chars():
    assert safe_dirname("Acme / Pizza") == "Acme - Pizza"


def test_find_latest_strategist_run_dir(tmp_path, monkeypatch):
    root = tmp_path / "Strategist" / safe_dirname("Test Operator")
    older = root / "20260101_120000"
    newer = root / "20260201_120000"
    older.mkdir(parents=True)
    newer.mkdir(parents=True)
    (older / "campaigns.xlsx").write_bytes(b"old")
    (newer / combined_analysis_filename(timestamp="20260201_120001")).write_bytes(b"new")

    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: "Test Operator" if oid == "op1" else oid,
    )

    found = find_latest_strategist_run_dir("op1")
    assert found == newer


def test_load_offers_combos_from_combined_analysis(tmp_path, monkeypatch):
    import openpyxl

    root = tmp_path / "Strategist" / safe_dirname("Bican")
    run_dir = root / "20260301_100000"
    run_dir.mkdir(parents=True)
    wb_path = run_dir / combined_analysis_filename(timestamp="20260301_100000")
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(CAMPAIGN_MAPPINGS_SHEET)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    headers = ["Store ID", "Store Name", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="123")
    ws.cell(row=2, column=2, value="Store A")
    ws.cell(row=2, column=3, value=15)
    ws.cell(row=2, column=4, value="1,2,8")
    ws.cell(row=2, column=5, value="TODC-123-$15")
    ws.cell(row=2, column=6, value="Pending")
    ws.cell(row=3, column=1, value="456")
    ws.cell(row=3, column=4, value="3")
    ws.cell(row=3, column=5, value="TODC-456-$10")
    ws.cell(row=3, column=6, value="Successful")
    wb.save(wb_path)

    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: "Bican",
    )

    combos, workbook, found_run = load_offers_combos("Bican")
    assert workbook == wb_path
    assert found_run == run_dir
    assert len(combos) == 1
    assert combos[0]["store_id"] == "123"
    assert combos[0]["slot_tags"] == [1, 2, 8]
    assert combos[0]["min_subtotal"] == 15


def test_load_offers_combos_from_legacy_workbook(tmp_path, monkeypatch):
    import openpyxl

    root = tmp_path / "Strategist" / safe_dirname("Bican")
    run_dir = root / "20260301_100000"
    run_dir.mkdir(parents=True)
    wb_path = run_dir / "campaigns.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offers Campaigns"
    headers = ["Store ID", "Store Name", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="123")
    ws.cell(row=2, column=2, value="Store A")
    ws.cell(row=2, column=3, value=15)
    ws.cell(row=2, column=4, value="1,2,8")
    ws.cell(row=2, column=5, value="TODC-123-$15")
    ws.cell(row=2, column=6, value="Pending")
    wb.save(wb_path)

    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: "Bican",
    )

    combos, workbook, _ = load_offers_combos("Bican")
    assert workbook == wb_path
    assert len(combos) == 1
    assert combos[0]["store_id"] == "123"


def test_load_offers_combos_retries_skipped_duplicate(tmp_path, monkeypatch):
    import openpyxl

    root = tmp_path / "Strategist" / safe_dirname("Bican")
    run_dir = root / "20260301_100000"
    run_dir.mkdir(parents=True)
    wb_path = run_dir / combined_analysis_filename(timestamp="20260301_100000")
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(CAMPAIGN_MAPPINGS_SHEET)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    headers = ["Store ID", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="10661")
    ws.cell(row=2, column=2, value=20)
    ws.cell(row=2, column=3, value="2,5,6")
    ws.cell(row=2, column=4, value="TODC-10661-$20")
    ws.cell(row=2, column=5, value="Successful")
    ws.cell(row=3, column=1, value="10661")
    ws.cell(row=3, column=2, value=30)
    ws.cell(row=3, column=3, value="28,31,32")
    ws.cell(row=3, column=4, value="TODC-10661-$30")
    ws.cell(row=3, column=5, value="Skipped (duplicate)")
    ws.cell(row=4, column=1, value="11399")
    ws.cell(row=4, column=2, value=25)
    ws.cell(row=4, column=3, value="1,2,3")
    ws.cell(row=4, column=4, value="TODC-11399-$25")
    ws.cell(row=4, column=5, value="Pending")
    wb.save(wb_path)

    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: "Bican",
    )

    combos, _, _ = load_offers_combos("Bican")
    names = {c["campaign_name"] for c in combos}
    assert "TODC-10661-$20" not in names
    assert "TODC-10661-$30" in names
    assert "TODC-11399-$25" in names
    assert len(combos) == 2


def test_load_ads_rows_from_combined_analysis(tmp_path, monkeypatch):
    import openpyxl

    from shared.campaign_workbook_format import ADS_CAMPAIGN_MAPPINGS_SHEET

    root = tmp_path / "Strategist" / safe_dirname("Bican")
    run_dir = root / "20260301_100000"
    run_dir.mkdir(parents=True)
    wb_path = run_dir / combined_analysis_filename(timestamp="20260301_100000")
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(ADS_CAMPAIGN_MAPPINGS_SHEET)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    headers = [
        "Store ID",
        "Store Name",
        "Minimum Bid",
        "Weekly Budget",
        "Slot Tags",
        "Campaign Name",
        "Status",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="99")
    ws.cell(row=2, column=2, value="Store Z")
    ws.cell(row=2, column=3, value=3)
    ws.cell(row=2, column=4, value=140)
    ws.cell(row=2, column=5, value="5,6,7")
    ws.cell(row=2, column=6, value="TODC-ADS-99")
    ws.cell(row=2, column=7, value="Pending")
    wb.save(wb_path)

    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: "Bican",
    )

    rows, workbook, _ = load_ads_rows("Bican")
    assert workbook == wb_path
    assert len(rows) == 1
    assert rows[0]["store_id"] == "99"
    assert rows[0]["slot_tags"] == [5, 6, 7]
    assert rows[0]["bid_strategy"] == 3.0
    assert rows[0]["budget"] == 140.0


def test_load_offers_combos_from_csv_path(tmp_path):
    csv_path = tmp_path / "offers.csv"
    csv_path.write_text(
        "Merchant store ID,Slots,Minimum Subtotal,Campaign Name,Status\n"
        "123,1;2;8,15,TODC-123-$15,Pending\n"
        "456,3,10,TODC-456-$10,Successful\n",
        encoding="utf-8",
    )
    combos = load_offers_combos_from_path(csv_path)
    assert len(combos) == 1
    assert combos[0]["store_id"] == "123"
    assert combos[0]["slot_tags"] == [1, 2, 8]
    assert combos[0]["min_subtotal"] == 15


def test_update_campaign_workbook_and_slot_info_status(tmp_path):
    import csv
    import openpyxl

    wb_path = tmp_path / combined_analysis_filename(timestamp="20260301_100000")
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(CAMPAIGN_MAPPINGS_SHEET)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    headers = ["Store ID", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="10661")
    ws.cell(row=2, column=2, value=20)
    ws.cell(row=2, column=3, value="2,5,6")
    ws.cell(row=2, column=4, value="TODC-10661-$20")
    ws.cell(row=2, column=5, value="Pending")
    wb.save(wb_path)

    slot_path = tmp_path / "slot_info.csv"
    slot_path.write_text(
        "Store ID,Store Name,Day,Slot,Slot Tag,Orders,Sales,AOV,Campaign Type,Campaign Name,Ads Campaign Name,Minimum Subtotal,Minimum Bid,Status\n"
        "10661,Store,Sunday,Afternoon,28,1,30,30,Offer,TODC-10661-$30,,30,,Pending\n"
        "10661,Store,Wednesday,Dinner,31,1,30,30,Offer,TODC-10661-$30,,30,,Pending\n",
        encoding="utf-8",
    )

    item = {
        "store_id": "10661",
        "campaign_name": "TODC-10661-$20",
        "slot_tags": [2, 5, 6],
    }
    write_strategist_campaign_statuses(wb_path, slot_path, item, "Successful", kind="offers")

    wb2 = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws2 = wb2[CAMPAIGN_MAPPINGS_SHEET]
    assert ws2.cell(row=2, column=5).value == "Successful"
    wb2.close()

    updated = update_slot_info_campaign_status(
        slot_path,
        "TODC-10661-$30",
        "Skipped (duplicate)",
        store_id="10661",
        slot_tags=[28, 31],
    )
    assert updated == 2
    with slot_path.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    by_tag = {int(r["Slot Tag"]): r["Status"] for r in rows}
    assert by_tag[28] == "Skipped (duplicate)"
    assert by_tag[31] == "Skipped (duplicate)"


def test_update_campaign_workbook_status_missing_file():
    assert update_campaign_workbook_status("/no/such/file.xlsx", "TODC-1", "Failed") is False


def test_load_offers_combos_raises_when_no_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.STRATEGIST_ROOT",
        tmp_path / "Strategist",
    )
    monkeypatch.setattr(
        "shared.strategist_campaign_sheets.resolve_business_name",
        lambda oid: oid,
    )
    with pytest.raises(FileNotFoundError):
        load_offers_combos("missing-operator")
