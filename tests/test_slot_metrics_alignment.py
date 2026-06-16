"""Verify Strategist manual (financial) and reporting_browser_use slot campaigns align."""

from __future__ import annotations

import csv
import sys
import tempfile
import zipfile
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]
SAMPLE_ZIP = ROOT / "sample_data/new-sample-data/financial_2026-03-01_2026-05-31_ZfBpu_2026-06-10T06-14-22Z.zip"
SLOTS_CSV = ROOT / "reporting_browser_use/slots.csv"


def _offer_campaigns_from_workbook(path: Path) -> dict[tuple[str, int], set[int]]:
    xl = pd.ExcelFile(path)
    sheet = next(s for s in xl.sheet_names if "offer" in s.lower())
    df = pd.read_excel(xl, sheet_name=sheet)
    df.columns = df.columns.astype(str).str.strip()
    out: dict[tuple[str, int], set[int]] = {}
    for _, row in df.iterrows():
        store = str(row.get("Store ID") or "").strip()
        try:
            min_sub = int(round(float(row.get("Minimum Subtotal") or 0)))
        except (TypeError, ValueError):
            continue
        raw_tags = str(row.get("Slot Tags") or "")
        tags = {int(t.strip()) for t in raw_tags.replace("，", ",").split(",") if t.strip()}
        if store and min_sub > 0 and tags:
            out[(store, min_sub)] = tags
    return out


def _offer_campaigns_from_slot_info(path: Path) -> dict[tuple[str, int], set[int]]:
    out: dict[tuple[str, int], set[int]] = {}
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ctype = str(row.get("Campaign Type") or "")
            if "Offer" not in ctype:
                continue
            store = str(row.get("Store ID") or "").strip()
            name = str(row.get("Campaign Name") or "")
            if not store or not name.startswith("TODC-") or "$" not in name:
                continue
            try:
                min_sub = int(name.rsplit("$", 1)[-1])
            except ValueError:
                continue
            tag_raw = str(row.get("Slot Tag") or "").strip()
            if not tag_raw:
                continue
            tag = int(float(tag_raw))
            out.setdefault((store, min_sub), set()).add(tag)
    return out


def _reporting_offer_campaigns(combined_xlsx: Path) -> dict[tuple[str, int], set[int]]:
    from shared.reporting_imports import import_reporting_agents_module

    get_campaign_combos_from_slots_and_combined = import_reporting_agents_module(
        "campaign_params"
    ).get_campaign_combos_from_slots_and_combined

    combos = get_campaign_combos_from_slots_and_combined(SLOTS_CSV, combined_xlsx)
    out: dict[tuple[str, int], set[int]] = {}
    for combo in combos:
        store = str(combo.get("store_id") or "").strip()
        min_sub = int(combo.get("min_subtotal") or 0)
        tags = set(int(t) for t in combo.get("slot_tags") or [])
        if store and min_sub > 0 and tags:
            out[(store, min_sub)] = tags
    return out


def _write_combined_workbook_from_financial(zip_path: Path, out_dir: Path) -> Path:
    from shared.reporting_imports import import_reporting_agents_module

    analysis_run = import_reporting_agents_module("analysis_agent").run

    sheets = analysis_run(zip_path, out_dir, "2026-03-01", "2026-05-31", write_file=False)
    assert sheets

    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets:
        if df is None or df.empty:
            continue
        ws = wb.create_sheet(sheet_name[:31])
        ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=12)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)
    out_path = out_dir / "combined_analysis_test.xlsx"
    wb.save(out_path)
    return out_path


@pytest.mark.skipif(not SAMPLE_ZIP.is_file(), reason="sample financial zip missing")
def test_strategist_financial_matches_reporting_campaigns():
    from agents.strategist.campaign_workbook import write_campaigns_workbook_from_per_store
    from agents.strategist.register_reco import (
        ADS_MIN_BID,
        ADS_WEEKLY_BUDGET,
        build_recommendations_from_financial,
    )

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)

        built = build_recommendations_from_financial(
            SAMPLE_ZIP,
            slots_csv=SLOTS_CSV,
            start_date="2026-03-01",
            end_date="2026-05-31",
        )
        strategist_dir = tmp_path / "strategist"
        write_campaigns_workbook_from_per_store(
            strategist_dir,
            built["per_store"],
            built["store_names"],
            ads_min_bid=ADS_MIN_BID,
            ads_weekly_budget=ADS_WEEKLY_BUDGET,
        )

        reporting_dir = tmp_path / "reporting"
        reporting_dir.mkdir()
        combined = _write_combined_workbook_from_financial(SAMPLE_ZIP, reporting_dir)

        strategist_offers = _offer_campaigns_from_workbook(strategist_dir / "campaigns.xlsx")
        reporting_offers = _reporting_offer_campaigns(combined)

        assert strategist_offers == reporting_offers, (
            f"offer campaigns differ: "
            f"only strategist={set(strategist_offers) - set(reporting_offers)} "
            f"only reporting={set(reporting_offers) - set(strategist_offers)}"
        )

        slot_info_offers = _offer_campaigns_from_slot_info(strategist_dir / "slot_info.csv")
        assert slot_info_offers == strategist_offers
