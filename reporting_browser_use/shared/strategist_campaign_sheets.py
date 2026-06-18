"""Locate and parse Strategist campaign workbooks (Offers / Ads sheets)."""

from __future__ import annotations

import csv
import logging
import re
from pathlib import Path
from typing import Any, Literal

from shared.subprocess_env import repo_root

logger = logging.getLogger(__name__)

STRATEGIST_ROOT = repo_root() / "data" / "Strategist"

CampaignKind = Literal["offers", "ads"]

_OFFERS_SHEET_NAMES = ("offers campaigns", "offers")
_ADS_SHEET_NAMES = ("ads campaigns", "ads")


def safe_dirname(name: str) -> str:
    """Sanitize operator/business name for Strategist output folders."""
    safe = (name or "operator").strip()
    for ch in ("/", "\\", ":", "*", "?", '"', "<", ">", "|"):
        safe = safe.replace(ch, "-")
    safe = safe.strip(". ")
    return (safe[:100] if len(safe) > 100 else safe) or "operator"


def resolve_business_name(operator_id: str) -> str:
    return (operator_id or "").strip()


def _operator_roots(operator_id: str) -> list[Path]:
    """Candidate Strategist directories for an operator (business name + raw id)."""
    names: list[str] = []
    business = resolve_business_name(operator_id)
    for n in (business, (operator_id or "").strip()):
        if n and n not in names:
            names.append(n)
    roots: list[Path] = []
    for n in names:
        p = STRATEGIST_ROOT / safe_dirname(n)
        if p not in roots:
            roots.append(p)
    return roots


def find_latest_strategist_run_dir(operator_id: str) -> Path | None:
    """Latest timestamp folder under data/Strategist/<operator>/ with a campaign workbook."""
    candidates: list[tuple[str, Path]] = []
    for root in _operator_roots(operator_id):
        if not root.is_dir():
            continue
        for child in root.iterdir():
            if not child.is_dir():
                continue
            if (child / "campaigns.xlsx").is_file() or (child / "marketing_plan.xlsx").is_file():
                candidates.append((child.name, child))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def find_campaign_workbook(operator_id: str) -> tuple[Path, Path]:
    """
    Return (workbook_path, run_dir) for the latest Strategist output for this operator.

    Prefers campaigns.xlsx (auto Strategist) over marketing_plan.xlsx (manual).
    """
    run_dir = find_latest_strategist_run_dir(operator_id)
    if run_dir is None:
        raise FileNotFoundError(
            f"No Strategist campaign workbook found under {STRATEGIST_ROOT} for operator {operator_id!r}. "
            "Run Strategist first."
        )
    auto = run_dir / "campaigns.xlsx"
    manual = run_dir / "marketing_plan.xlsx"
    if auto.is_file():
        return auto, run_dir
    if manual.is_file():
        return manual, run_dir
    raise FileNotFoundError(f"No campaigns.xlsx or marketing_plan.xlsx in {run_dir}")


def _pick_sheet(sheet_names: list[str], kind: CampaignKind) -> str:
    lowered = {s: s.strip().lower() for s in sheet_names}
    targets = _OFFERS_SHEET_NAMES if kind == "offers" else _ADS_SHEET_NAMES
    for target in targets:
        for original, low in lowered.items():
            if low == target:
                return original
    raise ValueError(
        f'Workbook has no {"Offers" if kind == "offers" else "Ads"} sheet '
        f'(looked for: {", ".join(targets)}). Sheets: {sheet_names}'
    )


def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").strip().lower())


def _row_val(row: dict[str, Any], *candidates: str) -> Any:
    if not row:
        return None
    by_norm = {_norm_col(k): v for k, v in row.items()}
    for c in candidates:
        key = _norm_col(c)
        if key in by_norm:
            return by_norm[key]
    return None


def _parse_slot_tags(raw: Any) -> list[int]:
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        parts = raw
    else:
        parts = re.split(r"[,;\s]+", str(raw).strip())
    out: list[int] = []
    for p in parts:
        s = str(p).strip()
        if not s:
            continue
        try:
            tag = int(float(s))
        except ValueError:
            continue
        if not (1 <= tag <= 42):
            logger.warning("_parse_slot_tags: tag %s is outside valid range 1–42, skipping", s)
            continue
        out.append(tag)
    return sorted(set(out))


def _status_skip(status: Any) -> bool:
    """Only completed campaigns are skipped on reload; duplicates are retried."""
    s = str(status or "").strip().lower()
    return s in ("successful", "success")


def _read_sheet_rows(workbook: Path, kind: CampaignKind) -> list[dict[str, Any]]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas is required to read Strategist campaign workbooks") from exc

    xl = pd.ExcelFile(workbook)
    sheet = _pick_sheet(list(xl.sheet_names), kind)
    df = pd.read_excel(xl, sheet_name=sheet)
    if df is None or df.empty:
        return []
    rows: list[dict[str, Any]] = []
    for rec in df.to_dict(orient="records"):
        cleaned = {k: ("" if pd.isna(v) else v) for k, v in rec.items()}
        rows.append(cleaned)
    return rows


def load_offers_combos(operator_id: str) -> tuple[list[dict[str, Any]], Path, Path]:
    """Parse offer rows into browser-use combo dicts (reporting_browser_use format)."""
    workbook, run_dir = find_campaign_workbook(operator_id)
    combos: list[dict[str, Any]] = []
    for row in _read_sheet_rows(workbook, "offers"):
        combo = _parse_offer_row(row, workbook=workbook)
        if combo:
            combos.append(combo)
    if not combos:
        raise ValueError(f"No pending offer campaigns in {workbook}")
    return combos, workbook, run_dir


def _parse_offer_row(row: dict[str, Any], *, workbook: Path | None = None) -> dict[str, Any] | None:
    status = _row_val(row, "Status")
    if _status_skip(status):
        return None
    store_id = str(
        _row_val(row, "Store ID", "Merchant store ID", "Merchant Store ID") or ""
    ).strip()
    if not store_id:
        return None
    slot_tags = _parse_slot_tags(_row_val(row, "Slot Tags", "Slots"))
    if not slot_tags:
        return None
    _raw_sub = _row_val(row, "Minimum Subtotal", "Min subtotal")
    if _raw_sub is None or str(_raw_sub).strip() in ("", "nan", "None"):
        logger.error("_parse_offer_row: missing Minimum Subtotal for store %s — rejecting row", store_id)
        return None
    try:
        min_sub = int(round(float(str(_raw_sub).strip().replace("$", "").replace(",", ""))))
    except (TypeError, ValueError):
        logger.error("_parse_offer_row: unparseable Minimum Subtotal %r for store %s — rejecting row", _raw_sub, store_id)
        return None
    campaign_name = str(
        _row_val(row, "Campaign Name", "Campaign name")
        or f"TODC-{store_id}-${min_sub}"
    ).strip()
    combo: dict[str, Any] = {
        "store_id": store_id,
        "store_name": str(_row_val(row, "Store Name", "Store name") or "").strip(),
        "min_subtotal": min_sub,
        "slot_tags": slot_tags,
        "campaign_name": campaign_name,
        "status": str(status or "Pending"),
        "_source_row": row,
    }
    if workbook is not None:
        combo["_source_workbook"] = str(workbook)
    return combo


def load_offers_combos_from_path(sheet_path: Path) -> list[dict[str, Any]]:
    """Parse offer combos from an uploaded CSV/Excel (Manual Offers mode)."""
    path = Path(sheet_path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".csv":
        try:
            import pandas as pd
        except ImportError as exc:
            raise RuntimeError("pandas is required") from exc
        df = pd.read_csv(path)
        raw_rows = df.to_dict(orient="records")
    else:
        raw_rows = _read_sheet_rows(path, "offers")

    combos: list[dict[str, Any]] = []
    for row in raw_rows:
        # NaN-clean all values (mirrors Excel path; CSV path previously skipped this)
        cleaned = {k: ("" if (hasattr(v, "__class__") and v.__class__.__name__ == "float" and v != v) else v) for k, v in row.items()}
        # Normalize store ID: "14351.0" → "14351", "nan" → reject
        for sid_key in ("Store ID", "Merchant store ID", "Merchant Store ID"):
            if sid_key in cleaned:
                raw_sid = str(cleaned[sid_key]).strip()
                if raw_sid.lower() in ("nan", "none", ""):
                    cleaned[sid_key] = ""
                else:
                    try:
                        cleaned[sid_key] = str(int(float(raw_sid)))
                    except (ValueError, TypeError):
                        pass
        combo = _parse_offer_row(cleaned)
        if combo:
            combos.append(combo)
    if not combos:
        raise ValueError(f"No offer rows found in {path}")
    return combos


def load_ads_rows(operator_id: str) -> tuple[list[dict[str, Any]], Path, Path]:
    """Parse ads rows for sponsored-listing browser automation."""
    workbook, run_dir = find_campaign_workbook(operator_id)
    rows_out: list[dict[str, Any]] = []
    for row in _read_sheet_rows(workbook, "ads"):
        status = _row_val(row, "Status")
        if _status_skip(status):
            continue
        store_id = str(
            _row_val(
                row,
                "Store ID",
                "Merchant store ID",
                "Merchant Store ID",
                "Merchant store ID",
            )
            or ""
        ).strip()
        if not store_id:
            continue
        slot_tags = _parse_slot_tags(_row_val(row, "Slot Tags", "Slots"))
        if not slot_tags:
            continue
        _raw_bid = _row_val(row, "Minimum Bid", "Bid strategy", "Bid Strategy")
        if _raw_bid is None or str(_raw_bid).strip() in ("", "nan", "None"):
            logger.error("load_ads_rows: missing Minimum Bid for store %s — rejecting row", store_id)
            continue
        try:
            bid = float(str(_raw_bid).strip().replace("$", "").replace(",", ""))
        except (TypeError, ValueError):
            logger.error("load_ads_rows: unparseable Minimum Bid %r for store %s — rejecting row", _raw_bid, store_id)
            continue
        _raw_budget = _row_val(row, "Weekly Budget", "Budget", "weekly_budget", "Weekly budget")
        try:
            budget = float(str(_raw_budget).strip().replace("$", "").replace(",", "")) if _raw_budget and str(_raw_budget).strip() not in ("", "nan", "None") else 0.0
        except (TypeError, ValueError):
            logger.error("load_ads_rows: unparseable Weekly Budget %r for store %s — rejecting row", _raw_budget, store_id)
            continue
        if budget == 0.0:
            logger.warning("load_ads_rows: budget is 0 for store %s — LLM will NOT pick spend (row rejected)", store_id)
            continue
        campaign_name = str(
            _row_val(row, "Campaign Name", "Campaign name") or f"TODC-ADS-{store_id}"
        ).strip()
        rows_out.append(
            {
                "store_id": store_id,
                "store_name": str(_row_val(row, "Store Name", "Store name") or "").strip(),
                "slot_tags": slot_tags,
                "bid_strategy": bid,
                "budget": budget,
                "campaign_name": campaign_name,
                "status": str(status or "Pending"),
                "_source_workbook": str(workbook),
                "_source_row": row,
            }
        )
    if not rows_out:
        raise ValueError(f"No pending ads campaigns in {workbook}")
    return rows_out, workbook, run_dir


def load_ads_rows_from_path(sheet_path: Path) -> list[dict[str, Any]]:
    """Parse ads rows from an uploaded CSV/Excel (Manual Ads mode)."""
    path = Path(sheet_path)
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".csv":
        try:
            import pandas as pd
        except ImportError as exc:
            raise RuntimeError("pandas is required") from exc
        df = pd.read_csv(path)
        raw_rows = df.to_dict(orient="records")
    else:
        raw_rows = _read_sheet_rows(path, "ads")

    rows_out: list[dict[str, Any]] = []
    for row in raw_rows:
        cleaned = {k: v for k, v in row.items()}
        store_id = str(
            _row_val(cleaned, "Store ID", "Merchant store ID", "Merchant Store ID") or ""
        ).strip()
        if not store_id:
            continue
        slot_tags = _parse_slot_tags(_row_val(cleaned, "Slots", "Slot Tags"))
        if not slot_tags:
            continue
        try:
            bid = float(_row_val(cleaned, "Bid strategy", "Bid Strategy", "Minimum Bid") or 3)
        except (TypeError, ValueError):
            bid = 3.0
        try:
            budget = float(
                _row_val(cleaned, "Weekly Budget", "Budget", "weekly_budget", "Weekly budget") or 0
            )
        except (TypeError, ValueError):
            budget = 0.0
        campaign_name = str(
            _row_val(cleaned, "Campaign Name", "Campaign name") or f"TODC-ADS-{store_id}"
        ).strip()
        rows_out.append(
            {
                "store_id": store_id,
                "store_name": str(_row_val(cleaned, "Store Name", "Store name") or "").strip(),
                "slot_tags": slot_tags,
                "bid_strategy": bid,
                "budget": budget,
                "campaign_name": campaign_name,
            }
        )
    if not rows_out:
        raise ValueError(f"No ads rows found in {path}")
    return rows_out


def resolve_slot_info_csv(workbook_path: Path | str) -> Path | None:
    """Return sibling slot_info.csv for a Strategist campaigns workbook, if present."""
    candidate = Path(workbook_path).resolve().parent / "slot_info.csv"
    return candidate if candidate.is_file() else None


def _workbook_header_map(ws) -> dict[str, int]:
    """Map normalized header → 1-based column index for row 1."""
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip():
            headers[_norm_col(str(cell.value))] = cell.column
    return headers


def update_campaign_workbook_status(
    workbook_path: Path | str,
    campaign_name: str,
    status: str,
    *,
    kind: CampaignKind = "offers",
) -> bool:
    """
    Write ``status`` to the Offers/Ads sheet row whose Campaign Name matches.

    Returns True when a row was updated.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed; cannot update campaign workbook status")
        return False

    path = Path(workbook_path)
    if not path.is_file():
        logger.warning("Campaign workbook not found for status update: %s", path)
        return False

    name = str(campaign_name or "").strip()
    if not name:
        return False

    try:
        wb = openpyxl.load_workbook(path, read_only=False)
    except Exception as exc:
        logger.warning("Could not open %s for status update: %s", path, exc)
        return False

    try:
        sheet_name = _pick_sheet(list(wb.sheetnames), kind)
        ws = wb[sheet_name]
        headers = _workbook_header_map(ws)
        name_col = headers.get(_norm_col("Campaign Name")) or headers.get(_norm_col("Campaign name"))
        status_col = headers.get(_norm_col("Status"))
        if not name_col or not status_col:
            logger.warning(
                "Campaign workbook %s missing Campaign Name or Status column (sheet %s)",
                path.name,
                sheet_name,
            )
            return False

        for row_idx in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=name_col).value
            if str(cell_name or "").strip() == name:
                ws.cell(row=row_idx, column=status_col).value = status
                wb.save(path)
                logger.info("Strategist workbook: %s → %s (%s)", name, status, path.name)
                return True

        logger.warning("Campaign '%s' not found in %s for status update", name, path.name)
        return False
    except Exception as exc:
        logger.warning("Status update failed for %s in %s: %s", name, path, exc)
        return False
    finally:
        wb.close()


_SLOT_INFO_COLUMNS = [
    "Store ID",
    "Store Name",
    "Day",
    "Slot",
    "Slot Tag",
    "Orders",
    "Sales",
    "AOV",
    "Campaign Type",
    "Campaign Name",
    "Ads Campaign Name",
    "Minimum Subtotal",
    "Minimum Bid",
    "Status",
]


def update_slot_info_campaign_status(
    slot_info_path: Path | str,
    campaign_name: str,
    status: str,
    *,
    store_id: str | None = None,
    slot_tags: list[int] | None = None,
    kind: CampaignKind = "offers",
) -> int:
    """
    Update Status in slot_info.csv for rows assigned to ``campaign_name``.

    When ``slot_tags`` is provided, only rows with matching Slot Tag are updated.
    Returns the number of rows changed.
    """
    path = Path(slot_info_path)
    if not path.is_file():
        logger.warning("slot_info.csv not found for status update: %s", path)
        return 0

    name = str(campaign_name or "").strip()
    if not name:
        return 0

    tag_filter: set[int] | None = None
    if slot_tags:
        tag_filter = {int(t) for t in slot_tags}

    rows: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))

    updated = 0
    for row in rows:
        row_name = str(row.get("Campaign Name") or "").strip()
        row_type = str(row.get("Campaign Type") or "").strip()
        name_matches = row_name == name
        ads_slot_match = (
            kind == "ads"
            and tag_filter is not None
            and "Ads" in row_type
        )
        if not name_matches and not ads_slot_match:
            continue
        if store_id and str(row.get("Store ID") or "").strip() != str(store_id).strip():
            continue
        if tag_filter is not None:
            try:
                row_tag = int(float(str(row.get("Slot Tag") or "").strip()))
            except ValueError:
                continue
            if row_tag not in tag_filter:
                continue
        if str(row.get("Status") or "").strip() == status:
            continue
        row["Status"] = status
        updated += 1

    if updated:
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=_SLOT_INFO_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        logger.info("slot_info.csv: %s → %s (%d row(s))", name, status, updated)

    return updated


def write_strategist_campaign_statuses(
    workbook_path: Path | str,
    slot_info_path: Path | str | None,
    item: dict[str, Any],
    status: str,
    *,
    kind: CampaignKind = "offers",
) -> None:
    """Persist one campaign result to campaigns.xlsx and optional slot_info.csv."""
    campaign_name = str(item.get("campaign_name") or "").strip()
    if not campaign_name:
        return

    update_campaign_workbook_status(workbook_path, campaign_name, status, kind=kind)

    slot_path = slot_info_path or resolve_slot_info_csv(workbook_path)
    if not slot_path:
        return

    store_id = str(item.get("store_id") or "").strip() or None
    slot_tags = item.get("slot_tags")
    if slot_tags is not None and not isinstance(slot_tags, list):
        slot_tags = _parse_slot_tags(slot_tags)

    update_slot_info_campaign_status(
        slot_path,
        campaign_name,
        status,
        store_id=store_id,
        slot_tags=slot_tags if slot_tags else None,
        kind=kind,
    )
