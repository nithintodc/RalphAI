"""
CombinedReportAgent: build one Excel workbook with all financial + marketing analysis sheets.
Can write from in-memory sheet data (DataFrames) or from existing xlsx files.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

logger = logging.getLogger(__name__)

# Sheet name and headers for campaign mappings (store → min_subtotal → slot tags → status)
CAMPAIGN_MAPPINGS_SHEET = "Campaign Mappings"
CAMPAIGN_MAPPINGS_HEADERS = ("Store ID", "Store Name", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status")
# Column indices (1-based) inside Campaign Mappings sheet
_COL_CAMPAIGN_NAME = 5   # E
_COL_STATUS = 6          # F

try:
    import pandas as pd
except ImportError:
    pd = None


def _copy_sheet_from_book(src_wb, sheet_name, dest_wb, new_name=None):
    """Copy a sheet from src_wb to dest_wb (values only)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")
    if sheet_name not in src_wb.sheetnames:
        return
    src_ws = src_wb[sheet_name]
    name = (new_name or sheet_name)[:31]
    if name in dest_wb.sheetnames:
        base, n = name, 1
        while f"{base}_{n}"[:31] in dest_wb.sheetnames:
            n += 1
        name = f"{base}_{n}"[:31]
    dest_ws = dest_wb.create_sheet(name)
    for row in src_ws.iter_rows():
        for cell in row:
            dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)


def write_combined_report(
    financial_xlsx_path: Optional[Path] = None,
    marketing_xlsx_path: Optional[Path] = None,
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Optional[Path]:
    """
    Create one Excel workbook with all sheets from financial and marketing workbooks.
    Financial sheets first (with "Financial - " prefix on sheet names if needed to avoid clashes),
    then marketing sheets. Saves to output_dir/combined_analysis_{timestamp}.xlsx unless output_filename set.
    Returns path to the combined file or None.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl required for combined report")
        return None

    output_dir = Path(output_dir) if output_dir else Path("downloads")
    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_filename:
        output_filename = f"combined_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = output_dir / output_filename

    wb_out = openpyxl.Workbook()
    # Remove default sheet after we add first real sheet
    default_sheet = wb_out.active
    sheet_count = 0

    for label, xlsx_path in [("Financial", financial_xlsx_path), ("Marketing", marketing_xlsx_path)]:
        if not xlsx_path or not Path(xlsx_path).is_file():
            continue
        try:
            wb_src = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            logger.warning("Could not open %s: %s", xlsx_path, e)
            continue
        for name in wb_src.sheetnames:
            safe_name = name[:31]
            if safe_name in wb_out.sheetnames:
                safe_name = f"{label}-{name}"[:31]
            _copy_sheet_from_book(wb_src, name, wb_out, safe_name)
            sheet_count += 1
        wb_src.close()

    if sheet_count == 0:
        wb_out.close()
        logger.warning("CombinedReportAgent: No sheets to write")
        return None

    wb_out.remove(default_sheet)
    wb_out.save(out_path)
    logger.info("CombinedReportAgent: Wrote %s (%s sheets)", out_path.name, sheet_count)
    return out_path


def _add_sheet_from_df(wb, sheet_name: str, df, title: str = None):
    """Add a sheet to openpyxl workbook from a pandas DataFrame."""
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    name = (sheet_name or "Sheet")[:31]
    if name in wb.sheetnames:
        base, n = name, 1
        while f"{base}_{n}"[:31] in wb.sheetnames:
            n += 1
        name = f"{base}_{n}"[:31]
    ws = wb.create_sheet(name)
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        start_row = 3
    else:
        start_row = 1
    if df is not None and not (getattr(df, "empty", True)):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)


def write_combined_from_sheets(
    financial_sheets: Optional[List[Tuple[str, object]]] = None,
    marketing_sheets: Optional[List[Tuple[str, object]]] = None,
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Optional[Path]:
    """
    Build one workbook from list of (sheet_name, DataFrame) for financial and marketing.
    Saves to output_dir/combined_analysis_{timestamp}.xlsx. Returns path or None.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl required for combined report")
        return None
    financial_sheets = financial_sheets or []
    marketing_sheets = marketing_sheets or []
    if not financial_sheets and not marketing_sheets:
        logger.warning("CombinedReportAgent: No sheets to write")
        return None

    output_dir = Path(output_dir) if output_dir else Path("downloads")
    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_filename:
        output_filename = f"combined_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = output_dir / output_filename

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    sheet_count = 0
    for name, df in financial_sheets:
        if df is not None and not (getattr(df, "empty", True)):
            _add_sheet_from_df(wb, name, df, name)
            sheet_count += 1
    for name, df in marketing_sheets:
        if df is not None and not (getattr(df, "empty", True)):
            safe = name[:31]
            if safe in wb.sheetnames:
                safe = f"Marketing-{name}"[:31]
            _add_sheet_from_df(wb, safe, df, name)
            sheet_count += 1
    if sheet_count == 0:
        wb.close()
        return None
    wb.remove(default_sheet)
    wb.save(out_path)
    logger.info("CombinedReportAgent: Wrote %s (%s sheets)", out_path.name, sheet_count)
    return out_path


def read_campaign_mapping_statuses(combined_xlsx_path: Path) -> Dict[str, str]:
    """
    Read the Campaign Mappings sheet and return {campaign_name: status}.
    Used before rewriting the sheet so existing statuses are preserved across retries.
    Returns empty dict if the sheet doesn't exist yet.
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    path = Path(combined_xlsx_path)
    if not path.is_file():
        return {}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        sheet_name = CAMPAIGN_MAPPINGS_SHEET[:31]
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        statuses: Dict[str, str] = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # skip header
            if not row or len(row) < _COL_CAMPAIGN_NAME or not row[_COL_CAMPAIGN_NAME - 1]:
                continue
            campaign_name = str(row[_COL_CAMPAIGN_NAME - 1]).strip()
            status = str(row[_COL_STATUS - 1]).strip() if len(row) >= _COL_STATUS and row[_COL_STATUS - 1] else "Pending"
            if campaign_name:
                statuses[campaign_name] = status
        return statuses
    finally:
        wb.close()


def update_campaign_mapping_status(combined_xlsx_path: Path, campaign_name: str, status: str) -> None:
    """
    Write status to column E for the matching campaign_name row in Campaign Mappings sheet.
    Called live after each campaign attempt so reruns can skip completed ones.
    """
    try:
        import openpyxl
    except ImportError:
        return

    path = Path(combined_xlsx_path)
    if not path.is_file():
        logger.warning("CombinedReportAgent: file not found for status update: %s", path)
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        logger.warning("CombinedReportAgent: could not open %s for status update: %s", path, e)
        return

    try:
        sheet_name = CAMPAIGN_MAPPINGS_SHEET[:31]
        if sheet_name not in wb.sheetnames:
            logger.warning("CombinedReportAgent: Campaign Mappings sheet not in %s", path.name)
            return
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            name_cell = row[_COL_CAMPAIGN_NAME - 1] if len(row) >= _COL_CAMPAIGN_NAME else None
            if name_cell and str(name_cell.value or "").strip() == campaign_name:
                # Extend row to column E if needed
                status_cell = ws.cell(row=name_cell.row, column=_COL_STATUS)
                status_cell.value = status
                wb.save(path)
                logger.info("CombinedReportAgent: %s → %s", campaign_name, status)
                return
        logger.warning("CombinedReportAgent: campaign '%s' not found in mappings for status update", campaign_name)
    except Exception as e:
        logger.warning("CombinedReportAgent: status update failed for %s: %s", campaign_name, e)
    finally:
        wb.close()


def append_campaign_mappings_to_workbook(
    combined_xlsx_path: Path,
    mappings: List[Dict[str, Any]],
) -> None:
    """
    Append a "Campaign Mappings" sheet to an existing combined_analysis workbook.

    Each item in mappings should have:
        store_id: str
        min_subtotal: int | float
        slot_tags: list of int/str (UI slot identifiers)
        campaign_name: str (e.g. TODC-{StoreID}-${min_subtotal})

    Slot tags are written as a comma-separated string. If a sheet with the same
    name exists, it is replaced.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        logger.warning("openpyxl required to append campaign mappings")
        return

    path = Path(combined_xlsx_path)
    if not path.is_file():
        logger.warning("CombinedReportAgent: combined analysis file not found: %s", path)
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=False)
    except Exception as e:
        logger.warning("CombinedReportAgent: could not open workbook %s: %s", path, e)
        return

    # Preserve any statuses already written from a previous (partial) run
    existing_statuses = read_campaign_mapping_statuses(path)

    sheet_name = CAMPAIGN_MAPPINGS_SHEET[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    for col, header in enumerate(CAMPAIGN_MAPPINGS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, m in enumerate(mappings, start=2):
        store_id = str(m.get("store_id", "")).strip()
        store_name = str(m.get("store_name", "")).strip()
        min_subtotal = m.get("min_subtotal", 0)
        slot_tags = m.get("slot_tags") or []
        if isinstance(slot_tags, (list, tuple)):
            slot_tags_str = ",".join(str(t).strip() for t in slot_tags)
        else:
            slot_tags_str = str(slot_tags).strip()
        campaign_name = str(m.get("campaign_name", "")).strip()
        # Carry forward existing status so a rerun doesn't reset completed campaigns
        status = existing_statuses.get(campaign_name, "Pending")
        ws.cell(row=row_idx, column=1, value=store_id)
        ws.cell(row=row_idx, column=2, value=store_name)
        ws.cell(row=row_idx, column=3, value=min_subtotal)
        ws.cell(row=row_idx, column=4, value=slot_tags_str)
        ws.cell(row=row_idx, column=5, value=campaign_name)
        ws.cell(row=row_idx, column=_COL_STATUS, value=status)

    try:
        wb.save(path)
        logger.info(
            "CombinedReportAgent: Pushed %s campaign mapping(s) to sheet %s in %s",
            len(mappings),
            sheet_name,
            path.name,
        )
    except Exception as e:
        logger.warning("CombinedReportAgent: could not save workbook %s: %s", path, e)
    finally:
        wb.close()


def copy_campaign_mappings_from_previous(src_xlsx: Path, dest_xlsx: Path) -> bool:
    """
    Copy the Campaign Mappings sheet wholesale from a previous combined analysis
    into the new one. Returns True if the sheet was successfully copied.
    """
    try:
        import openpyxl
    except ImportError:
        return False

    src_path = Path(src_xlsx)
    dest_path = Path(dest_xlsx)
    if not src_path.is_file() or not dest_path.is_file():
        return False

    try:
        src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    except Exception:
        return False

    sheet_name = CAMPAIGN_MAPPINGS_SHEET[:31]
    if sheet_name not in src_wb.sheetnames:
        src_wb.close()
        return False

    try:
        dest_wb = openpyxl.load_workbook(dest_path, read_only=False)
        if sheet_name in dest_wb.sheetnames:
            del dest_wb[sheet_name]
        _copy_sheet_from_book(src_wb, sheet_name, dest_wb, sheet_name)
        dest_wb.save(dest_path)
        dest_wb.close()
        src_wb.close()
        logger.info("CombinedReportAgent: Copied Campaign Mappings from %s to %s", src_path.name, dest_path.name)
        return True
    except Exception as e:
        logger.warning("CombinedReportAgent: Failed to copy Campaign Mappings: %s", e)
        src_wb.close()
        return False


def read_campaign_combos_from_mappings(combined_xlsx_path: Path) -> List[Dict[str, Any]]:
    """
    Read campaign combos from the Campaign Mappings sheet.
    Returns list of dicts: {store_id, store_name, min_subtotal, slot_tags, campaign_name, status}.
    """
    try:
        import openpyxl
    except ImportError:
        return []

    path = Path(combined_xlsx_path)
    if not path.is_file():
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    try:
        sheet_name = CAMPAIGN_MAPPINGS_SHEET[:31]
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        combos: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # skip header
            if not row or len(row) < 5:
                continue
            store_id = str(row[0] or "").strip()
            store_name = str(row[1] or "").strip()
            min_subtotal = row[2] if row[2] is not None else 10
            slot_tags_str = str(row[3] or "").strip()
            campaign_name = str(row[4] or "").strip()
            status = str(row[5] or "Pending").strip() if len(row) >= 6 and row[5] else "Pending"

            # Parse slot_tags from comma-separated string
            slot_tags = []
            if slot_tags_str:
                for t in slot_tags_str.split(","):
                    t = t.strip()
                    if t:
                        try:
                            slot_tags.append(int(float(t)))
                        except (ValueError, TypeError):
                            pass

            if campaign_name:
                combos.append({
                    "store_id": store_id,
                    "store_name": store_name,
                    "min_subtotal": min_subtotal,
                    "slot_tags": slot_tags,
                    "campaign_name": campaign_name,
                    "status": status,
                })
        return combos
    finally:
        wb.close()


def run(
    financial_xlsx_path: Optional[Path] = None,
    marketing_xlsx_path: Optional[Path] = None,
    financial_sheets: Optional[List[Tuple[str, object]]] = None,
    marketing_sheets: Optional[List[Tuple[str, object]]] = None,
    output_dir: Optional[Path] = None,
) -> Optional[Path]:
    """Write combined workbook from either xlsx paths or in-memory sheet lists. Returns path or None."""
    if financial_sheets is not None or marketing_sheets is not None:
        return write_combined_from_sheets(
            financial_sheets=financial_sheets,
            marketing_sheets=marketing_sheets,
            output_dir=output_dir,
        )
    return write_combined_report(
        financial_xlsx_path=financial_xlsx_path,
        marketing_xlsx_path=marketing_xlsx_path,
        output_dir=output_dir,
    )
