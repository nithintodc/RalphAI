"""Export functions for Excel and date exports"""
import pandas as pd
import streamlit as st
from datetime import datetime
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from config import ROOT_DIR
from gdrive_utils import get_drive_manager
from utils import normalize_store_id_column, filter_master_file_by_date_range, UE_DATE_COLUMN_VARIATIONS, DD_DATE_COLUMN_VARIATIONS
from table_generation import create_summary_tables
from data_processing import get_last_year_dates



# Canonical financial line-item keywords (order matches the CSV header).
# DD exports use pipe separators (e.g. "Marketing fees | (including…)") but the
# exact format varies across exports. We match by keyword prefix so both variants work.
_DD_FINANCIAL_KEYWORDS = [
    "Subtotal",
    "Subtotal tax",
    "Commission",
    "Payment processing fee",
    "Marketing fees",
    "Customer discounts from marketing",
    "DoorDash marketing credit",
    "Third-party contribution",
    "Ad fee",
    "Error charges",
    "Adjustments",
    "Net total",
    "Pre-adjusted subtotal",
    "Pre-adjusted tax subtotal",
    "Subtotal for tax",
    "Subtotal tax remitted",
    "DoorDash funded subtotal discount",
    "Merchant funded subtotal discount",
]

# Non-financial columns to skip when discovering line-items
_DD_SKIP_COLS = {
    "Merchant store ID", "Store ID", "Merchant Store ID",
    "DoorDash order ID", "Order received local time",
    "Timestamp local date", "Timestamp local time", "Timestamp Local Date",
    "Transaction type", "Business name", "Business ID",
    "Date", "date", "Timestamp", "timestamp",
}


def _discover_financial_columns(df):
    """
    Return an ordered list of financial line-item columns present in *df*.
    Uses keyword matching so both pipe-separated and plain names are found,
    preserving their original order in the DataFrame.
    """
    found = []
    subtotal_seen = False
    for col in df.columns:
        col_stripped = col.strip()
        if col_stripped in _DD_SKIP_COLS:
            continue
        col_lower = col_stripped.lower()
        if col_lower == "subtotal":
            subtotal_seen = True
        if subtotal_seen:
            found.append(col_stripped)
    return found


def _build_dd_financial_breakdown(dd_data_path, start_date, end_date, excluded_dates,
                                   aggregate_only=False):
    """
    Build a pivot DataFrame from the DD financial file for one time window.

    Rows   = every financial column from Subtotal onwards (auto-discovered)
    Columns = Store IDs + Total  (each with Value and %)
        Value = SUM of that column for the store
        %     = Value / Subtotal for the same store  (0 when Subtotal is 0)

    If *aggregate_only* is True, only the Total columns are returned (no per-store).

    Returns a flat DataFrame ready for openpyxl writing, or None on failure.
    """
    if not dd_data_path or not Path(dd_data_path).exists():
        return None
    try:
        filtered = filter_master_file_by_date_range(
            Path(dd_data_path), start_date, end_date,
            DD_DATE_COLUMN_VARIATIONS, excluded_dates,
        )
        if filtered.empty:
            return None

        store_col_name = "Merchant store ID"
        if store_col_name not in filtered.columns:
            for c in ("Store ID", "Merchant Store ID"):
                if c in filtered.columns:
                    store_col_name = c
                    break

        if store_col_name not in filtered.columns:
            return None

        present_rows = _discover_financial_columns(filtered)
        if not present_rows:
            return None

        subtotal_col = present_rows[0] if present_rows else None

        for col in present_rows:
            filtered[col] = pd.to_numeric(filtered[col], errors="coerce").fillna(0)

        store_ids = sorted(filtered[store_col_name].dropna().unique(), key=lambda x: str(x))

        records = []
        for row_name in present_rows:
            row_dict = {"Metric": row_name}
            total_val = filtered[row_name].sum()
            subtotal_total = filtered[subtotal_col].sum() if subtotal_col else 0

            if not aggregate_only:
                for sid in store_ids:
                    mask = filtered[store_col_name] == sid
                    val = filtered.loc[mask, row_name].sum()
                    sub = filtered.loc[mask, subtotal_col].sum() if subtotal_col else 0
                    pct = (val / sub * 100) if sub != 0 else 0
                    row_dict[f"{sid}_Value"] = round(val, 2)
                    row_dict[f"{sid}_%"] = round(pct, 1)

            pct_total = (total_val / subtotal_total * 100) if subtotal_total != 0 else 0
            row_dict["Total_Value"] = round(total_val, 2)
            row_dict["Total_%"] = round(pct_total, 1)
            records.append(row_dict)

        return pd.DataFrame(records)
    except Exception:
        return None


def _compute_financial_diff(post_df, pre_df):
    """
    Compute Post − Pre for each financial line-item.
    Returns a DataFrame with columns: Metric, <store>_Value, <store>_%, Total_Value, Total_%
    where Value = post − pre and % = percentage-point change in the % column.
    Works for both per-store and aggregate-only DataFrames.
    """
    if post_df is None or pre_df is None or post_df.empty or pre_df.empty:
        return None
    merged = post_df.merge(pre_df, on="Metric", how="outer", suffixes=("_post", "_pre"))
    merged["Metric"] = merged["Metric"].fillna("")

    value_cols = [c for c in post_df.columns if c.endswith("_Value")]
    pct_cols = [c for c in post_df.columns if c.endswith("_%")]

    records = []
    for _, r in merged.iterrows():
        row = {"Metric": r["Metric"]}
        for vc in value_cols:
            post_v = r.get(f"{vc}_post", 0) or 0
            pre_v = r.get(f"{vc}_pre", 0) or 0
            row[vc] = round(post_v - pre_v, 2)
        for pc in pct_cols:
            post_p = r.get(f"{pc}_post", 0) or 0
            pre_p = r.get(f"{pc}_pre", 0) or 0
            row[pc] = round(post_p - pre_p, 1)
        records.append(row)

    col_order = ["Metric"] + [c for c in post_df.columns if c != "Metric"]
    result = pd.DataFrame(records)
    return result[[c for c in col_order if c in result.columns]]


def _write_financial_breakdown_table(ws, title, df, start_row):
    """
    Write a DD financial breakdown table with merged multi-level headers to a worksheet.
    Returns the next available row.
    """
    if df is None or df.empty:
        return start_row

    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1

    store_ids = []
    for c in df.columns:
        if c.endswith("_Value") and c != "Total_Value":
            store_ids.append(c.replace("_Value", ""))

    header_row1 = start_row
    header_row2 = start_row + 1
    data_start = start_row + 2

    ws.cell(row=header_row1, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=header_row2, column=1, value="").font = Font(bold=True)
    ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)

    col_idx = 2
    for sid in store_ids:
        ws.cell(row=header_row1, column=col_idx, value=str(sid)).font = Font(bold=True)
        ws.cell(row=header_row1, column=col_idx).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=header_row1, start_column=col_idx, end_row=header_row1, end_column=col_idx + 1)
        ws.cell(row=header_row2, column=col_idx, value="Value").font = Font(bold=True)
        ws.cell(row=header_row2, column=col_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=header_row2, column=col_idx + 1, value="%").font = Font(bold=True)
        ws.cell(row=header_row2, column=col_idx + 1).alignment = Alignment(horizontal="center")
        col_idx += 2

    ws.cell(row=header_row1, column=col_idx, value="Total").font = Font(bold=True)
    ws.cell(row=header_row1, column=col_idx).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=header_row1, start_column=col_idx, end_row=header_row1, end_column=col_idx + 1)
    ws.cell(row=header_row2, column=col_idx, value="Value").font = Font(bold=True)
    ws.cell(row=header_row2, column=col_idx).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row2, column=col_idx + 1, value="%").font = Font(bold=True)
    ws.cell(row=header_row2, column=col_idx + 1).alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(df.iterrows()):
        r = data_start + ri
        ws.cell(row=r, column=1, value=row["Metric"])
        c = 2
        for sid in store_ids:
            val = row.get(f"{sid}_Value", 0)
            pct = row.get(f"{sid}_%", 0)
            cell_v = ws.cell(row=r, column=c, value=val)
            cell_v.number_format = '#,##0.00'
            cell_v.alignment = Alignment(horizontal="right")
            cell_p = ws.cell(row=r, column=c + 1, value=pct)
            cell_p.number_format = '0.0'
            cell_p.alignment = Alignment(horizontal="right")
            c += 2
        val = row.get("Total_Value", 0)
        pct = row.get("Total_%", 0)
        cell_v = ws.cell(row=r, column=c, value=val)
        cell_v.number_format = '#,##0.00'
        cell_v.alignment = Alignment(horizontal="right")
        cell_p = ws.cell(row=r, column=c + 1, value=pct)
        cell_p.number_format = '0.0'
        cell_p.alignment = Alignment(horizontal="right")

    total_cols = 1 + len(store_ids) * 2 + 2
    for ci in range(1, total_cols + 1):
        max_len = 0
        for ri in range(header_row1, data_start + len(df)):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 50)

    return data_start + len(df) + 2


# ── Financial Summary Table (aggregate + per-store) ──

_FINANCIAL_METRICS = [
    'Sales', 'DD Sales', 'DD Commission', 'DD Ads Spend', 'DD Promo Spend',
    'DD Error Charges', 'DD Adjustments', 'DD Payouts', 'DD Profitability%',
    'UE Sales', 'UE Error Charges', 'UE Promo', 'UE Commissions',
    'UE Payouts', 'UE Profitability%',
]


def _find_col_ci(df, name):
    """Find column by exact match, case-insensitive exact, then prefix match."""
    if df is None or df.empty or not name:
        return None
    if name in df.columns:
        return name
    nl = name.lower().strip()
    for c in df.columns:
        if c.strip().lower() == nl:
            return c
    for c in df.columns:
        if c.strip().lower().startswith(nl):
            return c
    return None


def _safe_col_sum(df, col):
    """Sum a column safely, returning 0 if column is None or missing."""
    if df is None or df.empty or col is None or col not in df.columns:
        return 0
    return pd.to_numeric(df[col], errors='coerce').fillna(0).sum()


def _get_dd_store_col_name(df):
    """Find the DD store ID column name in a DataFrame."""
    for c in ("Merchant store ID", "Merchant Store ID", "Store ID"):
        if c in df.columns:
            return c
    return None


def _compute_window_metrics(dd_df, ue_df, store_id=None):
    """Compute all financial metrics for one time window, optionally per-store."""
    if dd_df is None:
        dd_df = pd.DataFrame()
    if ue_df is None:
        ue_df = pd.DataFrame()

    if store_id is not None:
        if not dd_df.empty:
            sc = _get_dd_store_col_name(dd_df)
            dd_df = dd_df[dd_df[sc].astype(str) == str(store_id)] if sc else pd.DataFrame()
        if not ue_df.empty:
            ue_sc = next((c for c in ('Store ID', 'Shop ID') if c in ue_df.columns), None)
            ue_df = ue_df[ue_df[ue_sc].astype(str) == str(store_id)] if ue_sc else pd.DataFrame()

    dd_sales = dd_comm = dd_ads = dd_promo = dd_errors = dd_adj = dd_payouts = 0
    if not dd_df.empty:
        dd_sales = _safe_col_sum(dd_df, _find_col_ci(dd_df, 'Subtotal'))
        dd_comm = abs(_safe_col_sum(dd_df, _find_col_ci(dd_df, 'Commission')))
        dd_ads = abs(_safe_col_sum(dd_df, _find_col_ci(dd_df, 'Marketing fees')))
        dd_promo = abs(_safe_col_sum(dd_df, _find_col_ci(dd_df, 'Customer discounts from marketing')))
        dd_errors = abs(_safe_col_sum(dd_df, _find_col_ci(dd_df, 'Error charges')))
        dd_adj = abs(_safe_col_sum(dd_df, _find_col_ci(dd_df, 'Adjustments')))
        dd_payouts = _safe_col_sum(dd_df, _find_col_ci(dd_df, 'Net total'))

    ue_sales_val = ue_err = ue_promo_val = ue_comm = ue_pay = 0
    if not ue_df.empty:
        ue_sales_val = _safe_col_sum(ue_df, _find_col_ci(ue_df, 'Sales (excl. tax)') or _find_col_ci(ue_df, 'Sales'))
        ue_err = _safe_col_sum(ue_df, _find_col_ci(ue_df, 'Order Error Adjustments') or _find_col_ci(ue_df, 'Order Error'))
        ue_promo_val = _safe_col_sum(ue_df, _find_col_ci(ue_df, 'Offers on items'))
        ue_comm = abs(_safe_col_sum(ue_df, _find_col_ci(ue_df, 'Marketplace Fee') or _find_col_ci(ue_df, 'Marketplace fee')))
        ue_pay = _safe_col_sum(ue_df, _find_col_ci(ue_df, 'Total payout') or _find_col_ci(ue_df, 'Total Payout'))

    dd_prof = (dd_payouts / dd_sales * 100) if dd_sales != 0 else 0
    ue_prof = (ue_pay / ue_sales_val * 100) if ue_sales_val != 0 else 0

    return {
        'Sales': dd_sales + ue_sales_val,
        'DD Sales': dd_sales, 'DD Commission': dd_comm,
        'DD Ads Spend': dd_ads, 'DD Promo Spend': dd_promo,
        'DD Error Charges': dd_errors, 'DD Adjustments': dd_adj,
        'DD Payouts': dd_payouts, 'DD Profitability%': dd_prof,
        'UE Sales': ue_sales_val, 'UE Error Charges': ue_err,
        'UE Promo': ue_promo_val, 'UE Commissions': ue_comm,
        'UE Payouts': ue_pay, 'UE Profitability%': ue_prof,
    }


def _build_summary_df(pre_m, post_m, ly_pre_m, ly_post_m):
    """Build the financial summary DataFrame from four window metric dicts.

    Column order: Pre, Post, Pre vs Post, Linear Growth%, Last Year Pre/Post,
    LY Pre vs Post, LY Linear %, YoY, YoY%.
    """
    rows = []
    for m in _FINANCIAL_METRICS:
        pre = pre_m.get(m, 0)
        post = post_m.get(m, 0)
        ly_pre = ly_pre_m.get(m, 0)
        ly_post = ly_post_m.get(m, 0)
        pvp = post - pre
        ly_pvp = ly_post - ly_pre
        yoy = post - ly_post
        linear = (pvp / pre * 100) if pre != 0 else 0
        ly_linear = (ly_pvp / ly_pre * 100) if ly_pre != 0 else 0
        yoy_g = (yoy / ly_post * 100) if ly_post != 0 else 0
        rows.append({
            'Metric': m,
            'Pre': round(pre, 2),
            'Post': round(post, 2),
            'Pre vs Post': round(pvp, 2),
            'Linear Growth%': round(linear, 1),
            'Last Year Pre': round(ly_pre, 2),
            'Last Year Post': round(ly_post, 2),
            'LY Pre vs Post': round(ly_pvp, 2),
            'LY Linear %': round(ly_linear, 1),
            'YoY': round(yoy, 2),
            'YoY%': round(yoy_g, 1),
        })
    return pd.DataFrame(rows)


def _load_all_windows(dd_data_path, ue_data_path, pre_start, pre_end,
                       post_start, post_end, excluded_dates):
    """Load filtered DataFrames for all 4 time windows (8 DataFrames total)."""
    pre24_s, pre24_e = get_last_year_dates(pre_start, pre_end)
    post24_s, post24_e = get_last_year_dates(post_start, post_end)

    def _ld(p, s, e, dc):
        if p and Path(p).exists():
            try:
                return filter_master_file_by_date_range(Path(p), s, e, dc, excluded_dates)
            except Exception:
                pass
        return pd.DataFrame()

    return {
        'dd_pre': _ld(dd_data_path, pre_start, pre_end, DD_DATE_COLUMN_VARIATIONS),
        'dd_post': _ld(dd_data_path, post_start, post_end, DD_DATE_COLUMN_VARIATIONS),
        'dd_ly_pre': _ld(dd_data_path, pre24_s, pre24_e, DD_DATE_COLUMN_VARIATIONS),
        'dd_ly_post': _ld(dd_data_path, post24_s, post24_e, DD_DATE_COLUMN_VARIATIONS),
        'ue_pre': _ld(ue_data_path, pre_start, pre_end, UE_DATE_COLUMN_VARIATIONS),
        'ue_post': _ld(ue_data_path, post_start, post_end, UE_DATE_COLUMN_VARIATIONS),
        'ue_ly_pre': _ld(ue_data_path, pre24_s, pre24_e, UE_DATE_COLUMN_VARIATIONS),
        'ue_ly_post': _ld(ue_data_path, post24_s, post24_e, UE_DATE_COLUMN_VARIATIONS),
    }


def build_financial_summary_table(dd_data_path, ue_data_path,
                                    pre_start, pre_end, post_start, post_end,
                                    excluded_dates=None, store_id=None):
    """
    Build the unified financial summary table (aggregate or per-store).

    Rows: Sales, DD Sales/Commission/Ads/Promo/Errors/Adj/Payouts/Profitability%,
          UE Sales/Errors/Promo/Commissions/Payouts/Profitability%
    Columns: Pre, Post, Pre vs Post, Linear Growth%, Last Year Pre, Last Year Post,
             LY Pre vs Post, LY Linear %, YoY, YoY%
    """
    if not all([pre_start, pre_end, post_start, post_end]):
        return pd.DataFrame()
    w = _load_all_windows(dd_data_path, ue_data_path, pre_start, pre_end,
                           post_start, post_end, excluded_dates)
    pre_m = _compute_window_metrics(w['dd_pre'], w['ue_pre'], store_id)
    post_m = _compute_window_metrics(w['dd_post'], w['ue_post'], store_id)
    ly_pre_m = _compute_window_metrics(w['dd_ly_pre'], w['ue_ly_pre'], store_id)
    ly_post_m = _compute_window_metrics(w['dd_ly_post'], w['ue_ly_post'], store_id)
    return _build_summary_df(pre_m, post_m, ly_pre_m, ly_post_m)


def build_all_store_financial_tables(dd_data_path, ue_data_path,
                                      pre_start, pre_end, post_start, post_end,
                                      excluded_dates=None):
    """Build per-store financial summary tables. Returns list of (store_id, DataFrame) tuples."""
    if not all([pre_start, pre_end, post_start, post_end]):
        return []
    w = _load_all_windows(dd_data_path, ue_data_path, pre_start, pre_end,
                           post_start, post_end, excluded_dates)
    all_ids = set()
    for key in ('dd_pre', 'dd_post', 'dd_ly_pre', 'dd_ly_post'):
        if not w[key].empty:
            sc = _get_dd_store_col_name(w[key])
            if sc:
                all_ids.update(w[key][sc].dropna().astype(str).unique())
    for key in ('ue_pre', 'ue_post', 'ue_ly_pre', 'ue_ly_post'):
        if not w[key].empty:
            for sc in ('Store ID', 'Shop ID'):
                if sc in w[key].columns:
                    all_ids.update(w[key][sc].dropna().astype(str).unique())
                    break
    results = []
    for sid in sorted(all_ids, key=str):
        pre_m = _compute_window_metrics(w['dd_pre'], w['ue_pre'], sid)
        post_m = _compute_window_metrics(w['dd_post'], w['ue_post'], sid)
        ly_pre_m = _compute_window_metrics(w['dd_ly_pre'], w['ue_ly_pre'], sid)
        ly_post_m = _compute_window_metrics(w['dd_ly_post'], w['ue_ly_post'], sid)
        results.append((sid, _build_summary_df(pre_m, post_m, ly_pre_m, ly_post_m)))
    return results


def _write_financial_summary_sheet(ws, title, df, start_row):
    """Write a financial summary table to an openpyxl worksheet. Returns next row."""
    if df is None or df.empty:
        return start_row
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    start_row += 1
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=ci, value=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
    start_row += 1
    pct_cols = {'Linear Growth%', 'LY Linear %', 'YoY%'}
    for _, row in df.iterrows():
        metric = str(row['Metric'])
        is_pct = 'Profitability%' in metric
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=start_row, column=ci)
            cell.alignment = Alignment(horizontal='center' if col != 'Metric' else 'left')
            if col == 'Metric':
                cell.value = val
            elif col in pct_cols or is_pct:
                cell.value = f"{val:.1f}%" if isinstance(val, (int, float)) else val
            else:
                cell.value = f"${val:,.2f}" if isinstance(val, (int, float)) else val
        start_row += 1
    for ci in range(1, len(df.columns) + 1):
        mx = 0
        for r in range(start_row - len(df) - 2, start_row):
            cv = ws.cell(row=r, column=ci).value
            if cv:
                mx = max(mx, len(str(cv)))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 50)
    return start_row + 1


def export_to_excel(dd_table1, dd_table2, ue_table1, ue_table2, 
                     dd_sales_df, dd_payouts_df, dd_orders_df, dd_new_customers_df,
                     ue_sales_df, ue_payouts_df, ue_orders_df, ue_new_customers_df,
                     dd_selected_stores, ue_selected_stores,
                     combined_summary1, combined_summary2, combined_store_table1, combined_store_table2,
                     corporate_todc_table=None, promotion_table=None, sponsored_table=None,
                     summary_metrics_table=None, store_ids_markups_table=None, operator_name=None,
                     sales_pre_post_table=None, sales_yoy_table=None, payouts_pre_post_table=None, payouts_yoy_table=None,
                     ue_sales_pre_post_table=None, ue_sales_yoy_table=None, ue_payouts_pre_post_table=None, ue_payouts_yoy_table=None,
                     dd_data_path=None, ue_data_path=None,
                     pre_start_date=None, pre_end_date=None,
                     post_start_date=None, post_end_date=None, excluded_dates=None,
                     financial_summary_table=None):
    """Export all tables to an Excel file with sheets: Summary Tables, Store-Level Tables, Corporate vs TODC, and Financial Summary"""
    # Use temp directory for file creation (will be downloaded, not saved to disk)
    import tempfile
    temp_dir = Path(tempfile.gettempdir())
    outputs_dir = temp_dir / "streamlit_exports"
    outputs_dir.mkdir(exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get summary tables
    dd_summary1, dd_summary2 = None, None
    ue_summary1, ue_summary2 = None, None
    if dd_sales_df is not None and dd_payouts_df is not None and dd_orders_df is not None:
        dd_summary1, dd_summary2 = create_summary_tables(dd_sales_df, dd_payouts_df, dd_orders_df, dd_new_customers_df, dd_selected_stores, is_ue=False)
    if ue_sales_df is not None and ue_payouts_df is not None and ue_orders_df is not None:
        ue_summary1, ue_summary2 = create_summary_tables(ue_sales_df, ue_payouts_df, ue_orders_df, ue_new_customers_df, ue_selected_stores, is_ue=True)
    
    # Sheet 1: Summary Tables
    ws_summary = wb.create_sheet("Summary Tables")
    current_row = 1
    
    def add_table_to_sheet(ws, table_name, df, start_row, start_col=1):
        """Add a table with name header to the sheet and format it. start_col is 1-based column for table placement."""
        if df is None or df.empty:
            return start_row
        # Add table name
        ws.cell(row=start_row, column=start_col, value=table_name)
        ws.cell(row=start_row, column=start_col).font = Font(bold=True, size=12)
        start_row += 1
        # Add table data
        # Only reset index when it has a known name we want as a column (don't export default RangeIndex for slot/markup tables)
        if df.index.name in ['Store ID', 'Metric', 'Campaign', 'Is Self Serve Campaign']:
            df_display = df.reset_index()
        else:
            df_display = df.copy()
        
        # Write header row
        for col_idx, col_name in enumerate(df_display.columns, start=1):
            cell = ws.cell(row=start_row, column=start_col + col_idx - 1, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1
        
        # Write data rows
        for row_idx, row_data in df_display.iterrows():
            # Check if this row is for Orders or New Customers (for summary tables)
            is_orders_row = False
            is_new_customers_row = False
            metric_value = ''  # Initialize metric_value
            
            # Try to get metric value from column first (after reset_index, Metric becomes a column)
            if 'Metric' in df_display.columns:
                try:
                    metric_val = row_data['Metric']
                    metric_value = str(metric_val) if pd.notna(metric_val) else ''
                except (KeyError, IndexError, TypeError):
                    metric_value = ''
                
                if metric_value == 'Orders':
                    is_orders_row = True
                elif metric_value == 'New Customers':
                    is_new_customers_row = True
            # If Metric is still in the index (wasn't reset), use row_idx directly
            elif df_display.index.name == 'Metric':
                metric_value = str(row_idx) if pd.notna(row_idx) else ''
            # Last resort: if row_idx itself is the metric name (string)
            elif isinstance(row_idx, str):
                metric_value = str(row_idx)
            
            for col_idx, col_name in enumerate(df_display.columns, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=start_row, column=start_col + col_idx - 1, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format based on column name and row type
                if 'Growth%' in col_name or 'YoY%' in col_name:
                    # Format as percentage with % symbol
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.1f}%"
                elif col_name in ['Store ID', 'Metric', 'Campaign', 'Is Self Serve Campaign', 'Merchant Store IDs', 'Markups']:
                    # Keep as is (text)
                    pass
                elif col_name == 'Orders':
                    # Orders: format as integer with comma separators
                    if isinstance(value, (int, float)):
                        cell.value = f"{int(round(value)):,}"
                elif col_name in ['Sales', 'Spend', 'Cost per Order']:
                    # Sales, Spend, Cost per Order: format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.2f}"
                elif col_name == 'ROAS':
                    # ROAS: format as decimal
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.2f}"
                elif is_orders_row or is_new_customers_row:
                    # Orders and New Customers rows: format as integer with comma separators (no decimals, no dollar sign)
                    if isinstance(value, (int, float)):
                        cell.value = f"{int(round(value)):,}"
                elif metric_value == 'Profitability':
                    # Profitability: format as percentage
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.1f}%"
                elif metric_value == 'Average Check':
                    # Average Check: format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.1f}"
                else:
                    # Format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.1f}"
            
            start_row += 1
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df_display.columns, start=1):
            c = start_col + col_idx - 1
            col_series = df_display[col_name].astype(str)
            max_len = int(col_series.str.len().max()) if len(col_series) else 0
            max_length = max(len(str(col_name)), max_len)
            ws.column_dimensions[get_column_letter(c)].width = min(max_length + 2, 50)
        
        return start_row + 1  # Add blank row after table
    
    # Add Summary Metrics table first
    if summary_metrics_table is not None and not summary_metrics_table.empty:
        current_row = add_table_to_sheet(ws_summary, "Summary Metrics", summary_metrics_table, current_row)
    
    # Add Merchant Store IDs / Markups table beside Summary Metrics with 5 columns gap (export only)
    if store_ids_markups_table is not None and not store_ids_markups_table.empty:
        # Summary Metrics has 2 columns; gap = 5 columns; so start at column 1+2+5 = 8
        add_table_to_sheet(ws_summary, "Merchant Store IDs / Markups", store_ids_markups_table, start_row=1, start_col=8)
    
    # Add Combined Table 1
    if combined_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "Combined Table 1: Current Year Pre vs Post Analysis", combined_summary1, current_row)
    
    # Add Combined Table 2 (YoY) - second position
    if combined_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "Combined Table 2: Year-over-Year Analysis", combined_summary2, current_row)
    
    # Add DD Table 1
    if dd_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "DoorDash Table 1: Current Year Pre vs Post Analysis", dd_summary1, current_row)
    
    # Add DD Table 2
    if dd_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "DoorDash Table 2: Year-over-Year Analysis", dd_summary2, current_row)
    
    # Add UE Table 1
    if ue_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "UberEats Table 1: Current Year Pre vs Post Analysis", ue_summary1, current_row)
    
    # Add UE Table 2
    if ue_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "UberEats Table 2: Year-over-Year Analysis", ue_summary2, current_row)
    
    # Sheet 2: Store-Level Tables
    ws_store = wb.create_sheet("Store-Level Tables")
    current_row = 1
    
    # Add Combined Store Table 1
    if combined_store_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "Combined Table 1: Current Year Pre vs Post Analysis (Store-Level)", combined_store_table1, current_row)
    
    # Add Combined Store Table 2 (YoY) - second position
    if combined_store_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "Combined Table 2: Year-over-Year Analysis (Store-Level)", combined_store_table2, current_row)
    
    # Add DD Store Table 1
    if dd_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "DoorDash Table 1: Current Year Pre vs Post Analysis (Store-Level)", dd_table1, current_row)
    
    # Add DD Store Table 2
    if dd_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "DoorDash Table 2: Year-over-Year Analysis (Store-Level)", dd_table2, current_row)
    
    # Add UE Store Table 1
    if ue_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "UberEats Table 1: Current Year Pre vs Post Analysis (Store-Level)", ue_table1, current_row)
    
    # Add UE Store Table 2
    if ue_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "UberEats Table 2: Year-over-Year Analysis (Store-Level)", ue_table2, current_row)
    
    # Sheet 3: Corporate vs TODC Tables
    if corporate_todc_table is not None and not corporate_todc_table.empty:
        ws_corporate = wb.create_sheet("Corporate vs TODC")
        current_row = 1
        
        # Add Combined Corporate vs TODC table
        # Prepare the table for export (reset index to include Campaign as column)
        corporate_export = corporate_todc_table.copy()
        corporate_export.index.name = 'Campaign'
        corporate_export = corporate_export.reset_index()
        corporate_export['Campaign'] = corporate_export['Campaign'].apply(
            lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
        )
        corporate_export = corporate_export.set_index('Campaign')
        
        current_row = add_table_to_sheet(ws_corporate, "Combined: Corporate vs TODC", corporate_export, current_row)
        
        # Add Promotion table if available
        if promotion_table is not None and not promotion_table.empty:
            promo_export = promotion_table.copy()
            promo_export.index.name = 'Campaign'
            promo_export = promo_export.reset_index()
            promo_export['Campaign'] = promo_export['Campaign'].apply(
                lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
            )
            promo_export = promo_export.set_index('Campaign')
            current_row = add_table_to_sheet(ws_corporate, "Promotion: Corporate vs TODC", promo_export, current_row)
        
        # Add Sponsored Listing table if available
        if sponsored_table is not None and not sponsored_table.empty:
            sponsored_export = sponsored_table.copy()
            sponsored_export.index.name = 'Campaign'
            sponsored_export = sponsored_export.reset_index()
            sponsored_export['Campaign'] = sponsored_export['Campaign'].apply(
                lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
            )
            sponsored_export = sponsored_export.set_index('Campaign')
            current_row = add_table_to_sheet(ws_corporate, "Sponsored Listing: Corporate vs TODC", sponsored_export, current_row)
    
    # Add DD slot-wise sheet
    if sales_pre_post_table is not None or sales_yoy_table is not None or payouts_pre_post_table is not None or payouts_yoy_table is not None:
        ws_slots = wb.create_sheet("DD-slotWise")
        current_row = 1
        
        # Add Table 1: Sales Pre/Post
        if sales_pre_post_table is not None and not sales_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 1: Sales - Pre vs Post", sales_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 2: Sales YoY
        if sales_yoy_table is not None and not sales_yoy_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 2: Sales - Year over Year", sales_yoy_table, current_row)
            current_row += 2
        
        # Add Table 3: Payouts Pre/Post
        if payouts_pre_post_table is not None and not payouts_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 3: Payouts - Pre vs Post", payouts_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 4: Payouts YoY
        if payouts_yoy_table is not None and not payouts_yoy_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 4: Payouts - Year over Year", payouts_yoy_table, current_row)
    
    # Add UE slot-wise sheet
    if ue_sales_pre_post_table is not None or ue_sales_yoy_table is not None or ue_payouts_pre_post_table is not None or ue_payouts_yoy_table is not None:
        ws_ue_slots = wb.create_sheet("UE-slotWise")
        current_row = 1
        
        # Add Table 1: Sales Pre/Post
        if ue_sales_pre_post_table is not None and not ue_sales_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_ue_slots, "Table 1: Sales - Pre vs Post", ue_sales_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 2: Sales YoY
        if ue_sales_yoy_table is not None and not ue_sales_yoy_table.empty:
            current_row = add_table_to_sheet(ws_ue_slots, "Table 2: Sales - Year over Year", ue_sales_yoy_table, current_row)
            current_row += 2
        
        # Add Table 3: Payouts Pre/Post
        if ue_payouts_pre_post_table is not None and not ue_payouts_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_ue_slots, "Table 3: Payouts - Pre vs Post", ue_payouts_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 4: Payouts YoY
        if ue_payouts_yoy_table is not None and not ue_payouts_yoy_table.empty:
            current_row = add_table_to_sheet(ws_ue_slots, "Table 4: Payouts - Year over Year", ue_payouts_yoy_table, current_row)
    
    # ── Financial Summary Sheets (Aggregate + Per-Store) ──
    if pre_start_date and pre_end_date and post_start_date and post_end_date:
        ws_agg = wb.create_sheet("DD Financial-Aggregate")
        agg_row = 1
        if financial_summary_table is not None and not financial_summary_table.empty:
            agg_row = _write_financial_summary_sheet(
                ws_agg, "Financial Summary (Aggregate)", financial_summary_table, agg_row)
        else:
            _agg_tbl = build_financial_summary_table(
                dd_data_path, ue_data_path, pre_start_date, pre_end_date,
                post_start_date, post_end_date, excluded_dates)
            agg_row = _write_financial_summary_sheet(
                ws_agg, "Financial Summary (Aggregate)", _agg_tbl, agg_row)

        store_fin_tables = build_all_store_financial_tables(
            dd_data_path, ue_data_path, pre_start_date, pre_end_date,
            post_start_date, post_end_date, excluded_dates)
        if store_fin_tables:
            ws_bd = wb.create_sheet("DD Financial Breakdown")
            bd_row = 1
            for sid, tbl in store_fin_tables:
                bd_row = _write_financial_summary_sheet(ws_bd, f"Store {sid}", tbl, bd_row)

    # ── Insights Sheet ──
    ws_insights = wb.create_sheet("Insights")
    ins_row = 1
    ws_insights.cell(row=ins_row, column=1, value="Key Insights").font = Font(bold=True, size=14)
    ins_row += 2

    # Helper: write a titled section of insights rows
    def _write_insight_section(ws, title, rows_data, start_row):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
        start_row += 1
        if not rows_data:
            ws.cell(row=start_row, column=1, value="No data available")
            return start_row + 2
        headers = list(rows_data[0].keys())
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=ci, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
        start_row += 1
        for rd in rows_data:
            for ci, h in enumerate(headers, 1):
                ws.cell(row=start_row, column=ci, value=rd[h])
            start_row += 1
        return start_row + 1

    # 1. Platform with major loss/gain
    platform_insights = []
    if combined_summary1 is not None and not combined_summary1.empty:
        for metric in combined_summary1.index:
            try:
                pre_val = float(combined_summary1.loc[metric, 'Pre']) if 'Pre' in combined_summary1.columns else 0
                post_val = float(combined_summary1.loc[metric, 'Post']) if 'Post' in combined_summary1.columns else 0
                change = post_val - pre_val
                pct = (change / pre_val * 100) if pre_val != 0 else 0
                direction = "Gain" if change > 0 else ("Loss" if change < 0 else "No Change")
                platform_insights.append({
                    'Metric': metric, 'Pre': round(pre_val, 1), 'Post': round(post_val, 1),
                    'Change': round(change, 1), 'Change %': f"{pct:.1f}%", 'Direction': direction
                })
            except (ValueError, TypeError):
                continue
    else:
        # Build from individual platform summaries
        for label, s1 in [("DoorDash", dd_summary1), ("UberEats", ue_summary1)]:
            if s1 is not None and not s1.empty:
                for metric in s1.index:
                    try:
                        pre_val = float(s1.loc[metric, 'Pre']) if 'Pre' in s1.columns else 0
                        post_val = float(s1.loc[metric, 'Post']) if 'Post' in s1.columns else 0
                        change = post_val - pre_val
                        pct = (change / pre_val * 100) if pre_val != 0 else 0
                        direction = "Gain" if change > 0 else ("Loss" if change < 0 else "No Change")
                        platform_insights.append({
                            'Platform': label, 'Metric': metric, 'Pre': round(pre_val, 1),
                            'Post': round(post_val, 1), 'Change': round(change, 1),
                            'Change %': f"{pct:.1f}%", 'Direction': direction
                        })
                    except (ValueError, TypeError):
                        continue
    # Sort by absolute change descending to highlight major movers
    platform_insights.sort(key=lambda x: abs(x.get('Change', 0)), reverse=True)
    ins_row = _write_insight_section(ws_insights, "Platform – Major Loss/Gain (Pre vs Post)", platform_insights, ins_row)

    # 2. Stores with major loss/gain
    store_insights = []
    for label, tbl in [("Combined", combined_store_table1), ("DoorDash", dd_table1), ("UberEats", ue_table1)]:
        if tbl is not None and not tbl.empty:
            df_t = tbl.reset_index() if tbl.index.name else tbl.copy()
            id_col = 'Store ID' if 'Store ID' in df_t.columns else (df_t.columns[0] if len(df_t.columns) > 0 else None)
            if id_col and 'Pre' in df_t.columns and 'Post' in df_t.columns:
                for _, row in df_t.iterrows():
                    try:
                        pre_v = float(row['Pre'])
                        post_v = float(row['Post'])
                        chg = post_v - pre_v
                        pct = (chg / pre_v * 100) if pre_v != 0 else 0
                        direction = "Gain" if chg > 0 else ("Loss" if chg < 0 else "No Change")
                        store_insights.append({
                            'Source': label, 'Store': row[id_col], 'Pre': round(pre_v, 1),
                            'Post': round(post_v, 1), 'Change': round(chg, 1),
                            'Change %': f"{pct:.1f}%", 'Direction': direction
                        })
                    except (ValueError, TypeError):
                        continue
            # Only use the first available source
            if store_insights:
                break
    store_insights.sort(key=lambda x: abs(x.get('Change', 0)), reverse=True)
    ins_row = _write_insight_section(ws_insights, "Stores – Major Loss/Gain (Sales Pre vs Post)", store_insights[:20], ins_row)

    # 3. Slots with major loss/gain
    slot_insights = []
    if sales_pre_post_table is not None and not sales_pre_post_table.empty:
        slot_col = 'Slot' if 'Slot' in sales_pre_post_table.columns else sales_pre_post_table.columns[0]
        for _, row in sales_pre_post_table.iterrows():
            try:
                pre_v = float(row['Pre'])
                post_v = float(row['Post'])
                chg = post_v - pre_v
                pct = (chg / pre_v * 100) if pre_v != 0 else 0
                direction = "Gain" if chg > 0 else ("Loss" if chg < 0 else "No Change")
                slot_insights.append({
                    'Slot': row[slot_col], 'Pre': round(pre_v, 1), 'Post': round(post_v, 1),
                    'Change': round(chg, 1), 'Change %': f"{pct:.1f}%", 'Direction': direction
                })
            except (ValueError, TypeError):
                continue
    slot_insights.sort(key=lambda x: abs(x.get('Change', 0)), reverse=True)
    ins_row = _write_insight_section(ws_insights, "Slots – Major Loss/Gain (Sales Pre vs Post)", slot_insights, ins_row)

    # 4. Dates in post period with major loss/gain (daily sales from store-level data)
    date_insights = []
    # Try to build daily data from store tables if available
    for label, store_df in [("DoorDash", dd_sales_df), ("UberEats", ue_sales_df)]:
        if store_df is not None and not store_df.empty:
            # Store tables have pre_25, post_25 columns; compute per-store average for context
            try:
                total_pre = store_df['pre_25'].sum() if 'pre_25' in store_df.columns else 0
                total_post = store_df['post_25'].sum() if 'post_25' in store_df.columns else 0
                chg = total_post - total_pre
                pct = (chg / total_pre * 100) if total_pre != 0 else 0
                direction = "Gain" if chg > 0 else ("Loss" if chg < 0 else "No Change")
                date_insights.append({
                    'Platform': label, 'Pre Period Sales': round(total_pre, 1),
                    'Post Period Sales': round(total_post, 1), 'Change': round(chg, 1),
                    'Change %': f"{pct:.1f}%", 'Direction': direction
                })
            except (ValueError, TypeError):
                continue
    ins_row = _write_insight_section(ws_insights, "Post Period – Major Loss/Gain by Platform", date_insights, ins_row)

    # Auto-fit column widths for Insights sheet
    for col_cells in ws_insights.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_insights.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Generate filename with timestamp (use operator name if provided)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) and operator_name.strip() else None)
    filename = f"{tag}_analysis_export_{timestamp}.xlsx" if tag else f"analysis_export_{timestamp}.xlsx"
    filepath = outputs_dir / filename
    
    # Save workbook
    wb.save(filepath)
    
    # Read file as bytes for download
    with open(filepath, 'rb') as f:
        file_bytes = f.read()
    
    # Upload to Google Drive and create Google Doc with tables
    try:
        drive_manager = get_drive_manager()
        if drive_manager:
            # Upload Excel to "cloud-app-uploads" folder
            upload_result = drive_manager.upload_file_to_subfolder(
                file_path=filepath,
                root_folder_name="cloud-app-uploads",
                subfolder_name="outputs",
                file_name=filename
            )
            st.success(f"**Export successful!** Excel file ready for download and uploaded to Google Drive.")
            link = upload_result.get('webViewLink') or f"https://drive.google.com/file/d/{upload_result.get('file_id', '')}/view"
            st.info(f"File uploaded to Google Drive: [{upload_result['file_name']}]({link})")
    except Exception as e:
        st.warning(f"⚠️ Google Drive upload failed: {str(e)}")
    
    # Return file bytes and filename for download
    return file_bytes, filename


def create_date_export(dd_pre_24_path, dd_post_24_path, dd_pre_25_path, dd_post_25_path,
                      ue_pre_24_path, ue_post_24_path, ue_pre_25_path, ue_post_25_path,
                      dd_selected_stores, ue_selected_stores):
    """Create date pivot tables with Store IDs as columns and Sales, Payouts, Orders as values
    Processes only dd-pre/post and ue-pre/post files (8 files total)
    Returns a dictionary with file names as keys and dictionaries of Sales, Payouts, Orders as values
    Each file gets 3 separate pivot tables (Sales, Payouts, Orders)
    """
    
    def process_dd_file_for_date_export(file_path, selected_stores):
        """Process DD file and return data pivoted by date"""
        try:
            df = pd.read_csv(file_path)
            df.columns = df.columns.str.strip()
            
            # Use "Timestamp local date" for DD
            date_col = 'Timestamp local date'
            store_col = 'Merchant store ID'
            sales_col = 'Subtotal'
            
            # Determine payout column
            if '24' in file_path.name:
                payout_col = 'Net total (for historical reference only)'
            else:
                payout_col = 'Net total'
            
            order_col = 'DoorDash order ID'
            
            if date_col not in df.columns or store_col not in df.columns:
                st.warning(f"Missing required columns in {file_path.name}. Looking for '{date_col}' and '{store_col}'. Found: {list(df.columns)[:10]}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Check for payout column - try both names if needed
            if payout_col not in df.columns:
                # Try the other column name as fallback
                if payout_col == 'Net total':
                    payout_col = 'Net total (for historical reference only)'
                else:
                    payout_col = 'Net total'
                
                if payout_col not in df.columns:
                    st.warning(f"Payout column not found in {file_path.name}. Tried 'Net total' and 'Net total (for historical reference only)'. Available columns: {list(df.columns)[:10]}")
                    return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Process ALL data - no filtering by selected stores for date export
            # Convert date
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            df = df.dropna(subset=[date_col, store_col])
            
            if len(df) == 0:
                st.warning(f"No valid data after date conversion in {file_path.name}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert to numeric
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce')
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce')
            
            # Aggregate by Date and Store ID
            if len(df) == 0:
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            sales_pivot = df.groupby([date_col, store_col])[sales_col].sum().reset_index()
            payouts_pivot = df.groupby([date_col, store_col])[payout_col].sum().reset_index()
            orders_pivot = df.groupby([date_col, store_col])[order_col].nunique().reset_index()
            
            # Check if we have data to pivot
            if len(sales_pivot) == 0:
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert Store ID to string for consistent column names
            sales_pivot[store_col] = sales_pivot[store_col].astype(str)
            payouts_pivot[store_col] = payouts_pivot[store_col].astype(str)
            orders_pivot[store_col] = orders_pivot[store_col].astype(str)
            
            # Pivot: Date as index, Store ID as columns
            sales_pivot_table = sales_pivot.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            payouts_pivot_table = payouts_pivot.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            orders_pivot_table = orders_pivot.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            
            return sales_pivot_table, payouts_pivot_table, orders_pivot_table
        except Exception as e:
            st.error(f"Error processing {file_path.name} for date export: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    def process_ue_file_for_date_export(file_path, selected_stores):
        """Process UE file and return data pivoted by date"""
        try:
            df = pd.read_csv(file_path, skiprows=[0], header=0)
            df.columns = df.columns.str.strip()
            
            # Normalize store ID column (check for both 'Store ID' and 'Shop ID')
            df, store_col = normalize_store_id_column(df)
            
            # For UE files: hardcode to 9th column (index 8) as Order Date
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                st.warning(f"UE file {file_path.name} has fewer than 9 columns. Available columns: {list(df.columns)}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            sales_col = 'Sales (excl. tax)'
            payout_col = 'Total payout'
            order_col = 'Order ID'
            
            if date_col not in df.columns or store_col is None or store_col not in df.columns:
                st.warning(f"Missing required columns in {file_path.name}. Looking for '{date_col}' and 'Store ID' or 'Shop ID'. Found: {list(df.columns)[:10]}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Process ALL data - no filtering by selected stores for date export
            # Convert date - Store original values before parsing
            original_dates = df[date_col].copy()
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
            df = df.dropna(subset=[date_col, store_col])
            
            if len(df) == 0:
                st.warning(f"No valid data after date conversion in {file_path.name}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert to numeric
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce')
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce')
            
            # Aggregate by Date and Store ID
            sales_pivot = df.groupby([date_col, store_col])[sales_col].sum().reset_index()
            payouts_pivot = df.groupby([date_col, store_col])[payout_col].sum().reset_index()
            orders_pivot = df.groupby([date_col, store_col])[order_col].nunique().reset_index()
            
            # Convert Store ID to string for consistent column names
            sales_pivot[store_col] = sales_pivot[store_col].astype(str)
            payouts_pivot[store_col] = payouts_pivot[store_col].astype(str)
            orders_pivot[store_col] = orders_pivot[store_col].astype(str)
            
            # Pivot: Date as index, Store ID as columns
            sales_pivot_table = sales_pivot.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            payouts_pivot_table = payouts_pivot.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            orders_pivot_table = orders_pivot.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            
            return sales_pivot_table, payouts_pivot_table, orders_pivot_table
        except Exception as e:
            st.error(f"Error processing {file_path.name} for date export: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Process all 8 files separately - each file gets its own entry
    result = {}
    
    # Process DD files
    dd_files = [
        (dd_pre_24_path, 'DD_PRE_24', dd_selected_stores),
        (dd_post_24_path, 'DD_POST_24', dd_selected_stores),
        (dd_pre_25_path, 'DD_PRE_25', dd_selected_stores),
        (dd_post_25_path, 'DD_POST_25', dd_selected_stores),
    ]
    
    for file_path, file_key, selected_stores in dd_files:
        if not file_path.exists():
            st.warning(f"File not found: {file_path}")
            continue
        try:
            sales, payouts, orders = process_dd_file_for_date_export(file_path, selected_stores)
            # Always add the file entry if at least one metric has data
            if not sales.empty or not payouts.empty or not orders.empty:
                result[file_key] = {
                    'Sales': sales if not sales.empty else pd.DataFrame(),
                    'Payouts': payouts if not payouts.empty else pd.DataFrame(),
                    'Orders': orders if not orders.empty else pd.DataFrame()
                }
            else:
                st.warning(f"No data extracted from {file_path.name} - all pivot tables are empty")
        except Exception as e:
            st.error(f"Error processing {file_path.name}: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
    
    # Process UE files
    ue_files = [
        (ue_pre_24_path, 'UE_PRE_24', ue_selected_stores),
        (ue_post_24_path, 'UE_POST_24', ue_selected_stores),
        (ue_pre_25_path, 'UE_PRE_25', ue_selected_stores),
        (ue_post_25_path, 'UE_POST_25', ue_selected_stores),
    ]
    
    for file_path, file_key, selected_stores in ue_files:
        if not file_path.exists():
            st.warning(f"File not found: {file_path}")
            continue
        try:
            sales, payouts, orders = process_ue_file_for_date_export(file_path, selected_stores)
            # Always add the file entry if at least one metric has data
            if not sales.empty or not payouts.empty or not orders.empty:
                result[file_key] = {
                    'Sales': sales if not sales.empty else pd.DataFrame(),
                    'Payouts': payouts if not payouts.empty else pd.DataFrame(),
                    'Orders': orders if not orders.empty else pd.DataFrame()
                }
            else:
                st.warning(f"No data extracted from {file_path.name} - all pivot tables are empty")
        except Exception as e:
            st.error(f"Error processing {file_path.name}: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
    
    return result if result else None


def create_date_export_from_master_files(dd_data_path, ue_data_path, pre_start_date, pre_end_date, post_start_date, post_end_date, excluded_dates=None, operator_name=None):
    """
    Create date-wise exports of DD and UE financial data.
    Creates a single Excel file with 6 sheets (DD + UE x Sales/Payouts/Orders).
    Each sheet contains side-by-side blocks for 2025 and 2024 (Pre/Post).
    
    Args:
        dd_data_path: Path to DoorDash master file
        ue_data_path: Path to UberEats master file
        pre_start_date: Pre period start date (MM/DD/YYYY)
        pre_end_date: Pre period end date (MM/DD/YYYY)
        post_start_date: Post period start date (MM/DD/YYYY)
        post_end_date: Post period end date (MM/DD/YYYY)
        excluded_dates: List of dates to exclude
    
    Returns:
        Tuple of (excel_bytes, filename) for download
    """
    try:
        # Calculate last year dates
        pre_24_start, pre_24_end = get_last_year_dates(pre_start_date, pre_end_date)
        post_24_start, post_24_end = get_last_year_dates(post_start_date, post_end_date)
        
        # Create a single Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        GAP_COLUMNS = 1
        
        # Process in order: DD_25, UE_25, DD_24, UE_24 (each: Sales, Payouts, Orders)
        if dd_data_path and Path(dd_data_path).exists():
            dd_pre_25 = filter_master_file_by_date_range(Path(dd_data_path), pre_start_date, pre_end_date, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_post_25 = filter_master_file_by_date_range(Path(dd_data_path), post_start_date, post_end_date, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_pre_24 = filter_master_file_by_date_range(Path(dd_data_path), pre_24_start, pre_24_end, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_post_24 = filter_master_file_by_date_range(Path(dd_data_path), post_24_start, post_24_end, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
        else:
            dd_pre_25 = dd_post_25 = dd_pre_24 = dd_post_24 = pd.DataFrame()
        if ue_data_path and Path(ue_data_path).exists():
            ue_pre_25 = filter_master_file_by_date_range(Path(ue_data_path), pre_start_date, pre_end_date, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_post_25 = filter_master_file_by_date_range(Path(ue_data_path), post_start_date, post_end_date, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_pre_24 = filter_master_file_by_date_range(Path(ue_data_path), pre_24_start, pre_24_end, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_post_24 = filter_master_file_by_date_range(Path(ue_data_path), post_24_start, post_24_end, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
        else:
            ue_pre_25 = ue_post_25 = ue_pre_24 = ue_post_24 = pd.DataFrame()
        
        # Build only DD_25_* and UE_25_* sheets. Place 2024 blocks on those same sheets.
        def add_dd_sheets(base_sheet_label, pre25_df, post25_df, pre24_df, post24_df):
            pre25_sales, pre25_payouts, pre25_orders = _build_period_pivots(pre25_df, 'DD', 'Merchant store ID', 'Subtotal', 'Net total', 'DoorDash order ID')
            post25_sales, post25_payouts, post25_orders = _build_period_pivots(post25_df, 'DD', 'Merchant store ID', 'Subtotal', 'Net total', 'DoorDash order ID')
            pre24_sales, pre24_payouts, pre24_orders = _build_period_pivots(pre24_df, 'DD', 'Merchant store ID', 'Subtotal', 'Net total (for historical reference only)', 'DoorDash order ID')
            post24_sales, post24_payouts, post24_orders = _build_period_pivots(post24_df, 'DD', 'Merchant store ID', 'Subtotal', 'Net total (for historical reference only)', 'DoorDash order ID')

            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Sales", pre25_sales, post25_sales, pre24_sales, post24_sales, GAP_COLUMNS)
            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Payouts", pre25_payouts, post25_payouts, pre24_payouts, post24_payouts, GAP_COLUMNS)
            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Orders", pre25_orders, post25_orders, pre24_orders, post24_orders, GAP_COLUMNS)

        def add_ue_sheets(base_sheet_label, pre25_df, post25_df, pre24_df, post24_df):
            ref_df = post25_df if pre25_df.empty else pre25_df
            if ref_df.empty:
                return
            ref_norm, store_col = normalize_store_id_column(ref_df.copy())
            pre25_df_norm = normalize_store_id_column(pre25_df.copy())[0] if not pre25_df.empty else pre25_df
            post25_df_norm = normalize_store_id_column(post25_df.copy())[0] if not post25_df.empty else post25_df
            pre24_df_norm = normalize_store_id_column(pre24_df.copy())[0] if not pre24_df.empty else pre24_df
            post24_df_norm = normalize_store_id_column(post24_df.copy())[0] if not post24_df.empty else post24_df

            pre25_sales, pre25_payouts, pre25_orders = _build_period_pivots(pre25_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')
            post25_sales, post25_payouts, post25_orders = _build_period_pivots(post25_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')
            pre24_sales, pre24_payouts, pre24_orders = _build_period_pivots(pre24_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')
            post24_sales, post24_payouts, post24_orders = _build_period_pivots(post24_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')

            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Sales", pre25_sales, post25_sales, pre24_sales, post24_sales, GAP_COLUMNS)
            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Payouts", pre25_payouts, post25_payouts, pre24_payouts, post24_payouts, GAP_COLUMNS)
            _add_two_year_pre_post_sheet(wb, f"{base_sheet_label}_Orders", pre25_orders, post25_orders, pre24_orders, post24_orders, GAP_COLUMNS)
        
        add_dd_sheets('DD_25', dd_pre_25, dd_post_25, dd_pre_24, dd_post_24)
        add_ue_sheets('UE_25', ue_pre_25, ue_post_25, ue_pre_24, ue_post_24)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) and operator_name.strip() else None)
        filename = f"{tag}_date_export_{timestamp}.xlsx" if tag else f"date_export_{timestamp}.xlsx"
        
        return excel_buffer.read(), filename
    
    except Exception as e:
        st.error(f"Error creating date export: {str(e)}")
        import traceback
        with st.expander("Error details"):
            st.code(traceback.format_exc())
        return None, None


def _build_period_pivots(df, platform, store_col, sales_col, payout_col, order_col):
    """
    Build Sales, Payouts, and Orders pivot DataFrames for a single period (Date + store columns).
    Returns (sales_pivot_df, payouts_pivot_df, orders_pivot_df); each may be empty.
    """
    from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
    empty = pd.DataFrame()
    if df is None or df.empty or store_col is None or store_col not in df.columns:
        return empty.copy(), empty.copy(), empty.copy()
    
    date_col = None
    if platform == 'DD':
        date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
    else:
        if len(df.columns) > 8:
            date_col = df.columns[8]
    if date_col is None:
        return empty.copy(), empty.copy(), empty.copy()
    
    df = df.copy()
    original_dates = df[date_col].copy()
    if platform == 'UE':
        df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
        if df[date_col].isna().any():
            mask_na = df[date_col].isna()
            df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
    else:
        df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
        if df[date_col].isna().all():
            df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
        if df[date_col].isna().all():
            df[date_col] = pd.to_datetime(original_dates, errors='coerce')
    df = df.dropna(subset=[date_col, store_col])
    if df.empty:
        return empty.copy(), empty.copy(), empty.copy()
    
    if sales_col in df.columns:
        df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
    # DD payout column: prefer requested one; fallback to alternate if missing.
    # Also try case-insensitive / partial match (some DD exports have slight column name variations).
    effective_payout_col = payout_col
    if platform == 'DD':
        if payout_col not in df.columns:
            alt = 'Net total' if payout_col == 'Net total (for historical reference only)' else 'Net total (for historical reference only)'
            if alt in df.columns:
                effective_payout_col = alt
        if effective_payout_col not in df.columns:
            # Try partial match for historical column
            for c in df.columns:
                c_lower = str(c).lower().strip()
                if 'historical reference only' in c_lower and 'net total' in c_lower:
                    effective_payout_col = c
                    break
                if c_lower == 'net total' and payout_col == 'Net total':
                    effective_payout_col = c
                    break
    if effective_payout_col in df.columns:
        df[effective_payout_col] = pd.to_numeric(df[effective_payout_col], errors='coerce').fillna(0)
    
    sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
    payouts_agg = df.groupby([date_col, store_col])[effective_payout_col].sum().reset_index() if effective_payout_col in df.columns else pd.DataFrame()
    orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
    
    def _pivot(agg, date_col, store_col, value_col):
        if agg.empty or value_col not in agg.columns:
            return empty.copy()
        p = agg.pivot_table(index=date_col, columns=store_col, values=value_col, aggfunc='sum', fill_value=0)
        p.index = p.index.strftime('%Y-%m-%d')
        p.index.name = 'Date'
        return p.reset_index()
    
    sales_pivot = _pivot(sales_agg, date_col, store_col, sales_col)
    payouts_pivot = _pivot(payouts_agg, date_col, store_col, effective_payout_col)
    orders_pivot = _pivot(orders_agg, date_col, store_col, order_col)
    return sales_pivot, payouts_pivot, orders_pivot


def _add_totals_to_pivot(pivot_df):
    """
    Add a Total column (sum of each row), an Avg-Value column (daily average minus
    current day's total), and a Total row (sum of each column).
    First column is treated as label (Date); rest are numeric.
    Returns a new DataFrame.
    """
    if pivot_df is None or pivot_df.empty:
        return pivot_df
    df = pivot_df.copy()
    label_col = df.columns[0]
    numeric_cols = [c for c in df.columns if c != label_col]
    if not numeric_cols:
        return df
    df['Total'] = df[numeric_cols].sum(axis=1)
    num_days = len(df)
    daily_avg = round(df['Total'].sum() / num_days, 1) if num_days > 0 else 0
    df['Avg-Value'] = (daily_avg - df['Total']).round(1)
    total_row = {label_col: 'Total'}
    for c in numeric_cols:
        total_row[c] = df[c].sum()
    total_row['Total'] = df['Total'].sum()
    total_row['Avg-Value'] = ''
    order = [label_col] + numeric_cols + ['Total', 'Avg-Value']
    df = df[order]
    total_row_ordered = {c: total_row[c] for c in order}
    df = pd.concat([df, pd.DataFrame([total_row_ordered])], ignore_index=True)
    return df


def _add_pre_post_sheet(wb, sheet_name, pre_pivot, post_pivot, gap_cols=4):
    """
    Create one sheet with Pre data (left), gap_cols empty columns, then Post data (right).
    Adds Total column and Total row to each block.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows
    pre_pivot = _add_totals_to_pivot(pre_pivot) if pre_pivot is not None and not pre_pivot.empty else pre_pivot
    post_pivot = _add_totals_to_pivot(post_pivot) if post_pivot is not None and not post_pivot.empty else post_pivot
    ws = wb.create_sheet(sheet_name)
    start_col_post = 1
    if pre_pivot is not None and not pre_pivot.empty:
        pre_cols = pre_pivot.shape[1]
        pre_rows = 1 + len(pre_pivot)
        for row_idx, row in enumerate(dataframe_to_rows(pre_pivot, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == pre_rows:
                    cell.font = Font(bold=True)
                if row_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        start_col_post = pre_cols + 1 + gap_cols
    if post_pivot is not None and not post_pivot.empty:
        post_rows = 1 + len(post_pivot)
        for row_idx, row in enumerate(dataframe_to_rows(post_pivot, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=start_col_post):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == post_rows:
                    cell.font = Font(bold=True)
                if row_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')


def _add_two_year_pre_post_sheet(wb, sheet_name, pre25_pivot, post25_pivot, pre24_pivot, post24_pivot, gap_cols=4):
    """
    Create one sheet with four side-by-side blocks:
    Pre 25 | Post 25 | Pre 24 | Post 24
    Each block includes totals and a small header label.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows

    blocks = [
        ("Pre 25", _add_totals_to_pivot(pre25_pivot) if pre25_pivot is not None and not pre25_pivot.empty else None),
        ("Post 25", _add_totals_to_pivot(post25_pivot) if post25_pivot is not None and not post25_pivot.empty else None),
        ("Pre 24", _add_totals_to_pivot(pre24_pivot) if pre24_pivot is not None and not pre24_pivot.empty else None),
        ("Post 24", _add_totals_to_pivot(post24_pivot) if post24_pivot is not None and not post24_pivot.empty else None),
    ]

    ws = wb.create_sheet(sheet_name)
    start_col = 1
    start_row = 2  # Row 1 reserved for block titles

    for block_title, pivot_df in blocks:
        if pivot_df is None or pivot_df.empty:
            continue

        ws.cell(row=1, column=start_col, value=block_title).font = Font(bold=True, size=12)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='left', vertical='center')

        block_rows = 1 + len(pivot_df)
        block_cols = pivot_df.shape[1]

        for row_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), start=start_row):
            for col_offset, value in enumerate(row):
                col_idx = start_col + col_offset
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == start_row or row_idx == (start_row + block_rows - 1):
                    cell.font = Font(bold=True)
                if row_idx == start_row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        start_col += block_cols + gap_cols

def _add_period_sheets_to_workbook(wb, df, platform, period_name, store_col, sales_col, payout_col, order_col):
    """
    Add Sales, Payouts, and Orders sheets for a specific period to an existing workbook.
    
    Args:
        wb: openpyxl Workbook object
        df: DataFrame with data
        platform: 'DD' or 'UE'
        period_name: Period name like 'DD_Pre_25', 'UE_Post_24', etc.
        store_col: Name of store ID column
        sales_col: Name of sales column
        payout_col: Name of payout column
        order_col: Name of order ID column
    """
    try:
        # Find date column
        from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
        date_col = None
        if platform == 'DD':
            date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
        else:  # UE - hardcode to 9th column (index 8)
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                return
        
        if date_col is None or store_col is None or store_col not in df.columns:
            return
        
        # Convert date column - Store original values before parsing
        original_dates = df[date_col].copy()
        if platform == 'UE':
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
        else:
            # DD files: Try MM/DD/YYYY format first (most common), then YYYY-MM-DD
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            if df[date_col].isna().all():
                # If all failed, try YYYY-MM-DD format using original values
                df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
            # Fall back to auto parsing if format doesn't match
            if df[date_col].isna().all():
                df[date_col] = pd.to_datetime(original_dates, errors='coerce')
        df = df.dropna(subset=[date_col, store_col])
        
        if df.empty:
            return
        
        # Convert to numeric
        if sales_col in df.columns:
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
        if payout_col in df.columns:
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce').fillna(0)
        
        # Aggregate by date and store
        sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
        payouts_agg = df.groupby([date_col, store_col])[payout_col].sum().reset_index() if payout_col in df.columns else pd.DataFrame()
        orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Sheet 1: Sales
        if not sales_agg.empty:
            sales_pivot = sales_agg.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            sales_pivot.index = sales_pivot.index.strftime('%Y-%m-%d')
            sales_pivot.index.name = 'Date'
            sales_pivot = sales_pivot.reset_index()
            
            ws_sales = wb.create_sheet(f"{period_name}_Sales")
            for r in dataframe_to_rows(sales_pivot, index=False, header=True):
                ws_sales.append(r)
            # Format header row
            for cell in ws_sales[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 2: Payouts
        if not payouts_agg.empty:
            payouts_pivot = payouts_agg.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            payouts_pivot.index = payouts_pivot.index.strftime('%Y-%m-%d')
            payouts_pivot.index.name = 'Date'
            payouts_pivot = payouts_pivot.reset_index()
            
            ws_payouts = wb.create_sheet(f"{period_name}_Payouts")
            for r in dataframe_to_rows(payouts_pivot, index=False, header=True):
                ws_payouts.append(r)
            # Format header row
            for cell in ws_payouts[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 3: Orders
        if not orders_agg.empty:
            orders_pivot = orders_agg.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            orders_pivot.index = orders_pivot.index.strftime('%Y-%m-%d')
            orders_pivot.index.name = 'Date'
            orders_pivot = orders_pivot.reset_index()
            
            ws_orders = wb.create_sheet(f"{period_name}_Orders")
            for r in dataframe_to_rows(orders_pivot, index=False, header=True):
                ws_orders.append(r)
            # Format header row
            for cell in ws_orders[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    except Exception as e:
        st.warning(f"Error adding sheets for {period_name}: {str(e)}")
        import traceback
        st.error(traceback.format_exc())


def _create_period_excel_file(df, platform, period_name, store_col, sales_col, payout_col, order_col):
    """
    Create a single Excel file for a specific period with 3 sheets: Sales, Payouts, Orders.
    
    Args:
        df: DataFrame with data
        platform: 'DD' or 'UE'
        period_name: Period name like 'DD_Pre_25', 'UE_Post_24', etc.
        store_col: Name of store ID column
        sales_col: Name of sales column
        payout_col: Name of payout column
        order_col: Name of order ID column
    
    Returns:
        Bytes of Excel file
    """
    try:
        # Find date column
        from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
        date_col = None
        if platform == 'DD':
            date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
        else:  # UE - hardcode to 9th column (index 8)
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                return None
        
        if date_col is None or store_col is None or store_col not in df.columns:
            return None
        
        # Convert date column - Store original values before parsing
        original_dates = df[date_col].copy()
        if platform == 'UE':
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
        else:
            # DD files: Try MM/DD/YYYY format first (most common), then YYYY-MM-DD
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            if df[date_col].isna().all():
                # If all failed, try YYYY-MM-DD format using original values
                df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
            # Fall back to auto parsing if format doesn't match
            if df[date_col].isna().all():
                df[date_col] = pd.to_datetime(original_dates, errors='coerce')
        df = df.dropna(subset=[date_col, store_col])
        
        if df.empty:
            return None
        
        # Convert to numeric
        if sales_col in df.columns:
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
        if payout_col in df.columns:
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce').fillna(0)
        
        # Aggregate by date and store
        sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
        payouts_agg = df.groupby([date_col, store_col])[payout_col].sum().reset_index() if payout_col in df.columns else pd.DataFrame()
        orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
        
        # Create Excel workbook with 3 sheets
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Sheet 1: Sales
        if not sales_agg.empty:
            sales_pivot = sales_agg.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            sales_pivot.index = sales_pivot.index.strftime('%Y-%m-%d')
            sales_pivot.index.name = 'Date'
            sales_pivot = sales_pivot.reset_index()
            
            ws_sales = wb.create_sheet("Sales")
            for r in dataframe_to_rows(sales_pivot, index=False, header=True):
                ws_sales.append(r)
            # Format header row
            for cell in ws_sales[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 2: Payouts
        if not payouts_agg.empty:
            payouts_pivot = payouts_agg.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            payouts_pivot.index = payouts_pivot.index.strftime('%Y-%m-%d')
            payouts_pivot.index.name = 'Date'
            payouts_pivot = payouts_pivot.reset_index()
            
            ws_payouts = wb.create_sheet("Payouts")
            for r in dataframe_to_rows(payouts_pivot, index=False, header=True):
                ws_payouts.append(r)
            # Format header row
            for cell in ws_payouts[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 3: Orders
        if not orders_agg.empty:
            orders_pivot = orders_agg.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            orders_pivot.index = orders_pivot.index.strftime('%Y-%m-%d')
            orders_pivot.index.name = 'Date'
            orders_pivot = orders_pivot.reset_index()
            
            ws_orders = wb.create_sheet("Orders")
            for r in dataframe_to_rows(orders_pivot, index=False, header=True):
                ws_orders.append(r)
            # Format header row
            for cell in ws_orders[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.read()
    
    except Exception as e:
        st.warning(f"Error creating Excel file for {period_name}: {str(e)}")
        return None


_BUCKET_METRIC_COLS = [
    "Sales", "Payouts", "Mkt Spend", "Customer Discounts", "Orders",
    "GC $0-15", "GC $15-20", "GC $20-25", "GC $25-30", "GC $30-$35",
    "GC $35-$40", "GC $40+", "Count of Orders Mktg Driven",
]


def _bucket_agg(df, group_keys):
    """Aggregate bucketing metrics by *group_keys*, recompute Profitability_% and AOV."""
    import numpy as np
    agg_map = {c: "sum" for c in _BUCKET_METRIC_COLS if c in df.columns}
    if group_keys:
        out = df.groupby(group_keys, dropna=False).agg(agg_map).reset_index()
    else:
        out = pd.DataFrame([{c: df[c].sum() for c in agg_map}])
    for c in ["Sales", "Payouts", "Mkt Spend", "Customer Discounts"]:
        if c in out.columns:
            out[c] = out[c].round(1)
    out["Profitability_%"] = np.where(
        out["Sales"].abs() > 1e-9,
        np.round((out["Payouts"] / out["Sales"]) * 100.0, 1), np.nan)
    out["AOV"] = np.where(
        out["Orders"] > 0, np.round(out["Sales"] / out["Orders"], 1), np.nan)
    return out


def _bucket_diff_table(post_df, pre_df, group_keys):
    """
    Compute post − pre for every metric and add a % column for each.
    Returns a single DataFrame with <metric>_pre, <metric>_post, <metric>_diff, <metric>_% columns.
    """
    import numpy as np
    m_cols = [c for c in _BUCKET_METRIC_COLS if c in post_df.columns and c in pre_df.columns]
    post_a = _bucket_agg(post_df, group_keys)
    pre_a = _bucket_agg(pre_df, group_keys)
    merged = post_a.merge(pre_a, on=group_keys, how="outer", suffixes=("_post", "_pre"))
    rows = []
    for _, r in merged.iterrows():
        row = {k: r[k] for k in group_keys}
        for m in m_cols:
            post_v = r.get(f"{m}_post", 0) or 0
            pre_v = r.get(f"{m}_pre", 0) or 0
            diff = post_v - pre_v
            pct = (diff / pre_v * 100) if pre_v != 0 else 0
            row[f"{m}_Pre"] = round(pre_v, 2)
            row[f"{m}_Post"] = round(post_v, 2)
            row[f"{m}_Diff"] = round(diff, 2)
            row[f"{m}_%"] = round(pct, 1)
        # Recompute derived metrics for post and pre
        for tag in ("_Pre", "_Post"):
            s = row.get(f"Sales{tag}", 0)
            p = row.get(f"Payouts{tag}", 0)
            o = row.get(f"Orders{tag}", 0)
            row[f"Profitability_%{tag}"] = round((p / s * 100), 1) if s else 0
            row[f"AOV{tag}"] = round(s / o, 1) if o else 0
        s_post = row.get("Sales_Post", 0)
        p_post = row.get("Payouts_Post", 0)
        o_post = row.get("Orders_Post", 0)
        s_pre = row.get("Sales_Pre", 0)
        p_pre = row.get("Payouts_Pre", 0)
        o_pre = row.get("Orders_Pre", 0)
        prof_post = round((p_post / s_post * 100), 1) if s_post else 0
        prof_pre = round((p_pre / s_pre * 100), 1) if s_pre else 0
        aov_post = round((s_post / o_post), 1) if o_post else 0
        aov_pre = round((s_pre / o_pre), 1) if o_pre else 0
        row["Profitability_%_Diff"] = round(prof_post - prof_pre, 1)
        row["Profitability_%_%"] = ""
        row["AOV_Diff"] = round(aov_post - aov_pre, 1)
        row["AOV_%"] = round(((aov_post - aov_pre) / aov_pre * 100) if aov_pre else 0, 1)
        rows.append(row)
    out_cols = list(group_keys)
    for m in m_cols + ["Profitability_%", "AOV"]:
        out_cols += [f"{m}_Pre", f"{m}_Post", f"{m}_Diff", f"{m}_%"]
    return pd.DataFrame(rows, columns=[c for c in out_cols if c in pd.DataFrame(rows).columns])


def _daypart_gc_order_table(agg_tbl):
    """
    One row per day part + Grand Total; columns: Day part, GC buckets, Orders.
    Aligned to bucketing_analysis.DAY_PARTS order.
    """
    from bucketing_analysis import DAY_PARTS

    gc_cols = [
        "GC $0-15", "GC $15-20", "GC $20-25", "GC $25-30",
        "GC $30-$35", "GC $35-$40", "GC $40+",
    ]
    if agg_tbl is None or agg_tbl.empty or "Day part" not in agg_tbl.columns:
        return pd.DataFrame(columns=["Day part"] + gc_cols + ["Orders"])

    present_gc = [c for c in gc_cols if c in agg_tbl.columns]
    if not present_gc or "Orders" not in agg_tbl.columns:
        return pd.DataFrame(columns=["Day part"] + gc_cols + ["Orders"])

    g = agg_tbl.groupby("Day part", dropna=False)[present_gc + ["Orders"]].sum().reset_index()
    g[present_gc] = g[present_gc].fillna(0).astype(int)
    g["Orders"] = g["Orders"].fillna(0).astype(int)

    for c in gc_cols:
        if c not in g.columns:
            g[c] = 0

    rows = []
    for dp in DAY_PARTS:
        r = g[g["Day part"] == dp]
        if r.empty:
            rows.append({"Day part": dp, **{c: 0 for c in gc_cols}, "Orders": 0})
        else:
            row = r.iloc[0].to_dict()
            rows.append({**{"Day part": dp}, **{c: int(row.get(c, 0) or 0) for c in gc_cols}, "Orders": int(row.get("Orders", 0) or 0)})

    unk = g[g["Day part"] == "Unknown"]
    if not unk.empty:
        row = unk.iloc[0].to_dict()
        rows.append({**{"Day part": "Unknown"}, **{c: int(row.get(c, 0) or 0) for c in gc_cols}, "Orders": int(row.get("Orders", 0) or 0)})

    out = pd.DataFrame(rows)
    total = {"Day part": "Grand Total"}
    for c in gc_cols + ["Orders"]:
        total[c] = int(out[c].sum())
    out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
    return out


def _daypart_delta_tables(post_tbl, pre_tbl):
    """Return (delta, delta_pct) with same shape as post_tbl; delta = post - pre, pct = delta/pre."""
    import numpy as np

    gc_cols = [
        "GC $0-15", "GC $15-20", "GC $20-25", "GC $25-30",
        "GC $30-$35", "GC $35-$40", "GC $40+",
    ]
    if post_tbl.empty and pre_tbl.empty:
        empty = pd.DataFrame(columns=["Day part"] + gc_cols + ["Orders"])
        return empty.copy(), empty.copy()

    merged = post_tbl.merge(
        pre_tbl, on="Day part", how="outer", suffixes=("_post", "_pre")
    )
    merged = merged.fillna(0)
    delta = merged[["Day part"]].copy()
    pct = merged[["Day part"]].copy()
    for c in gc_cols + ["Orders"]:
        pc = f"{c}_post"
        pr = f"{c}_pre"
        if pc in merged.columns and pr in merged.columns:
            pv = merged[pc].astype(float)
            rv = merged[pr].astype(float)
            d = pv - rv
            delta[c] = d.round(0).astype(int)
            pct[c] = np.where(rv != 0, (d / rv) * 100.0, 0.0)
        elif pc in merged.columns:
            delta[c] = merged[pc].astype(float).round(0).astype(int)
            pct[c] = 0.0
        elif pr in merged.columns:
            delta[c] = (-merged[pr]).astype(float).round(0).astype(int)
            pct[c] = np.where(merged[pr] != 0, 100.0, 0.0)
        else:
            delta[c] = 0
            pct[c] = 0.0

    # Preserve row order from post_tbl where possible
    order = list(post_tbl["Day part"]) if not post_tbl.empty else list(delta["Day part"])
    extra = [x for x in delta["Day part"].tolist() if x not in order]
    order = order + extra
    seen = set()
    order = [x for x in order if not (x in seen or seen.add(x))]
    delta = delta.set_index("Day part").reindex(order).reset_index()
    pct = pct.set_index("Day part").reindex(order).reset_index()
    return delta, pct


_STORE_DAYPART_COL = "Merchant Store ID"


def _store_daypart_gc_order_table(agg_tbl, store_ids=None):
    """
    Rows: each store, sub-rows = day parts (slots); columns: Store ID, Day part, GC buckets, Orders.
    Grand Total row at bottom.

    store_ids: optional fixed store order/list (union of Post/Pre) so Post and Pre sheets align row-for-row.
    """
    from bucketing_analysis import DAY_PARTS

    gc_cols = [
        "GC $0-15", "GC $15-20", "GC $20-25", "GC $25-30",
        "GC $30-$35", "GC $35-$40", "GC $40+",
    ]
    cols_out = [_STORE_DAYPART_COL, "Day part"] + gc_cols + ["Orders"]
    if agg_tbl is None or agg_tbl.empty or _STORE_DAYPART_COL not in agg_tbl.columns:
        return pd.DataFrame(columns=cols_out)
    if "Day part" not in agg_tbl.columns:
        return pd.DataFrame(columns=cols_out)

    present_gc = [c for c in gc_cols if c in agg_tbl.columns]
    if not present_gc or "Orders" not in agg_tbl.columns:
        return pd.DataFrame(columns=cols_out)

    g = agg_tbl.groupby([_STORE_DAYPART_COL, "Day part"], dropna=False)[present_gc + ["Orders"]].sum().reset_index()
    g[present_gc] = g[present_gc].fillna(0).astype(int)
    g["Orders"] = g["Orders"].fillna(0).astype(int)
    for c in gc_cols:
        if c not in g.columns:
            g[c] = 0

    if store_ids is not None:
        stores = list(store_ids)
    else:
        stores = sorted(g[_STORE_DAYPART_COL].dropna().unique(), key=lambda x: str(x))
    rows = []
    for sid in stores:
        sid_key = str(sid).strip()
        for dp in DAY_PARTS:
            r = g[(g[_STORE_DAYPART_COL].astype(str) == sid_key) & (g["Day part"] == dp)]
            if r.empty:
                rows.append(
                    {_STORE_DAYPART_COL: sid, "Day part": dp, **{c: 0 for c in gc_cols}, "Orders": 0}
                )
            else:
                row = r.iloc[0].to_dict()
                rows.append({
                    _STORE_DAYPART_COL: sid,
                    "Day part": dp,
                    **{c: int(row.get(c, 0) or 0) for c in gc_cols},
                    "Orders": int(row.get("Orders", 0) or 0),
                })
        unk = g[(g[_STORE_DAYPART_COL].astype(str) == sid_key) & (g["Day part"] == "Unknown")]
        if not unk.empty:
            row = unk.iloc[0].to_dict()
            rows.append({
                _STORE_DAYPART_COL: sid,
                "Day part": "Unknown",
                **{c: int(row.get(c, 0) or 0) for c in gc_cols},
                "Orders": int(row.get("Orders", 0) or 0),
            })

    out = pd.DataFrame(rows)
    total = {_STORE_DAYPART_COL: "Grand Total", "Day part": ""}
    for c in gc_cols + ["Orders"]:
        total[c] = int(g[c].sum())
    out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
    return out


def _store_daypart_delta_tables(post_tbl, pre_tbl):
    """Delta and Delta % for store × day part; merge keys: Merchant Store ID, Day part."""
    import numpy as np

    gc_cols = [
        "GC $0-15", "GC $15-20", "GC $20-25", "GC $25-30",
        "GC $30-$35", "GC $35-$40", "GC $40+",
    ]
    key_cols = [_STORE_DAYPART_COL, "Day part"]
    empty = pd.DataFrame(columns=key_cols + gc_cols + ["Orders"])
    if post_tbl.empty and pre_tbl.empty:
        return empty.copy(), empty.copy()

    pt = post_tbl.copy()
    pr = pre_tbl.copy()
    pt["Day part"] = pt["Day part"].fillna("")
    pr["Day part"] = pr["Day part"].fillna("")

    merged = pt.merge(pr, on=key_cols, how="outer", suffixes=("_post", "_pre"))
    merged = merged.fillna(0)

    delta = merged[key_cols].copy()
    pct = merged[key_cols].copy()
    for c in gc_cols + ["Orders"]:
        pc = f"{c}_post"
        prc = f"{c}_pre"
        if pc in merged.columns and prc in merged.columns:
            pv = merged[pc].astype(float)
            rv = merged[prc].astype(float)
            d = pv - rv
            delta[c] = d.round(0).astype(int)
            pct[c] = np.where(rv != 0, (d / rv) * 100.0, 0.0)
        elif pc in merged.columns:
            delta[c] = merged[pc].astype(float).round(0).astype(int)
            pct[c] = 0.0
        elif prc in merged.columns:
            delta[c] = (-merged[prc]).astype(float).round(0).astype(int)
            pct[c] = np.where(merged[prc] != 0, 100.0, 0.0)
        else:
            delta[c] = 0
            pct[c] = 0.0

    from bucketing_analysis import DAY_PARTS

    dp_order = {dp: i for i, dp in enumerate(list(DAY_PARTS) + ["Unknown", ""])}
    for _df in (delta, pct):
        _df["_gt"] = (_df[_STORE_DAYPART_COL].astype(str) == "Grand Total").astype(int)
        _df["_sd"] = _df[_STORE_DAYPART_COL].astype(str)
        _df["_dp"] = _df["Day part"].astype(str).map(lambda x: dp_order.get(x, 999))
    delta = delta.sort_values(["_gt", "_sd", "_dp"]).drop(columns=["_gt", "_sd", "_dp"])
    pct = pct.sort_values(["_gt", "_sd", "_dp"]).drop(columns=["_gt", "_sd", "_dp"])

    return delta.reset_index(drop=True), pct.reset_index(drop=True)


def _write_daypart_gc_2x2_sheet(ws, post_df, pre_df, delta_df, delta_pct_df, pct_label_cols=1):
    """Write Post | Pre on top, Delta | Delta % below (2x2 grid with gap columns).

    pct_label_cols: leading columns in Delta % table shown as raw text (not %), e.g. 1 for Day part only,
    2 for Merchant Store ID + Day part.
    """

    def _fmt_pct(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
        try:
            return f"{float(val):.1f}%"
        except (TypeError, ValueError):
            return str(val)

    gap_cols = 2
    ncols = max(len(post_df.columns), len(pre_df.columns), 1)
    pre_start_col = 1 + ncols + gap_cols

    r = 1
    ws.cell(row=r, column=1, value="Post").font = Font(bold=True, size=12)
    ws.cell(row=r, column=pre_start_col, value="Pre").font = Font(bold=True, size=12)
    r += 1

    hdr_row = r
    for j, col in enumerate(post_df.columns, start=1):
        ws.cell(row=hdr_row, column=j, value=col).font = Font(bold=True)
    for j, col in enumerate(pre_df.columns, start=pre_start_col):
        ws.cell(row=hdr_row, column=j, value=col).font = Font(bold=True)
    r += 1

    max_top = max(len(post_df), len(pre_df))
    for i in range(max_top):
        ri = r + i
        if i < len(post_df):
            for j, col in enumerate(post_df.columns):
                ws.cell(row=ri, column=1 + j, value=post_df.iloc[i, j])
        if i < len(pre_df):
            for j, col in enumerate(pre_df.columns):
                ws.cell(row=ri, column=pre_start_col + j, value=pre_df.iloc[i, j])

    last_top_row = r + max_top - 1
    r = last_top_row + 2
    ws.cell(row=r, column=1, value="Delta").font = Font(bold=True, size=12)
    ws.cell(row=r, column=pre_start_col, value="Delta %").font = Font(bold=True, size=12)
    r += 1

    hdr2 = r
    for j, col in enumerate(delta_df.columns, start=1):
        ws.cell(row=hdr2, column=j, value=col).font = Font(bold=True)
    for j, col in enumerate(delta_pct_df.columns, start=pre_start_col):
        ws.cell(row=hdr2, column=j, value=col).font = Font(bold=True)
    r += 1

    max_bot = max(len(delta_df), len(delta_pct_df))
    for i in range(max_bot):
        ri = r + i
        if i < len(delta_df):
            for j in range(len(delta_df.columns)):
                ws.cell(row=ri, column=1 + j, value=delta_df.iloc[i, j])
        if i < len(delta_pct_df):
            for j in range(len(delta_pct_df.columns)):
                val = delta_pct_df.iloc[i, j]
                if j < pct_label_cols:
                    ws.cell(row=ri, column=pre_start_col + j, value=val)
                else:
                    try:
                        ws.cell(row=ri, column=pre_start_col + j, value=_fmt_pct(float(val)))
                    except (TypeError, ValueError):
                        ws.cell(row=ri, column=pre_start_col + j, value=val)

    for col in range(1, pre_start_col + ncols + 5):
        letter = get_column_letter(col)
        if pct_label_cols >= 2 and (col <= 2 or (pre_start_col <= col < pre_start_col + 2)):
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 14


def create_bucketing_export(dd_data_path, operator_name=None,
                            pre_start_date=None, pre_end_date=None,
                            post_start_date=None, post_end_date=None,
                            excluded_dates=None):
    """
    Run the AITF bucketing analysis on the DD financial file and produce a
    multi-sheet Excel workbook.

    Includes **Daypart GC**: 2×2 Post | Pre, Delta | Delta % (rows = day parts, columns = GC
    buckets + Orders). **Store Daypart GC**: same 2×2 layout with rows = store ID and
    sub-rows = day-part slots, same columns.

    Returns:
        Tuple of (excel_bytes, filename) or (None, None) on failure.
    """
    from bucketing_analysis import load_and_prepare, aggregate_slot_table
    import numpy as np

    csv_path = Path(dd_data_path)
    if not csv_path.exists():
        st.error(f"DD financial file not found: {csv_path}")
        return None, None

    try:
        has_dates = all([pre_start_date, pre_end_date, post_start_date, post_end_date])

        def _ts(d):
            if d is None:
                return None
            return pd.Timestamp(d)

        def _prepare_window(start, end):
            """Load + prepare for one window; returns (prepared_df, store_operator) or (empty, {})."""
            try:
                p, so = load_and_prepare(csv_path, start_date=_ts(start), end_date=_ts(end))
                return p, so
            except ValueError:
                return pd.DataFrame(), {}

        if has_dates:
            pre24_s, pre24_e = get_last_year_dates(pre_start_date, pre_end_date)
            post24_s, post24_e = get_last_year_dates(post_start_date, post_end_date)
            pre25, so_pre25 = _prepare_window(pre_start_date, pre_end_date)
            post25, so_post25 = _prepare_window(post_start_date, post_end_date)
            pre24, so_pre24 = _prepare_window(pre24_s, pre24_e)
            post24, so_post24 = _prepare_window(post24_s, post24_e)
            store_operator = {**so_pre24, **so_post24, **so_pre25, **so_post25}
        else:
            full, store_operator = load_and_prepare(csv_path, start_date=None, end_date=None)
            pre25 = post25 = pre24 = post24 = full

        windows = {"Pre": pre25, "Post": post25, "LY Pre": pre24, "LY Post": post24}
        agg_tables = {}
        for label, prep in windows.items():
            if prep is not None and not prep.empty:
                agg_tables[label] = aggregate_slot_table(prep, store_operator)

        STORE_COL = "Merchant Store ID"

        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:

            # Sheet 1: Aggregated by store × period
            period_rows = []
            for label, tbl in agg_tables.items():
                agg = _bucket_agg(tbl, [STORE_COL])
                agg.insert(0, "Period", label)
                period_rows.append(agg)
            if period_rows:
                pd.concat(period_rows, ignore_index=True).to_excel(
                    writer, sheet_name="By Period", index=False)

            # Sheet 2: Aggregated by store × slot × period
            slot_rows = []
            for label, tbl in agg_tables.items():
                agg = _bucket_agg(tbl, [STORE_COL, "Day part"])
                agg.insert(0, "Period", label)
                slot_rows.append(agg)
            if slot_rows:
                pd.concat(slot_rows, ignore_index=True).to_excel(
                    writer, sheet_name="By Slot-Period", index=False)

            # Sheet 3: Aggregated by store × day
            all_data = [t for t in agg_tables.values()]
            if all_data:
                combined = pd.concat(all_data, ignore_index=True)
                _bucket_agg(combined, [STORE_COL, "Day"]).to_excel(
                    writer, sheet_name="By Day", index=False)

            # Sheet 4: Post − Pre on store × slot (values + %)
            if "Post" in agg_tables and "Pre" in agg_tables:
                _bucket_diff_table(
                    agg_tables["Post"], agg_tables["Pre"], [STORE_COL, "Day part"]
                ).to_excel(writer, sheet_name="Slot Pre vs Post", index=False)

            # Sheet 5: Post − LY Post on store × slot (YoY values + %)
            if "Post" in agg_tables and "LY Post" in agg_tables:
                _bucket_diff_table(
                    agg_tables["Post"], agg_tables["LY Post"], [STORE_COL, "Day part"]
                ).to_excel(writer, sheet_name="Slot YoY", index=False)

            # Sheet 6: Post − Pre on store × day × slot (values + %)
            if "Post" in agg_tables and "Pre" in agg_tables:
                _bucket_diff_table(
                    agg_tables["Post"], agg_tables["Pre"], [STORE_COL, "Day", "Day part"]
                ).to_excel(writer, sheet_name="Day-Slot Pre vs Post", index=False)

            # Sheet 7: Post − LY Post on store × day × slot (YoY values + %)
            if "Post" in agg_tables and "LY Post" in agg_tables:
                _bucket_diff_table(
                    agg_tables["Post"], agg_tables["LY Post"], [STORE_COL, "Day", "Day part"]
                ).to_excel(writer, sheet_name="Day-Slot YoY", index=False)

            # Sheet: 2×2 Day part × GC buckets (Post | Pre, Delta | Delta %)
            if "Post" in agg_tables and "Pre" in agg_tables:
                post_tp = _daypart_gc_order_table(agg_tables["Post"])
                pre_tp = _daypart_gc_order_table(agg_tables["Pre"])
                if not post_tp.empty and not pre_tp.empty:
                    delta_tp, pct_tp = _daypart_delta_tables(post_tp, pre_tp)
                    ws_gc = writer.book.create_sheet("Daypart GC")
                    _write_daypart_gc_2x2_sheet(ws_gc, post_tp, pre_tp, delta_tp, pct_tp, pct_label_cols=1)

                _post_a = agg_tables["Post"]
                _pre_a = agg_tables["Pre"]
                if _STORE_DAYPART_COL in _post_a.columns and _STORE_DAYPART_COL in _pre_a.columns:
                    _stores_union = sorted(
                        set(_post_a[_STORE_DAYPART_COL].dropna().astype(str).unique())
                        | set(_pre_a[_STORE_DAYPART_COL].dropna().astype(str).unique()),
                        key=str,
                    )
                else:
                    _stores_union = None
                post_st = _store_daypart_gc_order_table(_post_a, store_ids=_stores_union)
                pre_st = _store_daypart_gc_order_table(_pre_a, store_ids=_stores_union)
                if not post_st.empty and not pre_st.empty:
                    delta_st, pct_st = _store_daypart_delta_tables(post_st, pre_st)
                    ws_st = writer.book.create_sheet("Store Daypart GC")
                    _write_daypart_gc_2x2_sheet(
                        ws_st, post_st, pre_st, delta_st, pct_st, pct_label_cols=2
                    )

        excel_buffer.seek(0)
        excel_bytes = excel_buffer.read()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag = (operator_name.strip()
               if operator_name and isinstance(operator_name, str) and operator_name.strip()
               else None)
        filename = (f"{tag}_bucketing_export_{timestamp}.xlsx" if tag
                    else f"bucketing_export_{timestamp}.xlsx")

        return excel_bytes, filename

    except Exception as e:
        st.error(f"Bucketing export failed: {e}")
        import traceback
        with st.expander("View Error Details"):
            st.code(traceback.format_exc())
        return None, None
