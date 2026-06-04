"""
MarketingAgent: runs after browser downloads the marketing report. Extracts the download
(ZIP or folder), runs marketing analysis (create_corporate_vs_todc_table and related
tables), and writes the final marketing analysis Excel to the downloads folder.
"""

import logging
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    import pandas as pd
except ImportError:
    pd = None


def _is_zip(path: Path) -> bool:
    if not path.is_file() or path.stat().st_size < 4:
        return False
    with open(path, "rb") as f:
        return f.read(4) == b"PK\x03\x04"


def _extract_marketing_zip(zip_path: Path, output_dir: Path) -> Optional[Path]:
    """
    Extract marketing ZIP to output_dir/marketing_extract_<timestamp>.
    If ZIP has marketing_* folder(s), return output_dir so find_marketing_folders finds them.
    If ZIP has only CSVs at root, create marketing_report/ and put CSVs there, return extract_dir.
    Returns path to use as marketing_folder_path for marketing analysis.
    """
    extract_dir = output_dir / f"marketing_extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(extract_dir)
    # Check structure: any dir named marketing_*?
    subdirs = [d for d in extract_dir.iterdir() if d.is_dir() and d.name.lower().startswith("marketing_")]
    if subdirs:
        logger.info("MarketingAgent: Found marketing_* dirs in ZIP, using %s", extract_dir)
        return extract_dir
    # CSVs at root or one level down?
    promotion_files = list(extract_dir.glob("MARKETING_PROMOTION*.csv"))
    sponsored_files = list(extract_dir.glob("MARKETING_SPONSORED_LISTING*.csv"))
    for d in extract_dir.iterdir():
        if d.is_dir():
            promotion_files.extend(d.glob("MARKETING_PROMOTION*.csv"))
            sponsored_files.extend(d.glob("MARKETING_SPONSORED_LISTING*.csv"))
    if promotion_files or sponsored_files:
        one_dir = extract_dir / "marketing_report"
        one_dir.mkdir(exist_ok=True)
        for f in promotion_files + sponsored_files:
            dest = one_dir / f.name
            dest.write_bytes(f.read_bytes())
        logger.info("MarketingAgent: Placed CSVs in %s", one_dir.name)
    return extract_dir


def _write_marketing_excel(
    output_dir: Path,
    promotion_table,
    sponsored_table,
    combined_table,
    promotion_by_campaign=None,
    promotion_by_store=None,
    sponsored_by_campaign=None,
    sponsored_by_store=None,
    store_wise_marketing=None,
    operator_name: Optional[str] = None,
) -> Path:
    """Write promotion, sponsored, combined, by-campaign, and by-store tables to an Excel file. Returns path to file."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) else None)
    filename = f"{tag}_marketing_analysis_{timestamp}.xlsx" if tag else f"marketing_analysis_{timestamp}.xlsx"
    filepath = output_dir / filename

    def _normalize_store_column(df):
        """Rename Store ID / Merchant store ID to Merchant Store ID for consistent output."""
        if df is None or not hasattr(df, "columns"):
            return df
        out = df.copy()
        for old in ("Store ID", "Merchant store ID"):
            if old in out.columns and old != "Merchant Store ID":
                out = out.rename(columns={old: "Merchant Store ID"})
                break
        return out

    def add_sheet(ws, df, title):
        if df is None or (hasattr(df, "empty") and df.empty):
            return
        if hasattr(df, "reset_index") and df.index.name:
            df_export = df.reset_index()
        else:
            df_export = df
        df_export = _normalize_store_column(df_export)
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)

    if combined_table is not None and not (getattr(combined_table, "empty", True)):
        ws = wb.create_sheet("Corporate vs TODC (Combined)")
        ws.cell(row=1, column=1, value="Combined: Corporate vs TODC").font = Font(bold=True, size=12)
        add_sheet(ws, combined_table, "Combined")
    if promotion_table is not None and not (getattr(promotion_table, "empty", True)):
        ws = wb.create_sheet("Promotion")
        ws.cell(row=1, column=1, value="Promotion by Campaign").font = Font(bold=True, size=12)
        add_sheet(ws, promotion_table, "Promotion")
    if sponsored_table is not None and not (getattr(sponsored_table, "empty", True)):
        ws = wb.create_sheet("Sponsored Listing")
        ws.cell(row=1, column=1, value="Sponsored Listing by Campaign").font = Font(bold=True, size=12)
        add_sheet(ws, sponsored_table, "Sponsored")
    if promotion_by_campaign is not None and not (getattr(promotion_by_campaign, "empty", True)):
        ws = wb.create_sheet("Promotion by Campaign Name")
        ws.cell(row=1, column=1, value="Promotion: Campaign name, Spend, Sales, Orders, ROAS, Cost per Order").font = Font(bold=True, size=12)
        add_sheet(ws, promotion_by_campaign, "Promotion by Campaign")
    if promotion_by_store is not None and not (getattr(promotion_by_store, "empty", True)):
        ws = wb.create_sheet("Promotion by Store")
        ws.cell(row=1, column=1, value="Promotion: Merchant Store ID, Spend, Sales, Orders, ROAS, Cost per Order").font = Font(bold=True, size=12)
        add_sheet(ws, promotion_by_store, "Promotion by Store")
    if sponsored_by_campaign is not None and not (getattr(sponsored_by_campaign, "empty", True)):
        ws = wb.create_sheet("Sponsored by Campaign Name")
        ws.cell(row=1, column=1, value="Sponsored: Campaign name, Spend, Sales, Orders, ROAS, Cost per Order").font = Font(bold=True, size=12)
        add_sheet(ws, sponsored_by_campaign, "Sponsored by Campaign")
    if sponsored_by_store is not None and not (getattr(sponsored_by_store, "empty", True)):
        ws = wb.create_sheet("Sponsored by Store")
        ws.cell(row=1, column=1, value="Sponsored: Merchant Store ID, Spend, Sales, Orders, ROAS, Cost per Order").font = Font(bold=True, size=12)
        add_sheet(ws, sponsored_by_store, "Sponsored by Store")
    if store_wise_marketing is not None and not (getattr(store_wise_marketing, "empty", True)):
        ws = wb.create_sheet("Store-wise")
        ws.cell(row=1, column=1, value="Store-wise (Combined): Merchant Store ID, Orders, Sales, Spend, ROAS, Cost per Order").font = Font(bold=True, size=12)
        add_sheet(ws, store_wise_marketing, "Store-wise")
    if not wb.sheetnames:
        ws = wb.create_sheet("Summary")
        ws.cell(row=1, column=1, value="No marketing data found for the selected date range.")

    wb.save(filepath)
    logger.info("MarketingAgent: Wrote %s", filepath.name)
    return filepath


def run(
    downloaded_path: Path,
    output_dir: Path,
    post_start_date: str,
    post_end_date: str,
    excluded_dates: Optional[list] = None,
    operator_name: Optional[str] = None,
    write_file: bool = True,
):
    """
    Run marketing analysis on the downloaded report. If it's a ZIP, extract it first.
    Uses marketing_analysis.create_corporate_vs_todc_table (module must be on PYTHONPATH).
    If write_file=True, writes marketing_analysis_<timestamp>.xlsx and returns path.
    If write_file=False, returns list of (sheet_name, DataFrame) for combined report.
    """
    if pd is None:
        raise RuntimeError("pandas is required for MarketingAgent. Install with: pip install pandas")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    downloaded_path = Path(downloaded_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Resolve marketing folder: extract ZIP if needed
    if downloaded_path.is_file() and _is_zip(downloaded_path):
        marketing_folder = _extract_marketing_zip(downloaded_path, output_dir)
    elif downloaded_path.is_dir():
        marketing_folder = downloaded_path
    else:
        logger.warning("MarketingAgent: Expected a ZIP file or folder, got %s", downloaded_path)
        return None

    if not marketing_folder or not marketing_folder.is_dir():
        logger.warning("MarketingAgent: No marketing folder to process")
        return None

    from agents.marketing_data import create_corporate_vs_todc_table

    excluded_dates = excluded_dates or []
    try:
        promotion_table, sponsored_table, combined_table = create_corporate_vs_todc_table(
            excluded_dates=excluded_dates,
            pre_start_date=post_start_date,
            pre_end_date=post_end_date,
            post_start_date=post_start_date,
            post_end_date=post_end_date,
            marketing_folder_path=marketing_folder,
        )
    except Exception as e:
        logger.warning("MarketingAgent: create_corporate_vs_todc_table failed: %s", e)
        return None

    # By-campaign / by-store breakdowns not yet implemented
    promotion_by_campaign = promotion_by_store = sponsored_by_campaign = sponsored_by_store = store_wise_marketing = None

    sheets_list: List[Tuple[str, object]] = []
    if combined_table is not None and not (getattr(combined_table, "empty", True)):
        sheets_list.append(("Corporate vs TODC (Combined)", combined_table))
    if promotion_table is not None and not (getattr(promotion_table, "empty", True)):
        sheets_list.append(("Promotion", promotion_table))
    if sponsored_table is not None and not (getattr(sponsored_table, "empty", True)):
        sheets_list.append(("Sponsored Listing", sponsored_table))
    if promotion_by_campaign is not None and not (getattr(promotion_by_campaign, "empty", True)):
        sheets_list.append(("Promotion by Campaign Name", promotion_by_campaign))
    if promotion_by_store is not None and not (getattr(promotion_by_store, "empty", True)):
        sheets_list.append(("Promotion by Store", promotion_by_store))
    if sponsored_by_campaign is not None and not (getattr(sponsored_by_campaign, "empty", True)):
        sheets_list.append(("Sponsored by Campaign Name", sponsored_by_campaign))
    if sponsored_by_store is not None and not (getattr(sponsored_by_store, "empty", True)):
        sheets_list.append(("Sponsored by Store", sponsored_by_store))
    if store_wise_marketing is not None and not (getattr(store_wise_marketing, "empty", True)):
        sheets_list.append(("Store-wise", store_wise_marketing))
    if not sheets_list:
        sheets_list.append(("Summary", pd.DataFrame({"Message": ["No marketing data found for the selected date range."]})))

    if not write_file:
        return sheets_list
    try:
        report_path = _write_marketing_excel(
            output_dir,
            promotion_table,
            sponsored_table,
            combined_table,
            promotion_by_campaign=promotion_by_campaign,
            promotion_by_store=promotion_by_store,
            sponsored_by_campaign=sponsored_by_campaign,
            sponsored_by_store=sponsored_by_store,
            store_wise_marketing=store_wise_marketing,
            operator_name=operator_name,
        )
        return report_path
    except Exception as e:
        logger.warning("MarketingAgent: Failed to write Excel: %s", e)
        return None
