"""MarketingReco agent with DeepDive, manual upload, and auto modes."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Literal

from shared.config.settings import data_root, marketingreco_reporting_root
from shared.models.campaign import RecommendedCampaign
from shared.models.report import DeepDiveReport, MarketingPlan
from shared.utils.date_helpers import utc_now_iso

from .approval_handler import apply_command
from .plan_generator import generate_plan

MarketingRecoMode = Literal["deepdive", "manual", "auto"]


def _reporting_subprocess_env(reporting_root: Path) -> dict[str, str]:
    """Ensure ``python -c`` subprocesses resolve ``agents.*`` to the reporting tree, not repo ``agents/``."""
    env = os.environ.copy()
    env["PYTHONPATH"] = str(reporting_root.resolve())
    return env


def _deepdive_path(operator_id: str) -> Path:
    return data_root() / "operators" / operator_id / "reports" / "deepdive.json"


def _plan_path(operator_id: str) -> Path:
    return data_root() / "operators" / operator_id / "reports" / "marketing_plan.json"


def _save_plan(operator_id: str, plan: MarketingPlan) -> dict[str, Any]:
    path = _plan_path(operator_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(plan.model_dump_json(indent=2), encoding="utf-8")
    return json.loads(plan.model_dump_json())


def _read_campaign_mappings(combined_analysis_path: Path) -> list[dict[str, Any]]:
    """
    Read Campaign Mappings sheet from combined analysis (Reporting pipeline).

    Column A is the store key from Day-Slot sheets (DoorDash Merchant store ID).
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read campaign mappings") from exc

    if not combined_analysis_path.is_file():
        raise FileNotFoundError(f"combined analysis file not found: {combined_analysis_path}")
    wb = openpyxl.load_workbook(combined_analysis_path, read_only=True, data_only=True)
    try:
        if "Campaign Mappings" not in wb.sheetnames:
            return []
        ws = wb["Campaign Mappings"]
        out: list[dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1 or not row:
                continue
            if len(row) < 5:
                continue
            store_id = str(row[0] or "").strip()
            store_name = str(row[1] or "").strip() if len(row) > 1 else ""
            min_subtotal = float(row[2] or 0)
            slot_tags = str(row[3] or "").strip()
            campaign_name = str(row[4] or "").strip()
            status = str(row[5] or "Pending").strip() if len(row) > 5 else "Pending"
            if not campaign_name:
                continue
            out.append(
                {
                    "store_id": store_id,
                    "store_name": store_name,
                    "min_subtotal": min_subtotal,
                    "slot_tags": slot_tags,
                    "campaign_name": campaign_name,
                    "status": status,
                }
            )
        return out
    finally:
        wb.close()


def _slot_tags_to_target_parts(slot_tags: Any) -> list[str]:
    """Map Campaign Mappings slot_tags (list or comma string) to target_day_parts for UI / exports."""
    if isinstance(slot_tags, list):
        out: list[str] = []
        for t in slot_tags:
            if t is None:
                continue
            try:
                out.append(str(int(t)))
            except (TypeError, ValueError):
                s = str(t).strip()
                if s:
                    out.append(s)
        return out
    s = str(slot_tags or "").strip()
    if not s:
        return []
    return [t.strip() for t in s.replace("，", ",").split(",") if t.strip()]


def _campaigns_from_mappings(
    operator_id: str,
    mappings: list[dict[str, Any]],
) -> MarketingPlan:
    campaigns: list[RecommendedCampaign] = []
    for m in mappings:
        tags = _slot_tags_to_target_parts(m.get("slot_tags"))
        campaigns.append(
            RecommendedCampaign(
                campaign_type="promo",
                campaign_name=m.get("campaign_name", "Campaign"),
                budget=0.0,
                start_date=utc_now_iso(),
                duration_days=7,
                target_day_parts=tags,
                target_items=[],
                discount_pct=15.0,
                rationale=(
                    f"Built from mapping for store {m.get('store_id') or 'unknown'} "
                    f"(min subtotal {m.get('min_subtotal')}, slot tags {tags or m.get('slot_tags')}, "
                    f"status {m.get('status', 'Pending')})."
                )[:500],
            )
        )
    if not campaigns:
        campaigns.append(
            RecommendedCampaign(
                campaign_type="promo",
                campaign_name="Fallback mapping campaign",
                budget=0.0,
                start_date=utc_now_iso(),
                duration_days=7,
                target_day_parts=[],
                target_items=[],
                discount_pct=15.0,
                rationale="No campaign mappings found in combined analysis; created fallback recommendation.",
            )
        )
    return MarketingPlan(
        operator_id=operator_id,
        plan_date=utc_now_iso(),
        recommended_campaigns=campaigns,
        approval_status="pending",
        approver_notes="",
    )


def _align_store_ids_with_financial_mapping(
    financial_source: str | Path | None,
    ads_plan: dict[str, Any] | None,
) -> None:
    """Resolve Store ID -> Merchant store ID from FINANCIAL_DETAILED for Ads."""
    if not financial_source:
        return
    path = Path(financial_source)
    if not path.is_file():
        return
    from .ads_planner import (
        apply_financial_store_to_merchant_map,
        build_store_to_merchant_from_financial_path,
    )

    store_to_merchant = build_store_to_merchant_from_financial_path(path)
    if not store_to_merchant:
        return
    if ads_plan is not None:
        apply_financial_store_to_merchant_map(ads_plan, store_to_merchant)


def _try_build_ads_plan(csv_path: Path) -> dict[str, Any] | None:
    if not csv_path.is_file():
        return None
    try:
        from .ads_planner import build_ads_plan

        return build_ads_plan(str(csv_path))
    except Exception:
        return None


def _parse_subprocess_ads_csv(stdout: str) -> str | None:
    for line in (stdout or "").splitlines():
        if line.startswith("ADS_FINANCIAL_CSV="):
            return line.split("=", 1)[1].strip() or None
    return None


def _build_combined_from_financial_input(reporting_root: Path, financial_report_path: Path) -> tuple[Path, dict[str, Any] | None]:
    script = """
from pathlib import Path
from datetime import datetime, timedelta
from agents.analysis_agent import run as analysis_run
from agents.combined_report_agent import run as combined_run, append_campaign_mappings_to_workbook
from agents.campaign_params import get_campaign_mappings_for_combined

def _dates():
    today = datetime.now().date()
    first_this_month = today.replace(day=1)
    last_prev_month = first_this_month - timedelta(days=1)
    y, m = first_this_month.year, first_this_month.month - 3
    if m <= 0:
        m += 12
        y -= 1
    start = datetime(y, m, 1).date()
    return start.strftime("%m/%d/%Y"), last_prev_month.strftime("%m/%d/%Y")

root = Path(".")
run_dir = root / "downloads" / f"manual-financial-{datetime.now().strftime('%Y%m%d_%H%M%S')}"
run_dir.mkdir(parents=True, exist_ok=True)
start_date, end_date = _dates()
financial_zip = Path(__import__("os").environ["FINANCIAL_REPORT_ZIP"])
financial_sheets = analysis_run(financial_zip, output_dir=run_dir, report_start_date=start_date, report_end_date=end_date, write_file=False)
fc = run_dir / "financial_detailed_report.csv"
if not fc.is_file():
    for p in sorted(run_dir.glob("*FINANCIAL*.csv")):
        fc = p
        break
if fc.is_file():
    print(f"ADS_FINANCIAL_CSV={fc.resolve()}")
combined = combined_run(financial_sheets=financial_sheets, marketing_sheets=None, output_dir=run_dir)
if combined:
    slots_csv = root / "slots.csv"
    mappings = get_campaign_mappings_for_combined(Path(combined), slots_csv)
    if mappings:
        append_campaign_mappings_to_workbook(Path(combined), mappings)
    print(f"COMBINED_PATH={combined}")
else:
    print("COMBINED_PATH=")
"""
    with tempfile.TemporaryDirectory(prefix="marketingreco_fin_") as td:
        temp_zip = Path(td) / "financial_input.zip"
        if financial_report_path.suffix.lower() == ".zip":
            temp_zip.write_bytes(financial_report_path.read_bytes())
        else:
            with zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                zf.write(financial_report_path, arcname="FINANCIAL_DETAILED_upload.csv")
        env = _reporting_subprocess_env(reporting_root)
        env["FINANCIAL_REPORT_ZIP"] = str(temp_zip)
        proc = subprocess.run(
            [sys.executable, "-c", script],
            cwd=str(reporting_root),
            env=env,
            check=True,
            capture_output=True,
            text=True,
        )
    ads_csv = _parse_subprocess_ads_csv(proc.stdout or "")
    ads_plan: dict[str, Any] | None = None
    if ads_csv:
        ads_plan = _try_build_ads_plan(Path(ads_csv))
    if ads_plan is None and financial_report_path.suffix.lower() == ".csv":
        ads_plan = _try_build_ads_plan(financial_report_path)
    for line in reversed((proc.stdout or "").splitlines()):
        if line.startswith("COMBINED_PATH="):
            value = line.split("=", 1)[1].strip()
            if value:
                p = Path(value)
                combined = p if p.is_absolute() else reporting_root / p
                return combined, ads_plan
            break
    raise RuntimeError("Failed to build combined analysis from FINANCIAL_DETAILED input")


def _run_manual_from_register(
    operator_id: str,
    *,
    register_report_path: str,
    reporting_root: str,
) -> dict[str, Any]:
    from .register_reco import build_recommendations_from_register

    root = Path(reporting_root)
    slots_csv = root / "slots.csv"
    built = build_recommendations_from_register(
        Path(register_report_path),
        slots_csv=slots_csv,
    )
    mappings = built.get("campaign_mappings") or []
    ads_plan = built.get("ads_plan")
    plan = _campaigns_from_mappings(operator_id, mappings)
    out = _save_plan(operator_id, plan)
    out["campaign_mappings"] = mappings
    out["slot_recommendations"] = built.get("slot_recommendations") or []
    out["ads_plan"] = ads_plan
    out["input_type"] = "register"
    return out


def _run_manual_mode(
    operator_id: str,
    *,
    financial_report_path: str | None = None,
    register_report_path: str | None = None,
    reporting_root: str,
) -> dict[str, Any]:
    if register_report_path:
        return _run_manual_from_register(
            operator_id,
            register_report_path=register_report_path,
            reporting_root=reporting_root,
        )
    if not financial_report_path:
        raise ValueError(
            "manual mode requires register_report_path (DD register Excel/CSV) "
            "or financial_report_path (legacy FINANCIAL_DETAILED)"
        )
    combined_path, ads_plan = _build_combined_from_financial_input(
        Path(reporting_root),
        Path(financial_report_path),
    )
    mappings = _read_campaign_mappings(combined_path)
    _align_store_ids_with_financial_mapping(financial_report_path, ads_plan)
    plan = _campaigns_from_mappings(operator_id, mappings)
    out = _save_plan(operator_id, plan)
    out["campaign_mappings"] = mappings
    out["ads_plan"] = ads_plan
    out["input_type"] = "financial"
    return out


def _latest_combined_for_email(reporting_root: Path, email: str) -> Path | None:
    safe = (email or "run").strip()
    for c in ("@", ".", " ", "/", "\\"):
        safe = safe.replace(c, "_")
    safe = safe[:50] if len(safe) > 50 else safe
    downloads_root = reporting_root / "downloads"
    if not downloads_root.is_dir():
        return None
    runs = sorted(downloads_root.glob(f"{safe}-*"), reverse=True)
    for run_dir in runs:
        combined = sorted(run_dir.glob("combined_analysis_*.xlsx"), reverse=True)
        if combined:
            return combined[0]
    return None


def _run_auto_mode(
    operator_id: str,
    *,
    doordash_email: str,
    doordash_password: str,
    reporting_root: str,
) -> dict[str, Any]:
    root = Path(reporting_root)
    main_py = root / "main.py"
    if not main_py.is_file():
        raise FileNotFoundError(f"reporting workflow not found: {main_py}")
    env = _reporting_subprocess_env(root)
    env["DOORDASH_EMAIL"] = doordash_email
    env["DOORDASH_PASSWORD"] = doordash_password
    script = """
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
from agents.doordash_agent import run_reports_only
from agents.marketing_agent import run as marketing_run
from agents.analysis_agent import run as analysis_run
from agents.combined_report_agent import run as combined_run, append_campaign_mappings_to_workbook
from agents.campaign_params import get_campaign_mappings_for_combined

def _dates():
    today = datetime.now().date()
    first_this_month = today.replace(day=1)
    last_prev_month = first_this_month - timedelta(days=1)
    y, m = first_this_month.year, first_this_month.month - 3
    if m <= 0:
        m += 12
        y -= 1
    start = datetime(y, m, 1).date()
    return start.strftime("%m/%d/%Y"), last_prev_month.strftime("%m/%d/%Y")

def _run_dir(email: str) -> Path:
    safe = (email or "run").strip()
    for c in ("@", ".", " ", "/", "\\\\"):
        safe = safe.replace(c, "_")
    safe = safe[:50] if len(safe) > 50 else safe
    return Path("downloads") / f"{safe}-{datetime.now().strftime('%Y%m%d_%H%M%S')}"

async def _main():
    import os
    email = os.environ["DOORDASH_EMAIL"]
    password = os.environ["DOORDASH_PASSWORD"]
    start_date, end_date = _dates()
    run_dir = _run_dir(email)
    run_dir.mkdir(parents=True, exist_ok=True)
    marketing_path, financial_path = await run_reports_only(
        download_dir=run_dir,
        email=email,
        password=password,
        start_date=start_date,
        end_date=end_date,
    )
    marketing_sheets = marketing_run(Path(marketing_path), output_dir=run_dir, post_start_date=start_date, post_end_date=end_date, write_file=False) if marketing_path else None
    financial_sheets = analysis_run(Path(financial_path), output_dir=run_dir, report_start_date=start_date, report_end_date=end_date, write_file=False) if financial_path else None
    fc = run_dir / "financial_detailed_report.csv"
    if not fc.is_file():
        for p in sorted(run_dir.glob("*FINANCIAL*.csv")):
            fc = p
            break
    if fc.is_file():
        print(f"ADS_FINANCIAL_CSV={fc.resolve()}")
    combined = combined_run(financial_sheets=financial_sheets, marketing_sheets=marketing_sheets, output_dir=run_dir)
    if combined:
        slots_csv = Path("slots.csv")
        mappings = get_campaign_mappings_for_combined(Path(combined), slots_csv)
        if mappings:
            append_campaign_mappings_to_workbook(Path(combined), mappings)
        print(f"COMBINED_PATH={combined}")
    else:
        print("COMBINED_PATH=")

asyncio.run(_main())
"""
    proc = subprocess.run(
        [sys.executable, "-c", script],
        cwd=str(root),
        env=env,
        check=True,
        capture_output=True,
        text=True,
    )
    ads_csv = _parse_subprocess_ads_csv(proc.stdout or "")
    ads_plan: dict[str, Any] | None = _try_build_ads_plan(Path(ads_csv)) if ads_csv else None
    combined: Path | None = None
    for line in reversed((proc.stdout or "").splitlines()):
        if line.startswith("COMBINED_PATH="):
            value = line.split("=", 1)[1].strip()
            combined = Path(value) if value else None
            break
    if combined and not combined.is_absolute():
        combined = root / combined
    if not combined:
        combined = _latest_combined_for_email(root, doordash_email)
    if not combined:
        raise RuntimeError("auto mode finished but no combined_analysis file was found")
    mappings = _read_campaign_mappings(combined)
    _align_store_ids_with_financial_mapping(ads_csv, ads_plan)
    plan = _campaigns_from_mappings(operator_id, mappings)
    out = _save_plan(operator_id, plan)
    out["campaign_mappings"] = mappings
    out["ads_plan"] = ads_plan
    return out


def run(
    operator_id: str,
    *,
    mode: MarketingRecoMode = "deepdive",
    deepdive_report: dict[str, Any] | None = None,
    financial_report_path: str | None = None,
    register_report_path: str | None = None,
    doordash_email: str | None = None,
    doordash_password: str | None = None,
    reporting_root: str | None = None,
    operator_profile: dict[str, Any] | None = None,
    budget_cap: float | None = None,
    campaign_history: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if reporting_root is None:
        reporting_root = str(marketingreco_reporting_root())
    if mode == "manual":
        if not register_report_path and not financial_report_path:
            raise ValueError(
                "manual mode requires register_report_path (DD register Excel/CSV) "
                "or financial_report_path (legacy FINANCIAL_DETAILED)"
            )
        return _run_manual_mode(
            operator_id,
            financial_report_path=financial_report_path,
            register_report_path=register_report_path,
            reporting_root=reporting_root,
        )
    if mode == "auto":
        if not doordash_email or not doordash_password:
            raise ValueError("auto mode requires doordash_email and doordash_password")
        return _run_auto_mode(
            operator_id,
            doordash_email=doordash_email,
            doordash_password=doordash_password,
            reporting_root=reporting_root,
        )

    _ = operator_profile
    _ = campaign_history
    if deepdive_report is None:
        raw = _deepdive_path(operator_id).read_text(encoding="utf-8")
        dd = DeepDiveReport.model_validate_json(raw)
    else:
        dd = DeepDiveReport.model_validate(deepdive_report)
    plan = generate_plan(dd, budget_cap=budget_cap)
    return _save_plan(operator_id, plan)


def approve(operator_id: str, command: str, notes: str = "") -> dict[str, Any]:
    path = _plan_path(operator_id)
    plan = MarketingPlan.model_validate_json(path.read_text(encoding="utf-8"))
    if command not in ("approve", "reject", "modify"):
        raise ValueError("command must be approve|reject|modify")
    apply_command(plan, command, notes)  # type: ignore[arg-type]
    path.write_text(plan.model_dump_json(indent=2), encoding="utf-8")
    return json.loads(plan.model_dump_json())


if __name__ == "__main__":
    import sys

    oid = sys.argv[1] if len(sys.argv) > 1 else "dev_operator"
    print(json.dumps(run(oid), indent=2))
