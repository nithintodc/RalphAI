"""Ralph Analyse — DoorDash + UberEats pre/post period comparison dashboard."""

import streamlit as st
import pandas as pd
import numpy as np
import tempfile
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO

from data_loader import (
    load_dd_financial, load_ue_financial,
    load_marketing_promotion, load_marketing_sponsored,
    filter_by_date_range, filter_excluded_dates,
    aggregate_dd, aggregate_ue, get_last_year_dates, classify_file,
)
from analysis import (
    process_metric, create_summary_tables, create_combined_summary,
    get_store_table_prepost, get_store_table_yoy,
)
from marketing import create_corporate_vs_todc

st.set_page_config(page_title="Ralph Analyse", page_icon="📊", layout="wide")

# ── Global CSS ──────────────────────────────────────────────────────────────
st.markdown("""<style>
:root {
    --bg: #FAFAF9; --surface: #FFFFFF; --border: #E7E5E4;
    --text: #0C0A09; --muted: #57534E; --primary: #059669;
}
html,body,.stApp,p,label{font-family:Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif!important}
.stApp{background:var(--bg)!important}
#MainMenu,footer{visibility:hidden}
header{background:transparent!important}
h1,h2,h3,h4{color:var(--text)!important;letter-spacing:-0.02em}
.block-container{padding-top:1rem;padding-bottom:3rem; max-width: none !important;}
section[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border)!important}
.kpi-card{background:#fff;border:1px solid var(--border);border-radius:12px;padding:20px 24px;text-align:center}
.kpi-value{font-size:1.8rem;font-weight:700;color:var(--text);margin:4px 0}
.kpi-label{font-size:0.82rem;color:var(--muted);font-weight:500;text-transform:uppercase;letter-spacing:0.04em}
.kpi-delta{font-size:0.9rem;font-weight:600}
.kpi-up{color:#059669}.kpi-down{color:#DC2626}.kpi-flat{color:var(--muted)}
.section-hdr{font-size:1.1rem;font-weight:700;color:var(--text);margin:1.5rem 0 0.5rem;padding-bottom:6px;border-bottom:2px solid var(--primary)}
.dataframe{font-size:0.85rem!important}
div[data-testid="stDataFrame"]{border:1px solid var(--border);border-radius:8px;overflow:hidden}
.stButton>button[data-testid="baseButton-primary"]{
    background:linear-gradient(135deg,#059669,#047857)!important;color:#fff!important;
    border:none!important;border-radius:8px;font-weight:600;padding:0.55rem 1.5rem;
    box-shadow:0 2px 8px rgba(5,150,105,0.25)}
.stButton>button[data-testid="baseButton-primary"]:hover{
    background:linear-gradient(135deg,#047857,#065F46)!important;transform:translateY(-1px);
    box-shadow:0 4px 16px rgba(5,150,105,0.35)!important}
.step-row{display:flex;gap:12px;margin-bottom:1.2rem}
.step-badge{width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;
    font-weight:700;font-size:0.85rem;flex-shrink:0}
.step-active .step-badge{background:#059669;color:#fff}
.step-done .step-badge{background:#047857;color:#fff}
.step-pending .step-badge{background:#E7E5E4;color:#57534E}
.step-label{line-height:32px;font-size:0.9rem;color:var(--muted)}
.step-active .step-label{color:var(--text);font-weight:600}
.step-done .step-label{color:#047857;font-weight:500}

/* Force light backgrounds on all inputs */
.stDateInput input, .stTextInput input, .stNumberInput input, .stSelectbox select,
div[data-baseweb="input"] input, div[data-baseweb="select"] div,
.stDateInput div[data-baseweb="input"], .stTextInput div[data-baseweb="input"] {
    background-color: #FFFFFF !important;
    color: #101828 !important;
    border: 1.5px solid #D0D5DD !important;
    border-radius: 8px !important;
}
.stDateInput div[data-baseweb="input"]{background:#FFFFFF!important}
.stDateInput input{color:#101828!important;-webkit-text-fill-color:#101828!important}
.stTextInput input{color:#101828!important;background:#FFFFFF!important;-webkit-text-fill-color:#101828!important}

/* Labels and markdown bold text */
.stMarkdown strong, .stMarkdown b, label, .stDateInput label, .stTextInput label,
p, .stMarkdown p {
    color: #101828 !important;
    -webkit-text-fill-color: #101828 !important;
}

/* Fix excluded dates text input placeholder */
.stTextInput input::placeholder{color:#98A2B3!important;-webkit-text-fill-color:#98A2B3!important}

/* Info box text */
div[data-testid="stAlert"] p, div[data-testid="stAlert"] span,
.stAlert p, .stAlert span {
    color: #1D4ED8 !important;
    -webkit-text-fill-color: #1D4ED8 !important;
}

/* Caption text */
.stCaption, .stCaption p { color: #667085 !important; -webkit-text-fill-color: #667085 !important; }

/* Multiselect */
div[data-baseweb="select"] span, div[data-baseweb="tag"] span {
    color: #101828 !important; -webkit-text-fill-color: #101828 !important;
}
div[data-baseweb="popover"] li { color: #101828 !important; }
</style>""", unsafe_allow_html=True)


# ── Session state defaults ──────────────────────────────────────────────────
for key, default in [
    ("step", 1), ("files", {}), ("tmp_dir", None),
    ("dd_raw", None), ("ue_raw", None), ("promo_raw", None), ("sponsored_raw", None),
    ("analysis_done", False), ("results", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ═══════════════════════════════════════════════════════════════════════════
# Helper functions (defined before the if/elif flow)
# ═══════════════════════════════════════════════════════════════════════════

def _render_stepper(current):
    labels = ["Upload Files", "Configure Dates", "Select Filters", "Dashboard"]
    html = '<div class="step-row">'
    for i, label in enumerate(labels, 1):
        cls = "step-done" if i < current else ("step-active" if i == current else "step-pending")
        icon = "✓" if i < current else str(i)
        html += f'<div class="{cls}" style="display:flex;gap:8px;align-items:center"><div class="step-badge">{icon}</div><div class="step-label">{label}</div></div>'
        if i < len(labels):
            html += '<div style="flex:1;border-bottom:2px solid #E5E7EB;align-self:center;margin:0 4px"></div>'
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def _load_raw_data():
    """Parse uploaded CSVs to extract store lists for the filter step."""
    files = st.session_state.files
    if "dd_financial" in files:
        try:
            df, dc, sc, *_ = load_dd_financial(files["dd_financial"])
            st.session_state.dd_raw = (df, dc, sc)
            st.session_state.dd_stores = sorted(df[sc].unique().tolist())
        except Exception as e:
            st.error(f"Error loading DD financial: {e}")
            st.session_state.dd_stores = []
    else:
        st.session_state.dd_stores = []

    if "ue_financial" in files:
        try:
            df, dc, sc, *_ = load_ue_financial(files["ue_financial"])
            st.session_state.ue_raw = (df, dc, sc)
            st.session_state.ue_stores = sorted(df[sc].unique().tolist())
        except Exception as e:
            st.error(f"Error loading UE financial: {e}")
            st.session_state.ue_stores = []
    else:
        st.session_state.ue_stores = []

    if "dd_mkt_promo" in files:
        try:
            df, sc = load_marketing_promotion(files["dd_mkt_promo"])
            st.session_state.promo_raw = (df, sc)
        except Exception as e:
            st.error(f"Error loading marketing promo: {e}")

    if "dd_mkt_sponsored" in files:
        try:
            df, sc = load_marketing_sponsored(files["dd_mkt_sponsored"])
            st.session_state.sponsored_raw = (df, sc)
        except Exception as e:
            st.error(f"Error loading marketing sponsored: {e}")


def _run_analysis():
    """Run the full analysis pipeline."""
    dates = st.session_state.dates
    ps, pe = dates["pre_start"], dates["pre_end"]
    pos, poe = dates["post_start"], dates["post_end"]
    excluded = dates["excluded"]

    ly_ps, ly_pe = get_last_year_dates(ps, pe)
    ly_pos, ly_poe = get_last_year_dates(pos, poe)

    dd_selected = st.session_state.dd_selected
    ue_selected = st.session_state.ue_selected

    results = {}

    # ── DoorDash ────────────────────────────────────────────────────────
    if st.session_state.dd_raw is not None:
        full = load_dd_financial(st.session_state.files["dd_financial"])
        raw_df, dc, sc, sal_c, pay_c, ord_c = full

        def _dd_window(start, end):
            w = filter_by_date_range(raw_df, dc, start, end)
            w = filter_excluded_dates(w, dc, excluded)
            return aggregate_dd(w, sc, sal_c, pay_c, ord_c)

        dd_pre24 = _dd_window(ly_ps, ly_pe)
        dd_post24 = _dd_window(ly_pos, ly_poe)
        dd_pre25 = _dd_window(ps, pe)
        dd_post25 = _dd_window(pos, poe)

        dd_sales = process_metric(dd_pre24, dd_post24, dd_pre25, dd_post25, "Sales")
        dd_payouts = process_metric(dd_pre24, dd_post24, dd_pre25, dd_post25, "Payouts")
        dd_orders = process_metric(dd_pre24, dd_post24, dd_pre25, dd_post25, "Orders")

        results["dd_sales"] = dd_sales
        results["dd_payouts"] = dd_payouts
        results["dd_orders"] = dd_orders
        results["dd_selected"] = dd_selected

        dd_t1, dd_t2 = create_summary_tables(dd_sales, dd_payouts, dd_orders,
                                              pd.DataFrame(), dd_selected, include_nc=False)
        results["dd_summary_t1"] = dd_t1
        results["dd_summary_t2"] = dd_t2
    else:
        results["dd_sales"] = pd.DataFrame()
        results["dd_payouts"] = pd.DataFrame()
        results["dd_orders"] = pd.DataFrame()
        results["dd_selected"] = []
        results["dd_summary_t1"] = pd.DataFrame()
        results["dd_summary_t2"] = pd.DataFrame()

    # ── UberEats ────────────────────────────────────────────────────────
    if st.session_state.ue_raw is not None:
        full = load_ue_financial(st.session_state.files["ue_financial"])
        raw_df, dc, sc, sal_c, pay_c, ord_c = full

        def _ue_window(start, end):
            w = filter_by_date_range(raw_df, dc, start, end)
            w = filter_excluded_dates(w, dc, excluded)
            return aggregate_ue(w, sc, sal_c, pay_c, ord_c)

        ue_pre24 = _ue_window(ly_ps, ly_pe)
        ue_post24 = _ue_window(ly_pos, ly_poe)
        ue_pre25 = _ue_window(ps, pe)
        ue_post25 = _ue_window(pos, poe)

        ue_sales = process_metric(ue_pre24, ue_post24, ue_pre25, ue_post25, "Sales")
        ue_payouts = process_metric(ue_pre24, ue_post24, ue_pre25, ue_post25, "Payouts")
        ue_orders = process_metric(ue_pre24, ue_post24, ue_pre25, ue_post25, "Orders")

        results["ue_sales"] = ue_sales
        results["ue_payouts"] = ue_payouts
        results["ue_orders"] = ue_orders
        results["ue_selected"] = ue_selected

        ue_t1, ue_t2 = create_summary_tables(ue_sales, ue_payouts, ue_orders,
                                              pd.DataFrame(), ue_selected, include_nc=False)
        results["ue_summary_t1"] = ue_t1
        results["ue_summary_t2"] = ue_t2
    else:
        results["ue_sales"] = pd.DataFrame()
        results["ue_payouts"] = pd.DataFrame()
        results["ue_orders"] = pd.DataFrame()
        results["ue_selected"] = []
        results["ue_summary_t1"] = pd.DataFrame()
        results["ue_summary_t2"] = pd.DataFrame()

    # ── New Customers (DD marketing promo) ──────────────────────────────
    dd_nc = pd.DataFrame()
    if st.session_state.promo_raw is not None:
        promo_df, promo_sc = st.session_state.promo_raw
        nc_col = "New Customers" if "New Customers" in promo_df.columns else None
        if nc_col:
            def _nc_window(start, end):
                w = filter_by_date_range(promo_df, "Date", start, end)
                w = filter_excluded_dates(w, "Date", excluded)
                if w.empty:
                    return pd.DataFrame(columns=["Store ID", "New Customers"])
                w[nc_col] = pd.to_numeric(w[nc_col], errors="coerce").fillna(0)
                agg = w.groupby(promo_sc)[nc_col].sum().reset_index()
                agg.columns = ["Store ID", "New Customers"]
                agg["Store ID"] = agg["Store ID"].astype(str)
                return agg

            nc_pre24 = _nc_window(ly_ps, ly_pe)
            nc_post24 = _nc_window(ly_pos, ly_poe)
            nc_pre25 = _nc_window(ps, pe)
            nc_post25 = _nc_window(pos, poe)
            dd_nc = process_metric(nc_pre24, nc_post24, nc_pre25, nc_post25, "New Customers")
    results["dd_nc"] = dd_nc

    # Rebuild DD summary with NC
    if not dd_nc.empty and not results["dd_sales"].empty:
        dd_t1, dd_t2 = create_summary_tables(results["dd_sales"], results["dd_payouts"],
                                              results["dd_orders"], dd_nc,
                                              results["dd_selected"], include_nc=True)
        results["dd_summary_t1"] = dd_t1
        results["dd_summary_t2"] = dd_t2

    # ── Combined ────────────────────────────────────────────────────────
    comb_t1, comb_t2 = create_combined_summary(
        results["dd_sales"], results["dd_payouts"], results["dd_orders"], dd_nc,
        results["ue_sales"], results["ue_payouts"], results["ue_orders"],
        results.get("dd_selected", []), results.get("ue_selected", []),
    )
    results["combined_t1"] = comb_t1
    results["combined_t2"] = comb_t2

    # ── Marketing Corporate vs TODC ─────────────────────────────────────
    promo_for_mkt = st.session_state.promo_raw[0] if st.session_state.promo_raw else None
    spons_for_mkt = st.session_state.sponsored_raw[0] if st.session_state.sponsored_raw else None
    if promo_for_mkt is not None or spons_for_mkt is not None:
        results["marketing"] = create_corporate_vs_todc(
            promo_for_mkt, spons_for_mkt, pos, poe, excluded
        )
    else:
        results["marketing"] = None

    st.session_state.results = results
    st.session_state.analysis_done = True


def _compute_derived_metrics_inline(df):
    """Recompute Growth%/YoY% after combining stores from two platforms."""
    df = df.copy()
    for c in ["pre_24", "post_24", "pre_25", "post_25"]:
        if c not in df.columns:
            df[c] = 0.0
    df["PrevsPost"] = df["post_25"] - df["pre_25"]
    df["LastYear_Pre_vs_Post"] = df["post_24"] - df["pre_24"]
    df["YoY"] = df["post_25"] - df["post_24"]
    df["Growth%"] = (df["PrevsPost"] / df["pre_25"].replace(0, 1) * 100).replace(
        [float("inf"), -float("inf")], 0).fillna(0).round(1)
    df["YoY%"] = (df["YoY"] / df["post_24"].replace(0, 1) * 100).replace(
        [float("inf"), -float("inf")], 0).fillna(0).round(1)
    return df


def _build_export(r):
    """Build an Excel workbook from analysis results."""
    try:
        from openpyxl import Workbook
        wb = Workbook()

        def _write_df(ws, df, start_row=1, start_col=1):
            if df.empty:
                return
            has_index = df.index.name is not None or not isinstance(df.index, pd.RangeIndex)
            if has_index:
                df = df.reset_index()
            for j, col_name in enumerate(df.columns):
                ws.cell(row=start_row, column=start_col + j, value=col_name)
            for i, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
                for j, val in enumerate(row_data):
                    ws.cell(row=i, column=start_col + j, value=val)

        ws = wb.active
        ws.title = "Summary Tables"
        ws.cell(row=1, column=1, value="Combined Pre vs Post")
        _write_df(ws, r["combined_t1"], start_row=2)
        offset = len(r["combined_t1"]) + 5
        ws.cell(row=offset, column=1, value="Combined YoY")
        _write_df(ws, r["combined_t2"], start_row=offset + 1)

        if not r["dd_summary_t1"].empty:
            ws2 = wb.create_sheet("DD Summary")
            ws2.cell(row=1, column=1, value="DoorDash Pre vs Post")
            _write_df(ws2, r["dd_summary_t1"], start_row=2)
            off = len(r["dd_summary_t1"]) + 5
            ws2.cell(row=off, column=1, value="DoorDash YoY")
            _write_df(ws2, r["dd_summary_t2"], start_row=off + 1)

        if not r["ue_summary_t1"].empty:
            ws3 = wb.create_sheet("UE Summary")
            ws3.cell(row=1, column=1, value="UberEats Pre vs Post")
            _write_df(ws3, r["ue_summary_t1"], start_row=2)
            off = len(r["ue_summary_t1"]) + 5
            ws3.cell(row=off, column=1, value="UberEats YoY")
            _write_df(ws3, r["ue_summary_t2"], start_row=off + 1)

        for label, key in [("DD Sales", "dd_sales"), ("DD Payouts", "dd_payouts"),
                           ("DD Orders", "dd_orders"), ("UE Sales", "ue_sales"),
                           ("UE Payouts", "ue_payouts"), ("UE Orders", "ue_orders")]:
            df = r.get(key, pd.DataFrame())
            if not df.empty:
                ws_s = wb.create_sheet(label)
                _write_df(ws_s, df)

        mkt = r.get("marketing")
        if mkt:
            ws_m = wb.create_sheet("Corporate vs TODC")
            row_num = 1
            for lbl, key in [("Combined", "combined"), ("Promotion", "promotion"), ("Sponsored", "sponsored")]:
                if key in mkt and not mkt[key].empty:
                    ws_m.cell(row=row_num, column=1, value=lbl)
                    _write_df(ws_m, mkt[key], start_row=row_num + 1)
                    row_num += len(mkt[key]) + 4

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None


def _kpi_val(table, metric, col):
    try:
        return table.loc[metric, col]
    except (KeyError, IndexError):
        return 0


def _render_kpi(label, value, fmt="pct"):
    if fmt == "pct":
        display = f"{value:+.1f}%"
    elif fmt == "dollar":
        display = f"${value:,.1f}"
    elif fmt == "int":
        display = f"{int(round(value)):,}"
    else:
        display = str(value)
    if value > 0:
        cls, arrow = "kpi-up", "▲"
    elif value < 0:
        cls, arrow = "kpi-down", "▼"
    else:
        cls, arrow = "kpi-flat", "—"
    return f"""<div class="kpi-card">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{display}</div>
        <div class="kpi-delta {cls}">{arrow} {'growth' if value >= 0 else 'decline'}</div>
    </div>"""


def _display_summary_and_stores(tab, t1, t2, sales_df, payouts_df, orders_df, selected, platform_label):
    with tab:
        st.markdown(f'<div class="section-hdr">{platform_label} — Summary (Pre vs Post)</div>', unsafe_allow_html=True)
        if not t1.empty:
            st.dataframe(t1, use_container_width=True)
        else:
            st.info(f"No {platform_label} data available.")

        st.markdown(f'<div class="section-hdr">{platform_label} — Summary (YoY)</div>', unsafe_allow_html=True)
        if not t2.empty:
            st.dataframe(t2, use_container_width=True)

        if not sales_df.empty:
            with st.expander(f"{platform_label} Store-Level Sales", expanded=False):
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Pre vs Post**")
                    st.dataframe(get_store_table_prepost(sales_df, selected), use_container_width=True, hide_index=True)
                with c2:
                    st.markdown("**YoY**")
                    st.dataframe(get_store_table_yoy(sales_df, selected), use_container_width=True, hide_index=True)

        if not payouts_df.empty:
            with st.expander(f"{platform_label} Store-Level Payouts", expanded=False):
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Pre vs Post**")
                    st.dataframe(get_store_table_prepost(payouts_df, selected), use_container_width=True, hide_index=True)
                with c2:
                    st.markdown("**YoY**")
                    st.dataframe(get_store_table_yoy(payouts_df, selected), use_container_width=True, hide_index=True)

        if not orders_df.empty:
            with st.expander(f"{platform_label} Store-Level Orders", expanded=False):
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Pre vs Post**")
                    st.dataframe(get_store_table_prepost(orders_df, selected), use_container_width=True, hide_index=True)
                with c2:
                    st.markdown("**YoY**")
                    st.dataframe(get_store_table_yoy(orders_df, selected), use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════════════
# Page layout — Header + Stepper
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("## Ralph Analyse")
st.caption("DoorDash + UberEats  ·  Pre vs Post  ·  Year-over-Year  ·  Marketing")
_render_stepper(st.session_state.step)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 — File Upload
# ═══════════════════════════════════════════════════════════════════════════
if st.session_state.step == 1:
    st.markdown('<div class="section-hdr">Upload Data Files</div>', unsafe_allow_html=True)
    st.markdown("Upload up to **4 files**: DD Financial, UE Financial, Marketing Promotion, Marketing Sponsored Listing.")

    uploaded = st.file_uploader(
        "Drop CSV files here", type=["csv"], accept_multiple_files=True,
        key="file_uploader",
    )

    if uploaded:
        if st.session_state.tmp_dir is None:
            st.session_state.tmp_dir = tempfile.mkdtemp(prefix="ralph_")
        tmp = Path(st.session_state.tmp_dir)

        classified = {}
        for f in uploaded:
            ftype = classify_file(f.name)
            dest = tmp / f.name
            dest.write_bytes(f.read())
            classified[ftype] = str(dest)

        st.session_state.files = classified

        summary_rows = []
        for ftype, path in classified.items():
            nice = ftype.replace("_", " ").title()
            summary_rows.append({"Type": nice, "File": Path(path).name})
        st.table(pd.DataFrame(summary_rows))

        has_dd = "dd_financial" in classified
        has_ue = "ue_financial" in classified

        if has_dd or has_ue:
            if st.button("Next →  Configure Dates", type="primary"):
                st.session_state.step = 2
                st.rerun()
        else:
            st.warning("Please upload at least one financial file (DD or UE).")

# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — Date Configuration
# ═══════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 2:
    st.markdown('<div class="section-hdr">Configure Date Ranges</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Pre Period**")
        pre_start = st.date_input("Pre Start", value=datetime(2026, 3, 1), key="pre_start")
        pre_end = st.date_input("Pre End", value=datetime(2026, 4, 12), key="pre_end")
    with col2:
        st.markdown("**Post Period**")
        post_start = st.date_input("Post Start", value=datetime(2026, 4, 13), key="post_start")
        post_end = st.date_input("Post End", value=datetime(2026, 5, 17), key="post_end")

    st.markdown("**Excluded Dates** (comma-separated MM/DD/YYYY)")
    excluded_text = st.text_input("e.g. 04/15/2026,04/20/2026", value="", key="excluded_text")

    ly_pre_s, ly_pre_e = get_last_year_dates(pre_start, pre_end)
    ly_post_s, ly_post_e = get_last_year_dates(post_start, post_end)
    st.info(f"Last Year Pre: {ly_pre_s.strftime('%m/%d/%Y')} – {ly_pre_e.strftime('%m/%d/%Y')}  |  "
            f"Last Year Post: {ly_post_s.strftime('%m/%d/%Y')} – {ly_post_e.strftime('%m/%d/%Y')}")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back"):
            st.session_state.step = 1
            st.rerun()
    with c2:
        if st.button("Next →  Select Filters", type="primary"):
            st.session_state.dates = {
                "pre_start": pre_start, "pre_end": pre_end,
                "post_start": post_start, "post_end": post_end,
                "excluded": [s.strip() for s in excluded_text.split(",") if s.strip()],
            }
            _load_raw_data()
            st.session_state.step = 3
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — Filters (store exclusion)
# ═══════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 3:
    st.markdown('<div class="section-hdr">Select Filters</div>', unsafe_allow_html=True)

    dd_stores = st.session_state.get("dd_stores", [])
    ue_stores = st.session_state.get("ue_stores", [])

    if dd_stores:
        st.markdown(f"**DoorDash Stores** ({len(dd_stores)} found)")
        dd_excluded = st.multiselect("Stores to EXCLUDE from DD analysis",
                                     options=dd_stores, default=[], key="dd_exclude")
        dd_selected = [s for s in dd_stores if s not in dd_excluded]
        st.caption(f"{len(dd_selected)} stores selected")
    else:
        dd_selected = []

    if ue_stores:
        st.markdown(f"**UberEats Stores** ({len(ue_stores)} found)")
        ue_excluded = st.multiselect("Stores to EXCLUDE from UE analysis",
                                     options=ue_stores, default=[], key="ue_exclude")
        ue_selected = [s for s in ue_stores if s not in ue_excluded]
        st.caption(f"{len(ue_selected)} stores selected")
    else:
        ue_selected = []

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back"):
            st.session_state.step = 2
            st.rerun()
    with c2:
        if st.button("Analyse", type="primary"):
            st.session_state.dd_selected = dd_selected
            st.session_state.ue_selected = ue_selected
            _run_analysis()
            st.session_state.step = 4
            st.rerun()

# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 — Dashboard
# ═══════════════════════════════════════════════════════════════════════════
elif st.session_state.step == 4:
    r = st.session_state.results
    if r is None:
        st.warning("No analysis results. Please run the analysis first.")
        if st.button("← Start Over"):
            st.session_state.step = 1
            st.rerun()
        st.stop()

    dates = st.session_state.dates
    comb_t1 = r["combined_t1"]
    comb_t2 = r["combined_t2"]

    # ── KPI Cards ───────────────────────────────────────────────────────
    sales_growth_pp = _kpi_val(comb_t1, "Sales", "Growth%")
    sales_growth_yoy = _kpi_val(comb_t2, "Sales", "YoY%")
    order_growth = _kpi_val(comb_t1, "Orders", "Growth%")
    nc_growth = _kpi_val(comb_t1, "New Customers", "Growth%") if "New Customers" in comb_t1.index else 0
    payout_lift = _kpi_val(comb_t1, "Payouts", "PrevsPost")
    dd_count = len(r.get("dd_selected", []))
    ue_count = len(r.get("ue_selected", []))
    total_stores = dd_count + ue_count
    payout_per_store = payout_lift / total_stores if total_stores > 0 else 0

    cols = st.columns(5)
    kpis = [
        ("Sales Growth (Pre→Post)", sales_growth_pp, "pct"),
        ("Sales Growth (YoY)", sales_growth_yoy, "pct"),
        ("Order Growth", order_growth, "pct"),
        ("New Customer Growth", nc_growth, "pct"),
        ("Payout Lift / Store", payout_per_store, "dollar"),
    ]
    for col, (label, val, fmt) in zip(cols, kpis):
        col.markdown(_render_kpi(label, val, fmt), unsafe_allow_html=True)

    st.markdown(f"""<div style="text-align:center;margin:8px 0;color:#667085;font-size:0.85rem">
        Active Stores — DoorDash: <b>{dd_count}</b> | UberEats: <b>{ue_count}</b> &nbsp;&nbsp;|&nbsp;&nbsp;
        Pre: <b>{dates['pre_start'].strftime('%m/%d/%Y')} – {dates['pre_end'].strftime('%m/%d/%Y')}</b> &nbsp;&nbsp;
        Post: <b>{dates['post_start'].strftime('%m/%d/%Y')} – {dates['post_end'].strftime('%m/%d/%Y')}</b>
    </div>""", unsafe_allow_html=True)

    # ── Tabs ────────────────────────────────────────────────────────────
    tab_combined, tab_dd, tab_ue, tab_mkt = st.tabs([
        "Combined", "DoorDash", "UberEats", "Marketing"
    ])

    # Combined tab
    with tab_combined:
        st.markdown('<div class="section-hdr">Combined — Summary (Pre vs Post)</div>', unsafe_allow_html=True)
        if not comb_t1.empty:
            st.dataframe(comb_t1, use_container_width=True)
        st.markdown('<div class="section-hdr">Combined — Summary (YoY)</div>', unsafe_allow_html=True)
        if not comb_t2.empty:
            st.dataframe(comb_t2, use_container_width=True)

        all_sales = pd.concat([
            r["dd_sales"][r["dd_sales"]["Store ID"].isin(r.get("dd_selected", []))] if not r["dd_sales"].empty else pd.DataFrame(),
            r["ue_sales"][r["ue_sales"]["Store ID"].isin(r.get("ue_selected", []))] if not r["ue_sales"].empty else pd.DataFrame(),
        ], ignore_index=True)
        if not all_sales.empty:
            grouped = all_sales.groupby("Store ID", as_index=False).sum(numeric_only=True)
            grouped = _compute_derived_metrics_inline(grouped)
            with st.expander("Combined Store-Level Sales", expanded=False):
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("**Pre vs Post**")
                    st.dataframe(get_store_table_prepost(grouped, grouped["Store ID"].tolist()),
                                 use_container_width=True, hide_index=True)
                with c2:
                    st.markdown("**YoY**")
                    st.dataframe(get_store_table_yoy(grouped, grouped["Store ID"].tolist()),
                                 use_container_width=True, hide_index=True)

    # DD tab
    _display_summary_and_stores(
        tab_dd, r["dd_summary_t1"], r["dd_summary_t2"],
        r["dd_sales"], r["dd_payouts"], r["dd_orders"],
        r.get("dd_selected", []), "DoorDash",
    )

    # UE tab
    _display_summary_and_stores(
        tab_ue, r["ue_summary_t1"], r["ue_summary_t2"],
        r["ue_sales"], r["ue_payouts"], r["ue_orders"],
        r.get("ue_selected", []), "UberEats",
    )

    # Marketing tab
    with tab_mkt:
        mkt = r.get("marketing")
        if mkt is None:
            st.info("No marketing files uploaded. Upload MARKETING_PROMOTION and/or MARKETING_SPONSORED_LISTING CSVs.")
        else:
            st.markdown('<div class="section-hdr">Corporate vs TODC — Post Period Only</div>', unsafe_allow_html=True)
            for label, key in [("Combined", "combined"), ("Promotion", "promotion"), ("Sponsored Listing", "sponsored")]:
                if key in mkt and not mkt[key].empty:
                    st.markdown(f"**{label}**")
                    st.dataframe(mkt[key], use_container_width=True)

    # ── Export + Restart ────────────────────────────────────────────────
    st.markdown("---")
    col_export, col_restart = st.columns([3, 1])
    with col_export:
        xlsx = _build_export(r)
        if xlsx:
            st.download_button("Download Excel Report", data=xlsx,
                               file_name=f"ralph_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary")
    with col_restart:
        if st.button("← New Analysis"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
