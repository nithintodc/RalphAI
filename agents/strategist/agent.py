"""
Strategist agent: sequential browser-use login per operator from Airtable,
download financial + marketing reports for the last 3 calendar months, run them
through the full Reporting analysis pipeline, and produce
a campaigns Excel with two sheets — Offers Campaigns and Ads Campaigns (store-wise,
slot-tagged). Output saved to data/Strategist/<operatorName>/<timestamp>/.
"""

from __future__ import annotations

import json
import logging
import os
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Literal

from shared.config.settings import data_root, marketingreco_reporting_root
from shared.utils.account_directory import load_account_operators

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_ROOT = PROJECT_ROOT / "data" / "Strategist"
REPORTING_ROOT = marketingreco_reporting_root()

# Canonical Day-Slot → tag grid (slots.csv equivalent, built in-code so no external
# file is required). tag = slot_index * 7 + day_index + 1, giving tags 1..42.
#   Overnight Mon = 1, Breakfast Mon = 8, ... Late night Sun = 42.
_GRID_SLOTS = ["Overnight", "Breakfast", "Lunch", "Afternoon", "Dinner", "Late night"]
_GRID_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
# Below which average-order-value a slot gets an Ads campaign instead of an Offer.
ADS_AOV_THRESHOLD = 10.0
# Fixed minimum bid for every Ads campaign (per spec).
ADS_MIN_BID = 3


def _slot_tag(day_full: str, slot_name: str) -> int | None:
    """Map a (full day name, slot name) pair to its grid tag (1..42), or None."""
    try:
        di = _GRID_DAYS.index(str(day_full).strip().title())
        si = _GRID_SLOTS.index(str(slot_name).strip())
    except ValueError:
        return None
    return si * 7 + di + 1


@dataclass(frozen=True)
class StrategistOperator:
    operator_id: str
    business_name: str
    email: str
    password: str


def _date_range_90_days() -> tuple[str, str]:
    """Return (start_date, end_date) as MM/DD/YYYY for the last 3 complete calendar months."""
    today = datetime.now().date()
    end = today.replace(day=1) - timedelta(days=1)
    month = today.month - 3
    year = today.year
    if month <= 0:
        month += 12
        year -= 1
    start = today.replace(year=year, month=month, day=1)
    return start.strftime("%m/%d/%Y"), end.strftime("%m/%d/%Y")


def _safe_email(email: str) -> str:
    safe = (email or "operator").strip()
    for ch in ("@", ".", " ", "/", "\\"):
        safe = safe.replace(ch, "_")
    return safe[:80] if len(safe) > 80 else safe


def _safe_dirname(name: str) -> str:
    """Sanitize an operator/business name for use as a directory name (keeps spaces)."""
    safe = (name or "operator").strip()
    for ch in ('/', '\\', ':', '*', '?', '"', '<', '>', '|'):
        safe = safe.replace(ch, "-")
    safe = safe.strip(". ")
    return (safe[:100] if len(safe) > 100 else safe) or "operator"


def _resolve_selected_operators(selected_operator_ids: list[str]) -> list[StrategistOperator]:
    rows, warning = load_account_operators()
    if not rows:
        raise RuntimeError(
            warning or "No operators in Airtable account directory (check AIRTABLE_PAT)."
        )
    by_operator_id = {str(r.get("operator_id", "")).strip(): r for r in rows}
    out: list[StrategistOperator] = []
    for oid in selected_operator_ids:
        key = (oid or "").strip()
        if not key:
            continue
        row = by_operator_id.get(key)
        if not row:
            continue
        email = str(row.get("doordash_email", "")).strip()
        password = str(row.get("doordash_password", "")).strip()
        if not email:
            continue
        try:
            from shared.doordash_portal_tasks import resolve_doordash_credentials

            email, password = resolve_doordash_credentials(
                email,
                password or None,
                operator_name=str(row.get("business_name", key)).strip() or key,
            )
        except ValueError:
            if not password:
                continue
        out.append(
            StrategistOperator(
                operator_id=key,
                business_name=str(row.get("business_name", key)).strip() or key,
                email=email,
                password=password,
            )
        )
    return out


def _subprocess_script() -> str:
    """Browser download + full Reporting analysis pipeline in one subprocess.

    Logs go to stderr (streamed to terminal). Only the final STRATEGIST_RESULT
    line goes to stdout (captured by parent process).
    """
    return """
import asyncio
import json
import logging
import os
import sys
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

# ── Logging setup: all logs → stderr so parent can stream them to terminal ──
logging.basicConfig(
    level=logging.INFO,
    format="[STRATEGIST %(asctime)s] %(levelname)s %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stderr,
)
log = logging.getLogger("strategist.subprocess")


def _log(msg, *args):
    log.info(msg, *args)
    print(f"[STRATEGIST] {msg % args if args else msg}", file=sys.stderr, flush=True)


# ── Download helpers (copied from doordash_agent to keep browser open until files land) ──

def _peek_zip_type(path: Path) -> str:
    try:
        with zipfile.ZipFile(path, "r") as z:
            names_upper = " ".join(z.namelist()).upper()
        if "FINANCIAL_DETAILED" in names_upper or ("FINANCIAL" in names_upper and "MARKETING" not in names_upper):
            return "financial"
        if "MARKETING_PROMOTION" in names_upper or "MARKETING_SPONSORED" in names_upper or "MARKETING" in names_upper:
            return "marketing"
    except Exception:
        pass
    return ""


def _discover_downloads(download_dir: Path) -> Tuple[Optional[Path], Optional[Path]]:
    if not download_dir.is_dir():
        return (None, None)
    all_files = []
    for ext in ("*.csv", "*.zip", "*.xlsx"):
        for f in download_dir.glob(ext):
            if f.is_file() and not f.name.startswith("."):
                all_files.append((f.stat().st_mtime, f))
    all_files.sort(key=lambda x: x[0], reverse=True)

    financial_path = None
    marketing_path = None
    unmatched = []
    for _mtime, path in all_files:
        name_lower = path.name.lower()
        if "financial" in name_lower:
            if financial_path is None:
                financial_path = path
        elif "marketing" in name_lower:
            if marketing_path is None:
                marketing_path = path
        else:
            unmatched.append(path)
        if financial_path and marketing_path:
            break

    if (financial_path is None or marketing_path is None) and unmatched:
        for path in unmatched:
            if path.suffix.lower() == ".zip":
                kind = _peek_zip_type(path)
                if kind == "financial" and financial_path is None:
                    financial_path = path
                elif kind == "marketing" and marketing_path is None:
                    marketing_path = path
            if financial_path and marketing_path:
                break

    if financial_path is None and all_files:
        for _mtime, candidate in all_files:
            if candidate != marketing_path:
                financial_path = candidate
                break

    return (marketing_path, financial_path)


def _wait_for_downloads(download_dir: Path, max_wait: int = 120) -> Tuple[Optional[Path], Optional[Path]]:
    \"\"\"Poll download_dir until both reports land or timeout.\"\"\"
    _log("Waiting for downloads in %s (max %ds)...", download_dir, max_wait)
    start = time.time()
    marketing_path = None
    financial_path = None
    while time.time() - start < max_wait:
        marketing_path, financial_path = _discover_downloads(download_dir)
        if marketing_path and financial_path:
            _log("Both reports found: financial=%s, marketing=%s", financial_path.name, marketing_path.name)
            return (marketing_path, financial_path)
        found = []
        if financial_path:
            found.append(f"financial={financial_path.name}")
        if marketing_path:
            found.append(f"marketing={marketing_path.name}")
        _log("  ... %ds elapsed, found so far: %s", int(time.time() - start), ", ".join(found) or "none")
        time.sleep(5)
    _log("Download wait timed out after %ds. financial=%s, marketing=%s",
         max_wait, financial_path, marketing_path)
    return (marketing_path, financial_path)


# ── Main pipeline ──

async def _main():
    download_dir = Path(os.environ["STRATEGIST_DOWNLOAD_DIR"])
    download_dir.mkdir(parents=True, exist_ok=True)
    start_date = os.environ["STRATEGIST_START_DATE"]
    end_date = os.environ["STRATEGIST_END_DATE"]
    email = os.environ["DOORDASH_EMAIL"]
    password = os.environ["DOORDASH_PASSWORD"]

    _log("=" * 60)
    _log("PHASE 1: BROWSER LOGIN + REPORT DOWNLOAD")
    _log("  Email: %s", email)
    _log("  Date range: %s → %s", start_date, end_date)
    _log("  Download dir: %s", download_dir)
    _log("=" * 60)

    from browser_use import Agent

    # --- Browser + LLM setup (same as doordash_agent) ---
    from agents.doordash_agent import (
        _get_llm,
        _get_browser,
        _kill_browser,
        get_task_description_reports_only,
        _get_retry_download_task,
        AGENT_REPORTS_TIMEOUT,
    )

    llm = _get_llm()
    browser = _get_browser(download_dir, keep_alive=True)
    phase1_start = time.time()

    try:
        task = get_task_description_reports_only(
            email=email, password=password,
            start_date=start_date, end_date=end_date,
        )
        _log("Starting browser-use agent for reports...")
        agent = Agent(task=task, llm=llm, browser=browser)
        history = await asyncio.wait_for(agent.run(), timeout=AGENT_REPORTS_TIMEOUT)
        if history and history.final_result:
            _log("Agent result: %s", history.final_result)

        _log("Agent finished. Checking for downloaded files...")
        marketing_path, financial_path = _wait_for_downloads(download_dir, max_wait=60)

        # Retry missing reports (browser still open)
        if not marketing_path or not financial_path:
            missing = []
            if not financial_path:
                missing.append("Financial")
            if not marketing_path:
                missing.append("Marketing")
            _log("Missing reports: %s — retrying download with browser still open", missing)
            await asyncio.sleep(10)
            retry_task = _get_retry_download_task(missing)
            retry_agent = Agent(task=retry_task, llm=llm, browser=browser)
            await asyncio.wait_for(retry_agent.run(), timeout=300)
            marketing_path, financial_path = _wait_for_downloads(download_dir, max_wait=60)

        phase1_elapsed = time.time() - phase1_start
        _log("PHASE 1 COMPLETE (%.0fs). financial=%s, marketing=%s",
             phase1_elapsed, financial_path, marketing_path)

    except asyncio.TimeoutError:
        _log("PHASE 1 TIMED OUT after %ds", AGENT_REPORTS_TIMEOUT)
        marketing_path, financial_path = _discover_downloads(download_dir)
        _log("Files found after timeout: financial=%s, marketing=%s", financial_path, marketing_path)
    except Exception as exc:
        _log("PHASE 1 FAILED: %s", exc)
        marketing_path, financial_path = _discover_downloads(download_dir)
    finally:
        _log("Closing browser...")
        await _kill_browser(browser)
        _log("Browser closed.")

    # ── PHASE 2: Reporting analysis ──
    from agents.marketing_agent import run as marketing_run
    from agents.analysis_agent import run as analysis_run
    from agents.combined_report_agent import run as combined_run, append_campaign_mappings_to_workbook
    from agents.campaign_params import get_campaign_mappings_for_combined

    _log("=" * 60)
    _log("PHASE 2: ANALYSIS PIPELINE")
    _log("=" * 60)

    marketing_sheets = None
    if marketing_path:
        _log("Running marketing analysis on %s...", marketing_path)
        marketing_sheets = marketing_run(
            Path(marketing_path),
            output_dir=download_dir,
            post_start_date=start_date,
            post_end_date=end_date,
            write_file=False,
        )
        _log("Marketing analysis: %d sheets", len(marketing_sheets) if marketing_sheets else 0)
    else:
        _log("No marketing report found — skipping marketing analysis")

    financial_sheets = None
    if financial_path:
        _log("Running financial analysis on %s...", financial_path)
        financial_sheets = analysis_run(
            Path(financial_path),
            output_dir=download_dir,
            report_start_date=start_date,
            report_end_date=end_date,
            write_file=False,
        )
        _log("Financial analysis: %d sheets", len(financial_sheets) if financial_sheets else 0)
    else:
        _log("No financial report found — skipping financial analysis")

    fc = download_dir / "financial_detailed_report.csv"
    if not fc.is_file():
        for p in sorted(download_dir.glob("*FINANCIAL*.csv")):
            fc = p
            break
    _log("Financial CSV for ads plan: %s (exists=%s)", fc, fc.is_file())

    # ── PHASE 3: Combined analysis + campaign mappings ──
    _log("=" * 60)
    _log("PHASE 3: COMBINED ANALYSIS + CAMPAIGN MAPPINGS")
    _log("=" * 60)

    combined = combined_run(
        financial_sheets=financial_sheets,
        marketing_sheets=marketing_sheets,
        output_dir=download_dir,
    )
    _log("Combined analysis: %s", combined)

    if combined:
        slots_csv = Path("slots.csv")
        mappings = get_campaign_mappings_for_combined(Path(combined), slots_csv)
        _log("Campaign mappings derived: %d", len(mappings) if mappings else 0)
        if mappings:
            append_campaign_mappings_to_workbook(Path(combined), mappings)
            _log("Campaign mappings appended to %s", combined)

    payload = {
        "marketing_path": str(marketing_path) if marketing_path else "",
        "financial_path": str(financial_path) if financial_path else "",
        "financial_csv": str(fc) if fc.is_file() else "",
        "combined_path": str(combined) if combined else "",
    }
    _log("DONE. Result: %s", json.dumps(payload))
    print("STRATEGIST_RESULT=" + json.dumps(payload))

asyncio.run(_main())
"""


def _parse_money(val: Any) -> float | None:
    """Parse a possibly dollar-formatted value (e.g. '$12.34', '1,234.5') to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            import math as _m

            return None if (isinstance(val, float) and _m.isnan(val)) else float(val)
        except Exception:
            return None
    s = str(val).strip().replace("$", "").replace(",", "")
    if s in ("", "nan", "None"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _read_day_slots_per_store(combined_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read every ``Day-Slot - {store_id}`` sheet from the combined workbook.

    Returns ``{store_id: [{day, slot, aov, min_subtotal, sales, orders}, ...]}``.
    AOV is dollar-formatted in the sheet, so it is parsed back to float.
    """
    import re

    import pandas as pd

    pattern = re.compile(r"Day-Slot\s*-\s*(.+)", re.IGNORECASE)
    out: dict[str, list[dict[str, Any]]] = {}
    try:
        xl = pd.ExcelFile(combined_path)
    except Exception as e:
        logger.warning("strategist: could not open combined workbook %s: %s", combined_path, e)
        return out

    for sheet_name in xl.sheet_names:
        m = pattern.search(sheet_name)
        if not m:
            continue
        store_id = m.group(1).strip()
        if not store_id:
            continue
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name, header=2)
        except Exception:
            continue
        df.columns = df.columns.astype(str).str.strip()
        if any(c not in df.columns for c in ("Day", "Slot", "Min.Subtotal", "AOV")):
            continue
        rows = out.setdefault(store_id, [])
        for _, row in df.dropna(subset=["Day", "Slot"]).iterrows():
            aov = _parse_money(row.get("AOV"))
            sales = _parse_money(row.get("Sales")) or 0.0
            try:
                orders = int(float(row.get("Orders"))) if pd.notna(row.get("Orders")) else 0
            except (ValueError, TypeError):
                orders = 0
            min_sub = _parse_money(row.get("Min.Subtotal")) or 0.0
            rows.append({
                "day": str(row["Day"]).strip(),
                "slot": str(row["Slot"]).strip(),
                "aov": aov,
                "min_subtotal": int(round(min_sub)),
                "sales": sales,
                "orders": orders,
            })
    return out


def _read_store_names(combined_path: Path, financial_csv: Path | None) -> dict[str, str]:
    """Build ``store_id → store name`` from the combined workbook's Store-wise sheet,
    falling back to the financial CSV. Self-contained (runs in the parent process)."""
    import pandas as pd

    names: dict[str, str] = {}

    # Primary: Store-wise sheet (header at row 3 / header=2).
    try:
        xl = pd.ExcelFile(combined_path)
        for sheet_name in xl.sheet_names:
            norm = sheet_name.lower().replace("-", "").replace(" ", "")
            if norm not in ("storewise", "financialstorewise"):
                continue
            df = pd.read_excel(xl, sheet_name=sheet_name, header=2)
            df.columns = df.columns.astype(str).str.strip()
            id_col = next((c for c in df.columns if c.lower() in ("merchant store id", "store id")), None)
            name_col = next((c for c in df.columns if c.lower() == "store name"), None)
            if id_col and name_col:
                for _, row in df.dropna(subset=[id_col]).iterrows():
                    sid = str(row[id_col]).strip()
                    sname = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                    if sid and sname:
                        names[sid] = sname
                break
    except Exception as e:
        logger.debug("strategist: store-wise name lookup failed: %s", e)

    # Fallback: financial CSV (Merchant store ID + Store name columns).
    if financial_csv and Path(financial_csv).is_file():
        try:
            fdf = pd.read_csv(financial_csv)
            id_col = next((c for c in ["Merchant store ID", "Merchant Store ID", "Store ID"] if c in fdf.columns), None)
            name_col = next((c for c in ["Store name", "Store Name", "Merchant store name"] if c in fdf.columns), None)
            if id_col and name_col:
                for _, row in fdf.dropna(subset=[id_col]).iterrows():
                    sid = str(row[id_col]).strip()
                    if sid and sid not in names:
                        sname = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                        if sname:
                            names[sid] = sname
        except Exception as e:
            logger.debug("strategist: financial CSV name lookup failed: %s", e)

    return names


def _build_campaign_excel(
    operator_dir: Path,
    combined_path: Path,
    financial_csv: Path | None,
) -> Path | None:
    """Write the spec campaign workbook (exactly two sheets) to operator_dir.

    Reads the per-store ``Day-Slot - {store_id}`` sheets from the combined analysis
    and routes each populated slot:
      * AOV >= 10  → Offers Campaigns (grouped per store + Min.Subtotal, slot tags)
      * AOV  < 10  → Ads Campaigns    (TODC-ADS-<storeID>, fixed $3 min bid)
      * Orders == 0 and Sales == 0 → blank (slot has no row, so naturally skipped)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        logger.warning("openpyxl not installed — skipping campaign Excel")
        return None

    per_store = _read_day_slots_per_store(combined_path)
    store_names = _read_store_names(combined_path, financial_csv)

    # offers[store_id][min_subtotal] = sorted set of tags (AOV >= threshold)
    offers: dict[str, dict[int, set[int]]] = {}
    # ads[store_id] = set of tags (AOV < threshold)
    ads: dict[str, set[int]] = {}

    for store_id, rows in per_store.items():
        for r in rows:
            # Blank slot: no orders and no sales → no campaign.
            if r["orders"] == 0 and (r["sales"] or 0) == 0:
                continue
            tag = _slot_tag(r["day"], r["slot"])
            if tag is None:
                continue
            aov = r["aov"]
            if aov is not None and aov < ADS_AOV_THRESHOLD:
                ads.setdefault(store_id, set()).add(tag)
            else:
                min_sub = r["min_subtotal"]
                if min_sub <= 0:
                    continue
                offers.setdefault(store_id, {}).setdefault(min_sub, set()).add(tag)

    out_path = operator_dir / "campaigns.xlsx"
    wb = openpyxl.Workbook()

    def _tags_str(tags) -> str:
        return ",".join(str(t) for t in sorted(tags))

    # Sheet 1: Offers Campaigns
    ws = wb.active
    ws.title = "Offers Campaigns"
    offer_headers = ["Store ID", "Store Name", "Minimum Subtotal", "Slot Tags", "Campaign Name", "Status"]
    for idx, h in enumerate(offer_headers, start=1):
        ws.cell(row=1, column=idx, value=h).font = Font(bold=True)
    r = 2
    for store_id in sorted(offers):
        for min_sub in sorted(offers[store_id]):
            tags = offers[store_id][min_sub]
            ws.cell(row=r, column=1, value=store_id)
            ws.cell(row=r, column=2, value=store_names.get(store_id, ""))
            ws.cell(row=r, column=3, value=min_sub)
            ws.cell(row=r, column=4, value=_tags_str(tags))
            ws.cell(row=r, column=5, value=f"TODC-{store_id}-${min_sub}")
            ws.cell(row=r, column=6, value="Pending")
            r += 1
    offers_count = r - 2

    # Sheet 2: Ads Campaigns
    wsa = wb.create_sheet("Ads Campaigns")
    ads_headers = ["Store ID", "Store Name", "Minimum Bid", "Slot Tags", "Campaign Name", "Status"]
    for idx, h in enumerate(ads_headers, start=1):
        wsa.cell(row=1, column=idx, value=h).font = Font(bold=True)
    r = 2
    for store_id in sorted(ads):
        tags = ads[store_id]
        if not tags:
            continue
        wsa.cell(row=r, column=1, value=store_id)
        wsa.cell(row=r, column=2, value=store_names.get(store_id, ""))
        wsa.cell(row=r, column=3, value=ADS_MIN_BID)
        wsa.cell(row=r, column=4, value=_tags_str(tags))
        wsa.cell(row=r, column=5, value=f"TODC-ADS-{store_id}")
        wsa.cell(row=r, column=6, value="Pending")
        r += 1
    ads_count = r - 2

    wb.save(out_path)
    logger.info(
        "Wrote campaigns.xlsx (%d offer rows, %d ads rows) to %s",
        offers_count, ads_count, out_path,
    )
    return out_path


def _extract_failure_reason(output: str | None, *, max_chars: int = 500) -> str:
    """Pull a human-readable failure reason out of subprocess stderr/stdout.

    Prefers the last Python traceback's final line (e.g. ``TimeoutError: ...``),
    then any ``[STRATEGIST]`` error marker, then falls back to the last non-empty
    line. Returns an empty string when nothing useful is found.
    """
    if not output:
        return ""
    lines = [ln.rstrip() for ln in output.splitlines() if ln.strip()]
    if not lines:
        return ""

    # 1) Final line of the last traceback is usually the exception type + message.
    for idx in range(len(lines) - 1, -1, -1):
        if lines[idx].startswith("Traceback (most recent call last)"):
            exc_line = lines[-1]
            # Walk forward from the traceback to the first exception-looking line.
            for ln in lines[idx + 1:]:
                if ln and not ln.startswith((" ", "\t", "File \"")):
                    exc_line = ln
                    break
            return exc_line[:max_chars]

    # 2) Explicit strategist error markers.
    for ln in reversed(lines):
        low = ln.lower()
        if "[strategist]" in low and ("error" in low or "fail" in low or "timeout" in low):
            return ln[:max_chars]

    # 3) Any line that looks like an error/exception.
    for ln in reversed(lines):
        if any(tok in ln for tok in ("Error", "Exception", "Timeout", "Traceback", "Failed")):
            return ln[:max_chars]

    # 4) Last resort: the final line of output.
    return lines[-1][:max_chars]


def _run_for_operator(
    *,
    operator: StrategistOperator,
    start_date: str,
    end_date: str,
    run_timestamp: str,
) -> dict[str, Any]:
    """Download reports, run analysis pipeline, and build campaign Excel for a single operator.

    Output is stored at ``data/Strategist/<operatorName>/<timestamp>/``.
    """
    operator_dir = OUTPUT_ROOT / _safe_dirname(operator.business_name) / run_timestamp
    operator_dir.mkdir(parents=True, exist_ok=True)

    download_dir = operator_dir / "downloads"
    download_dir.mkdir(parents=True, exist_ok=True)

    from shared.subprocess_env import reporting_subprocess_env

    env = reporting_subprocess_env(REPORTING_ROOT)
    env["DOORDASH_EMAIL"] = operator.email
    env["DOORDASH_PASSWORD"] = operator.password
    env["STRATEGIST_START_DATE"] = start_date
    env["STRATEGIST_END_DATE"] = end_date
    env["STRATEGIST_DOWNLOAD_DIR"] = str(download_dir)

    logger.info("Starting browser subprocess for %s ...", operator.email)

    proc = subprocess.run(
        [sys.executable, "-u", "-c", _subprocess_script()],
        cwd=str(REPORTING_ROOT),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=1200,
    )

    # Subprocess logs go to stderr; echo them to our stderr so the terminal still
    # shows live-ish progress, and keep them around to surface the real failure reason.
    if proc.stderr:
        print(proc.stderr, file=sys.stderr, flush=True)

    if proc.returncode != 0:
        detail = _extract_failure_reason(proc.stderr) or _extract_failure_reason(proc.stdout)
        msg = f"Browser subprocess failed (rc={proc.returncode}) for {operator.email}"
        if detail:
            msg = f"{msg}: {detail}"
        raise RuntimeError(msg)

    result_line = ""
    for line in reversed((proc.stdout or "").splitlines()):
        if line.startswith("STRATEGIST_RESULT="):
            result_line = line.split("=", 1)[1].strip()
            break
    if not result_line:
        raise RuntimeError(f"No STRATEGIST_RESULT returned for {operator.email}")

    parsed = json.loads(result_line)
    financial_path = parsed.get("financial_path", "").strip()
    marketing_path = parsed.get("marketing_path", "").strip()
    financial_csv = parsed.get("financial_csv", "").strip()
    combined_path = parsed.get("combined_path", "").strip()

    campaigns_xlsx = None
    if combined_path:
        fc = Path(financial_csv) if financial_csv else None
        try:
            campaigns_xlsx = _build_campaign_excel(operator_dir, Path(combined_path), fc)
        except Exception as e:
            logger.warning("Campaign Excel generation failed for %s: %s", operator.email, e)

    return {
        "operator_id": operator.operator_id,
        "business_name": operator.business_name,
        "email": operator.email,
        "status": "success",
        "output_dir": str(operator_dir),
        "financial_path": financial_path or None,
        "marketing_path": marketing_path or None,
        "combined_analysis": combined_path or None,
        "campaigns_xlsx": str(campaigns_xlsx) if campaigns_xlsx else None,
    }


StrategistMode = Literal["auto", "manual"]


def _marketing_plan_path(operator_id: str) -> Path:
    return data_root() / "operators" / operator_id / "reports" / "marketing_plan.json"


def _save_marketing_plan(operator_id: str, plan) -> dict[str, Any]:
    path = _marketing_plan_path(operator_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(plan.model_dump_json(indent=2), encoding="utf-8")
    return json.loads(plan.model_dump_json())


def run_manual_from_register(
    operator_id: str,
    *,
    register_report_path: str | Path,
    business_name: str | None = None,
) -> dict[str, Any]:
    """Build a marketing plan from an uploaded DoorDash register file (no browser login)."""
    from .campaigns_excel import write_campaigns_excel
    from .plan_builder import build_marketing_plan
    from .register_reco import build_recommendations_from_register

    oid = (operator_id or "").strip()
    if not oid:
        raise ValueError("operator_id is required")
    register_path = Path(register_report_path)
    if not register_path.is_file():
        raise FileNotFoundError(f"register file not found: {register_path}")

    slots_csv = REPORTING_ROOT / "slots.csv"
    built = build_recommendations_from_register(register_path, slots_csv=slots_csv)
    mappings = built.get("campaign_mappings") or []
    ads_plan = built.get("ads_plan")
    plan = build_marketing_plan(oid, mappings=mappings, ads_plan=ads_plan)

    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dirname = _safe_dirname(business_name or oid)
    operator_dir = OUTPUT_ROOT / dirname / run_timestamp
    operator_dir.mkdir(parents=True, exist_ok=True)

    register_copy = operator_dir / register_path.name
    register_copy.write_bytes(register_path.read_bytes())

    plan_dict = _save_marketing_plan(oid, plan)
    out = {
        **plan_dict,
        "operator_id": oid,
        "business_name": business_name or oid,
        "status": "success",
        "mode": "manual",
        "input_type": "register",
        "output_dir": str(operator_dir),
        "register_path": str(register_copy),
        "campaign_mappings": mappings,
        "slot_recommendations": built.get("slot_recommendations") or [],
        "ads_plan": ads_plan,
    }

    campaigns_xlsx = operator_dir / "marketing_plan.xlsx"
    write_campaigns_excel(campaigns_xlsx, out)
    out["campaigns_xlsx"] = str(campaigns_xlsx)
    (operator_dir / "result.json").write_text(json.dumps(out, indent=2), encoding="utf-8")
    return out


def run(
    *,
    mode: StrategistMode = "auto",
    operator_ids: list[str] | None = None,
    operator_id: str | None = None,
    register_report_path: str | None = None,
    business_name: str | None = None,
) -> dict[str, Any]:
    """
    Auto: iterate selected operators, download reports, run analysis, generate campaigns Excel.
    Manual: single operator + DD register upload → marketing plan (no browser login).
    """
    if mode == "manual":
        if not operator_id or not register_report_path:
            raise ValueError("manual mode requires operator_id and register_report_path")
        result = run_manual_from_register(
            operator_id,
            register_report_path=register_report_path,
            business_name=business_name,
        )
        return {
            "status": "success",
            "mode": "manual",
            "output_root": str(OUTPUT_ROOT),
            "results": [result],
        }

    if not operator_ids:
        raise ValueError("auto mode requires operator_ids")
    raw_ids = [(oid or "").strip() for oid in operator_ids if (oid or "").strip()]
    operators = _resolve_selected_operators(raw_ids)
    operator_by_id = {op.operator_id: op for op in operators}
    start_date, end_date = _date_range_90_days()
    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not (REPORTING_ROOT / "main.py").is_file():
        raise FileNotFoundError(f"Reporting workflow not found at: {REPORTING_ROOT}")

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, Any]] = []
    for oid in raw_ids:
        op = operator_by_id.get(oid)
        if op is None:
            results.append({
                "operator_id": oid,
                "business_name": oid,
                "email": "",
                "status": "skipped",
                "error": "Missing DoorDash credentials or operator not found.",
            })
            continue
        try:
            results.append(_run_for_operator(
                operator=op,
                start_date=start_date,
                end_date=end_date,
                run_timestamp=run_timestamp,
            ))
        except Exception as exc:
            logger.error("Operator %s failed: %s", op.email, exc, exc_info=True)
            results.append({
                "operator_id": op.operator_id,
                "business_name": op.business_name,
                "email": op.email,
                "status": "failed",
                "error": str(exc),
            })

    return {
        "status": "success",
        "mode": "auto",
        "date_range": {"start": start_date, "end": end_date},
        "output_root": str(OUTPUT_ROOT),
        "results": results,
    }
