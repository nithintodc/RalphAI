"""Write Strategist marketing plan workbook (Offers, Ads, Register slots)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from shared.campaign_planning.ralph_ads_excel import ralph_ads_upload_rows


def write_campaigns_excel(path: Path, result: dict[str, Any]) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offers"
    headers = [
        "Store ID",
        "DoorDash Store ID",
        "Store Name",
        "Minimum Subtotal",
        "Slot Tags",
        "Campaign Name",
        "Status",
    ]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = Font(bold=True)
    mappings = result.get("campaign_mappings") or []
    for r, m in enumerate(mappings, start=2):
        tags = m.get("slot_tags", [])
        tags_str = ",".join(str(t) for t in tags) if isinstance(tags, list) else str(tags or "")
        ws.cell(row=r, column=1, value=m.get("store_id", ""))
        ws.cell(row=r, column=2, value=m.get("doordash_store_id", ""))
        ws.cell(row=r, column=3, value=m.get("store_name", ""))
        ws.cell(row=r, column=4, value=m.get("min_subtotal", 0))
        ws.cell(row=r, column=5, value=tags_str)
        ws.cell(row=r, column=6, value=m.get("campaign_name", ""))
        ws.cell(row=r, column=7, value=m.get("status", "Pending"))

    ads_plan = result.get("ads_plan") or {}
    slot_table = ads_plan.get("slot_table") or []
    if slot_table:
        wss = wb.create_sheet("Ads slots")
        sh = [
            "Merchant store ID",
            "Store name",
            "Slot",
            "Orders",
            "Sales",
            "Net total",
            "Profitability %",
            "Ad placement",
        ]
        for idx, h in enumerate(sh, start=1):
            cell = wss.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(slot_table, start=2):
            wss.cell(row=r, column=1, value=row.get("store_id"))
            wss.cell(row=r, column=2, value=row.get("store_name"))
            wss.cell(row=r, column=3, value=row.get("slot"))
            wss.cell(row=r, column=4, value=row.get("orders"))
            wss.cell(row=r, column=5, value=row.get("sales"))
            wss.cell(row=r, column=6, value=row.get("net_total"))
            wss.cell(row=r, column=7, value=row.get("profitability_pct"))
            wss.cell(row=r, column=8, value=row.get("ad_placement"))

    ralph_ads = ralph_ads_upload_rows(ads_plan)
    if ralph_ads:
        wsr = wb.create_sheet("Ads")
        rh = ["Merchant store ID", "Slots", "Bid strategy", "Campaign Name"]
        for idx, h in enumerate(rh, start=1):
            cell = wsr.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(ralph_ads, start=2):
            wsr.cell(row=r, column=1, value=row["store_id"])
            wsr.cell(row=r, column=2, value=row["slots"])
            wsr.cell(row=r, column=3, value=row["bid_strategy"])
            wsr.cell(row=r, column=4, value=row["campaign_name"])

    slot_recs = result.get("slot_recommendations") or []
    if slot_recs:
        wsrg = wb.create_sheet("Register slots")
        rgh = [
            "Store ID",
            "Day",
            "Daypart",
            "Orders",
            "Sales",
            "Payouts",
            "AOV",
            "Profitability %",
            "Action",
            "Min subtotal",
            "Slot tag",
            "Campaign name",
            "Rationale",
        ]
        for idx, h in enumerate(rgh, start=1):
            cell = wsrg.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(slot_recs, start=2):
            wsrg.cell(row=r, column=1, value=row.get("store_id"))
            wsrg.cell(row=r, column=2, value=row.get("day"))
            wsrg.cell(row=r, column=3, value=row.get("daypart"))
            wsrg.cell(row=r, column=4, value=row.get("orders"))
            wsrg.cell(row=r, column=5, value=row.get("sales"))
            wsrg.cell(row=r, column=6, value=row.get("payouts"))
            wsrg.cell(row=r, column=7, value=row.get("aov"))
            wsrg.cell(row=r, column=8, value=row.get("profitability_pct"))
            wsrg.cell(row=r, column=9, value=row.get("action"))
            wsrg.cell(row=r, column=10, value=row.get("min_subtotal"))
            wsrg.cell(row=r, column=11, value=row.get("slot_tag"))
            wsrg.cell(row=r, column=12, value=row.get("campaign_name"))
            wsrg.cell(row=r, column=13, value=row.get("rationale"))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
