"""
FastAPI server for Monthly Reporter (App2.0) + run history for the dashboard.

Run from repo root:
  PYTHONPATH=. uvicorn api.main:app --reload --port 8000

Vite proxies /api → 8000 (see dashboard/vite.config.ts).
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from agents.monthly_reporter.cloud_app.marketing_upload_layout import (  # noqa: E402
    write_marketing_csvs_to_work_dir,
)
from agents.monthly_reporter.cloud_app.ralph_runner import (  # noqa: E402
    ReportInputs,
    generate_monthly_report_bundle,
)
from agents.deepdive.agent import run as run_deepdive  # noqa: E402
from agents.marketingreco.agent import run as run_marketingreco  # noqa: E402
from agents.campaign_review.agent import run as run_campaign_review  # noqa: E402
from agents.campaign_review.agent import to_json_safe as campaign_review_to_json_safe  # noqa: E402
from agents.data_run.agent import run as run_data_run  # noqa: E402
from agents.marketingreco.ralph_ads_excel import ralph_ads_upload_rows  # noqa: E402
from shared.config.settings import account_information_csv_path  # noqa: E402
from shared.utils.account_directory import load_account_operators_csv  # noqa: E402

RUNS_BASE = ROOT / "data" / "runs" / "monthly_reporter"
RUNS_BASE.mkdir(parents=True, exist_ok=True)
INDEX_PATH = RUNS_BASE / "index.jsonl"

DD_RUNS_BASE = ROOT / "data" / "runs" / "deepdive"
DD_RUNS_BASE.mkdir(parents=True, exist_ok=True)
MRK_RUNS_BASE = ROOT / "data" / "runs" / "marketingreco"
MRK_RUNS_BASE.mkdir(parents=True, exist_ok=True)
CR_RUNS_BASE = ROOT / "data" / "runs" / "campaign_review"
CR_RUNS_BASE.mkdir(parents=True, exist_ok=True)
DATA_RUNS_BASE = ROOT / "data" / "runs" / "data_run"
DATA_RUNS_BASE.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="RalphAI API", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/account-directory")
def get_account_directory():
    """
    Unique operators from ``Business Name (original)`` with DoorDash login/password for dashboard autofill.
    Configure path via ``ACCOUNT_INFORMATION_CSV`` (defaults to repo-root ``Account Information-McDonalds.csv``).
    """
    path = account_information_csv_path()
    operators, warning = load_account_operators_csv(path)
    return {
        "path": str(path),
        "operators": operators,
        "warning": warning,
    }


def _append_index(rec: dict) -> None:
    with INDEX_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, default=str) + "\n")


def _prepare_ads_rows_file(input_path: Path, work_dir: Path) -> Path:
    """
    Normalize Ads manual upload into a CSV consumed by browser automation.

    - CSV: returned as-is.
    - Excel: reads sheet named "Ads" (case-insensitive), writes extracted CSV.
    """
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return input_path

    if suffix not in (".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"):
        raise HTTPException(400, "ads_sheet_file must be .csv or an Excel file.")

    try:
        import pandas as pd
    except ImportError as exc:
        raise HTTPException(500, "pandas is required to read Excel ads_sheet_file.") from exc

    try:
        xl = pd.ExcelFile(input_path)
    except Exception as exc:
        raise HTTPException(400, f"Failed to read Excel file: {exc}") from exc

    ads_sheet_name = next((s for s in xl.sheet_names if s.strip().lower() == "ads"), None)
    if not ads_sheet_name:
        raise HTTPException(400, 'Excel file must contain a sheet named "Ads".')

    try:
        ads_df = pd.read_excel(xl, sheet_name=ads_sheet_name)
    except Exception as exc:
        raise HTTPException(400, f'Failed to read "Ads" sheet: {exc}') from exc

    out_csv = work_dir / f"{input_path.stem}__ads_sheet.csv"
    ads_df.to_csv(out_csv, index=False)
    return out_csv


def _write_marketingreco_campaigns_excel(path: Path, result: dict) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offers"
    # store_id = National when FINANCIAL_DETAILED mapping exists; doordash_store_id = Reporting / Day-Slot key.
    headers = [
        "Store ID",
        "DoorDash Store ID",
        "Store Name",
        "Minimum Subtotal",
        "Slot Tags",
        "Campaign Name",
        "Status",
    ]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = Font(bold=True)
    mappings = result.get("campaign_mappings") or []
    for r, m in enumerate(mappings, start=2):
        tags = m.get("slot_tags", [])
        tags_str = ",".join(str(t) for t in tags) if isinstance(tags, list) else str(tags or "")
        ws.cell(row=r, column=1, value=m.get("store_id", ""))
        ws.cell(row=r, column=2, value=m.get("doordash_store_id", ""))
        ws.cell(row=r, column=3, value=m.get("store_name", ""))
        ws.cell(row=r, column=4, value=m.get("min_subtotal", 0))
        ws.cell(row=r, column=5, value=tags_str)
        ws.cell(row=r, column=6, value=m.get("campaign_name", ""))
        ws.cell(row=r, column=7, value=m.get("status", "Pending"))

    ads_plan = result.get("ads_plan") or {}
    slot_table = ads_plan.get("slot_table") or []
    if slot_table:
        wss = wb.create_sheet("Ads slots")
        sh = [
            "Merchant store ID",
            "Store name",
            "Slot",
            "Orders",
            "Sales",
            "Net total",
            "Profitability %",
            "Ad placement",
            "Budget estimate",
            "Weekly budget",
        ]
        for idx, h in enumerate(sh, start=1):
            cell = wss.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(slot_table, start=2):
            wss.cell(row=r, column=1, value=row.get("store_id"))
            wss.cell(row=r, column=2, value=row.get("store_name"))
            wss.cell(row=r, column=3, value=row.get("slot"))
            wss.cell(row=r, column=4, value=row.get("orders"))
            wss.cell(row=r, column=5, value=row.get("sales"))
            wss.cell(row=r, column=6, value=row.get("net_total"))
            wss.cell(row=r, column=7, value=row.get("profitability_pct"))
            wss.cell(row=r, column=8, value=row.get("ad_placement"))
            wss.cell(row=r, column=9, value=row.get("budget_estimate"))
            wss.cell(row=r, column=10, value=row.get("weekly_budget"))

    ralph_ads = ralph_ads_upload_rows(ads_plan)
    if ralph_ads:
        wsr = wb.create_sheet("Ads")
        rh = ["Merchant store ID", "Slots", "Bid strategy", "Budget", "Campaign Name"]
        for idx, h in enumerate(rh, start=1):
            cell = wsr.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, row in enumerate(ralph_ads, start=2):
            wsr.cell(row=r, column=1, value=row["store_id"])
            wsr.cell(row=r, column=2, value=row["slots"])
            wsr.cell(row=r, column=3, value=row["bid_strategy"])
            wsr.cell(row=r, column=4, value=row["budget"])
            wsr.cell(row=r, column=5, value=row["campaign_name"])

    campaigns = ads_plan.get("campaigns") or []
    if campaigns:
        wsa = wb.create_sheet("Ads planner")
        ah = [
            "store_id",
            "store_name",
            "day_of_week",
            "daypart",
            "tier",
            "priority_rank",
            "target_audience",
            "start_date",
            "end_date",
            "bid_strategy",
            "bid_amount",
            "bid_display",
            "budget_weight",
            "allocation_pct",
            "campaign_name",
            "rationale",
            "order_count",
            "avg_aov",
            "median_aov",
            "mode_basket",
            "avg_profitability",
            "profitability_pct",
            "ad_penetration",
            "composite_score",
        ]
        for idx, h in enumerate(ah, start=1):
            cell = wsa.cell(row=1, column=idx, value=h)
            cell.font = Font(bold=True)
        for r, c in enumerate(campaigns, start=2):
            m = c.get("metrics") or {}
            wsa.cell(row=r, column=1, value=c.get("store_id"))
            wsa.cell(row=r, column=2, value=c.get("store_name"))
            wsa.cell(row=r, column=3, value=c.get("day_of_week"))
            wsa.cell(row=r, column=4, value=c.get("daypart"))
            wsa.cell(row=r, column=5, value=c.get("tier"))
            wsa.cell(row=r, column=6, value=c.get("priority_rank"))
            wsa.cell(row=r, column=7, value=c.get("target_audience"))
            wsa.cell(row=r, column=8, value=c.get("start_date"))
            wsa.cell(row=r, column=9, value=c.get("end_date"))
            wsa.cell(row=r, column=10, value=c.get("bid_strategy"))
            wsa.cell(row=r, column=11, value=c.get("bid_amount"))
            wsa.cell(row=r, column=12, value=c.get("bid_display"))
            wsa.cell(row=r, column=13, value=c.get("budget_weight"))
            wsa.cell(row=r, column=14, value=c.get("allocation_pct"))
            wsa.cell(row=r, column=15, value=c.get("campaign_name"))
            wsa.cell(row=r, column=16, value=c.get("rationale"))
            wsa.cell(row=r, column=17, value=m.get("order_count"))
            wsa.cell(row=r, column=18, value=m.get("avg_aov"))
            wsa.cell(row=r, column=19, value=m.get("median_aov"))
            wsa.cell(row=r, column=20, value=m.get("mode_basket"))
            wsa.cell(row=r, column=21, value=m.get("avg_profitability"))
            wsa.cell(row=r, column=22, value=m.get("profitability_pct"))
            wsa.cell(row=r, column=23, value=m.get("ad_penetration"))
            wsa.cell(row=r, column=24, value=m.get("composite_score"))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _read_all_runs(limit: int = 200) -> list[dict]:
    if not INDEX_PATH.is_file():
        return []
    rows: list[dict] = []
    with INDEX_PATH.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return list(reversed(rows[-limit:]))


def _friendly_agent_name(agent: str) -> str:
    mapping = {
        "deepdive": "DeepDive",
        "marketingreco": "MarketingReco",
        "campaign_review": "Campaign Review",
        "data_run": "Data Run",
        "monthly_reporter": "Monthly Reporter",
        "offers": "RalphAI Offers",
        "ads": "RalphAI Ads",
    }
    return mapping.get((agent or "").strip().lower(), agent or "Unknown agent")


@app.get("/api/logs/live")
def get_live_logs(limit: int = 100) -> list[dict]:
    runs = _read_all_runs(limit=max(1, min(limit * 2, 500)))
    lines: list[dict] = []
    for run in runs:
        status = (run.get("status") or "").strip().lower()
        if status == "failed":
            level = "ERROR"
            status_text = "failed"
        elif status == "running":
            level = "WARN"
            status_text = "is running"
        else:
            level = "INFO"
            status_text = "completed"

        agent = _friendly_agent_name(str(run.get("agent") or ""))
        operator = str(run.get("operator") or "unknown operator")
        duration = str(run.get("duration") or "").strip()
        duration_suffix = f" in {duration}" if duration else ""
        lines.append(
            {
                "ts": str(run.get("started") or ""),
                "level": level,
                "msg": f"{agent}: run {status_text} for operator {operator}{duration_suffix}.",
            }
        )
        if len(lines) >= limit:
            break
    return lines


@app.get("/api/health")
def health() -> dict:
    return {"ok": True, "service": "ralphai-api"}


@app.get("/api/runs")
def list_runs() -> list[dict]:
    return _read_all_runs()


@app.get("/api/runs/{run_id}")
def get_run(run_id: str) -> dict:
    # Try monthly_reporter first
    meta_path = RUNS_BASE / run_id / "meta.json"
    if not meta_path.is_file():
        # Try deepdive
        meta_path = DD_RUNS_BASE / run_id / "meta.json"
        
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    return json.loads(meta_path.read_text(encoding="utf-8"))


@app.get("/api/runs/{run_id}/preview")
def get_preview(run_id: str) -> dict:
    p = RUNS_BASE / run_id / "preview.json"
    if not p.is_file():
        raise HTTPException(404, "Preview not found")
    return json.loads(p.read_text(encoding="utf-8"))


@app.get("/api/runs/{run_id}/download/full")
def download_full(run_id: str):
    folder = RUNS_BASE / run_id
    meta_path = folder / "meta.json"
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fn = meta.get("full_report_filename") or "report.xlsx"
    path = folder / fn
    if not path.is_file():
        raise HTTPException(404, "File missing")
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/runs/{run_id}/download/date")
def download_date(run_id: str):
    folder = RUNS_BASE / run_id
    meta_path = folder / "meta.json"
    if not meta_path.is_file():
        raise HTTPException(404, "Run not found")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    fn = meta.get("date_export_filename")
    if not fn:
        raise HTTPException(404, "Date export not available for this run")
    path = folder / fn
    if not path.is_file():
        raise HTTPException(404, "File missing")
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/runs/deepdive")
async def post_deepdive(
    operator_id: str = Form(...),
    zip_files: Optional[List[UploadFile]] = File(
        None,
        description="Required: one or more DoorDash export zip files for DeepDive.",
    ),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"dd_{run_id[:8]}_"))
    try:
        uploaded_names: list[str] = []
        if not zip_files:
            raise HTTPException(400, "Upload at least one DeepDive zip file.")

        for uploaded in zip_files:
            filename = (uploaded.filename or "").strip()
            if not filename:
                continue
            if not filename.lower().endswith(".zip"):
                raise HTTPException(400, f"Invalid file '{filename}'. Only .zip files are allowed.")
            raw = await uploaded.read()
            if not raw:
                continue
            safe_name = Path(filename).name
            (work / safe_name).write_bytes(raw)
            uploaded_names.append(safe_name)

        if not uploaded_names:
            raise HTTPException(400, "Upload at least one non-empty DeepDive zip file.")

        res = run_deepdive(operator_id=operator_id, data_dir=work)
        if res.get("status") != "success":
            raise HTTPException(400, res.get("message", "DeepDive failed"))

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        
        report_path = Path(res["report_html_path"])
        
        # Store in DD_RUNS_BASE / run_id
        out_dir = DD_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        
        shutil.copy(report_path, out_dir / "report.html")
        
        meta = {
            "run_id": run_id,
            "agent": "deepdive",
            "operator_id": operator_id,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "uploaded_files": uploaded_names,
            "datasets_loaded": res.get("datasets_loaded", []),
            "metric_hierarchy": (res.get("sections") or {}).get("metric_hierarchy") or {},
            "deepdive_json_path": res.get("deepdive_json_path"),
            "report_url": f"/api/runs/deepdive/{run_id}/report",
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        _append_index(
            {
                "id": run_id,
                "agent": "deepdive",
                "operator": operator_id,
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )

        return JSONResponse(meta)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.get("/api/runs/deepdive/{run_id}/report")
def get_deepdive_report(run_id: str):
    path = DD_RUNS_BASE / run_id / "report.html"
    if not path.is_file():
        raise HTTPException(404, "Report not found")
    return FileResponse(path)


@app.post("/api/runs/marketingreco")
async def post_marketingreco(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    financial_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mrk_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        kwargs: dict = {}

        if mode_norm == "manual":
            if not financial_file or not financial_file.filename:
                raise HTTPException(400, "Manual mode requires FINANCIAL_DETAILED file (.zip or .csv).")
            if not (
                financial_file.filename.lower().endswith(".zip")
                or financial_file.filename.lower().endswith(".csv")
            ):
                raise HTTPException(400, "financial_file must be .zip or .csv")
            raw = await financial_file.read()
            if not raw:
                raise HTTPException(400, "financial_file is empty.")
            in_path = work / Path(financial_file.filename).name
            in_path.write_bytes(raw)
            kwargs["financial_report_path"] = str(in_path)
            kwargs["reporting_root"] = str(ROOT / "Reporting-browser-use-claude-code")
        elif mode_norm == "auto":
            if not doordash_email.strip() or not doordash_password:
                raise HTTPException(400, "Auto mode requires doordash_email and doordash_password.")
            kwargs["doordash_email"] = doordash_email.strip()
            kwargs["doordash_password"] = doordash_password
            kwargs["reporting_root"] = str(ROOT / "Reporting-browser-use-claude-code")
        else:
            raise HTTPException(400, "mode must be 'manual' or 'auto'")

        result = run_marketingreco(operator_id=operator_id.strip(), mode=mode_norm, **kwargs)

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = MRK_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "marketingreco",
            "operator_id": operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "recommended_campaigns": len(result.get("recommended_campaigns") or []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
        campaigns_xlsx = out_dir / "marketingreco_campaigns.xlsx"
        _write_marketingreco_campaigns_excel(campaigns_xlsx, result)
        _append_index(
            {
                "id": run_id,
                "agent": "marketingreco",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        ads_plan_payload = result.get("ads_plan") or {}
        response = {
            **result,
            "run_id": run_id,
            "ads_upload_rows": ralph_ads_upload_rows(ads_plan_payload),
            "downloads": {
                "campaigns_excel": f"/api/runs/marketingreco/{run_id}/download/campaigns",
            },
        }
        return JSONResponse(response)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/runs/offers")
async def post_offers(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    campaign_mappings_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"offers_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        reporting_root = ROOT / "Reporting-browser-use-claude-code"
        env = os.environ.copy()
        if not doordash_email.strip() or not doordash_password:
            raise HTTPException(400, "Offers mode requires doordash_email and doordash_password.")
        env["DOORDASH_EMAIL"] = doordash_email.strip()
        env["DOORDASH_PASSWORD"] = doordash_password
        if mode_norm not in ("manual", "auto", "full"):
            raise HTTPException(400, "mode must be 'manual', 'auto', or 'full'")

        # Product behavior: Offers mode always runs the complete Reporting app pipeline
        # (download + analysis + campaign execution) with credentials provided via UI.
        _ = campaign_mappings_file
        subprocess.run(
            [sys.executable, "main.py"],
            cwd=str(reporting_root),
            env=env,
            check=True,
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "offers",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse({"status": "success", "mode": "full", "operator_id": operator_id.strip()})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/runs/ads")
async def post_ads(
    operator_id: str = Form(...),
    mode: str = Form("manual"),
    ads_sheet_file: Optional[UploadFile] = File(None),
    doordash_email: str = Form(""),
    doordash_password: str = Form(""),
):
    """
    Sponsored listing automation.

    Manual: CSV or any Excel file; for Excel, sheet "Ads" is read as input rows.
    Expected Ads columns: Merchant store ID (or Store ID) | Slots | Bid strategy | Budget | Campaign name.
    Auto: login → download financial + marketing reports → analysis + combined workbook (campaign
    recommendations) → build Ads upload rows from FINANCIAL_DETAILED → sponsored listing automation
    (same browser flow as Manual).
    """
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"ads_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        reporting_root = ROOT / "Reporting-browser-use-claude-code"
        if mode_norm not in ("manual", "auto"):
            raise HTTPException(400, "mode must be 'manual' or 'auto'")
        if not doordash_email.strip() or not doordash_password:
            raise HTTPException(400, "DoorDash email and password are required (browser login).")

        env = os.environ.copy()
        env["DOORDASH_EMAIL"] = doordash_email.strip()
        env["DOORDASH_PASSWORD"] = doordash_password

        rows_file: str | None = None

        if mode_norm == "manual":
            if not ads_sheet_file or not ads_sheet_file.filename:
                raise HTTPException(400, "Manual mode requires an ads sheet (.csv or Excel).")
            fn = ads_sheet_file.filename.lower()
            if not (
                fn.endswith(".csv")
                or fn.endswith(".xlsx")
                or fn.endswith(".xls")
                or fn.endswith(".xlsm")
                or fn.endswith(".xltx")
                or fn.endswith(".xltm")
            ):
                raise HTTPException(400, "ads_sheet_file must be .csv or an Excel file")

            raw = await ads_sheet_file.read()
            if not raw:
                raise HTTPException(400, "ads_sheet_file is empty.")

            sheet_path = work / Path(ads_sheet_file.filename).name
            sheet_path.write_bytes(raw)
            rows_path = _prepare_ads_rows_file(sheet_path, work)
            rows_file = rows_path.name

            env["ADS_DOWNLOAD_DIR"] = str(work)
            env["ADS_SHEET_PATH"] = str(rows_path)

            script = """
import asyncio
import os
from pathlib import Path
from agents.doordash_agent import run_ads_campaigns_from_sheet

async def _main():
    await run_ads_campaigns_from_sheet(
        download_dir=Path(os.environ["ADS_DOWNLOAD_DIR"]),
        email=os.environ["DOORDASH_EMAIL"],
        password=os.environ["DOORDASH_PASSWORD"],
        sheet_path=Path(os.environ["ADS_SHEET_PATH"]),
    )

asyncio.run(_main())
"""
            subprocess.run(
                [sys.executable, "-c", script],
                cwd=str(reporting_root),
                env=env,
                check=True,
            )
        elif mode_norm == "auto":
            env["RALPH_AI_ROOT"] = str(ROOT)
            ads_auto_script = """
import asyncio
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

from agents.doordash_agent import run_ads_campaigns_from_sheet, run_reports_only
from agents.marketing_agent import run as marketing_run
from agents.analysis_agent import run as analysis_run
from agents.combined_report_agent import run as combined_run, append_campaign_mappings_to_workbook
from agents.campaign_params import get_campaign_mappings_for_combined
import pandas as pd


def _dates():
    today = datetime.now().date()
    first_this_month = today.replace(day=1)
    last_prev_month = first_this_month - timedelta(days=1)
    y, m = first_this_month.year, first_this_month.month - 3
    if m <= 0:
        m += 12
        y -= 1
    start = datetime(y, m, 1).date()
    return start.strftime("%m/%d/%Y"), last_prev_month.strftime("%m/%d/%Y")


def _run_dir(email: str) -> Path:
    safe = (email or "run").strip()
    for c in ("@", ".", " ", "/", chr(92)):
        safe = safe.replace(c, "_")
    safe = safe[:50] if len(safe) > 50 else safe
    return Path("downloads") / f"{safe}-{datetime.now().strftime('%Y%m%d_%H%M%S')}"


async def _main():
    email = os.environ["DOORDASH_EMAIL"]
    password = os.environ["DOORDASH_PASSWORD"]
    ralph_root = Path(os.environ["RALPH_AI_ROOT"])
    sys.path.insert(0, str(ralph_root))
    from agents.marketingreco.ads_planner import build_ads_plan
    from agents.marketingreco.ralph_ads_excel import ralph_ads_upload_rows

    start_date, end_date = _dates()
    run_dir = _run_dir(email)
    run_dir.mkdir(parents=True, exist_ok=True)

    marketing_path, financial_path = await run_reports_only(
        download_dir=run_dir,
        email=email,
        password=password,
        start_date=start_date,
        end_date=end_date,
    )
    if not financial_path:
        raise SystemExit(
            "Ads auto: financial report was not downloaded. Check Browser Use / portal access."
        )

    marketing_sheets = (
        marketing_run(
            Path(marketing_path),
            output_dir=run_dir,
            post_start_date=start_date,
            post_end_date=end_date,
            write_file=False,
        )
        if marketing_path
        else None
    )
    financial_sheets = analysis_run(
        Path(financial_path),
        output_dir=run_dir,
        report_start_date=start_date,
        report_end_date=end_date,
        write_file=False,
    )

    combined = combined_run(
        financial_sheets=financial_sheets,
        marketing_sheets=marketing_sheets,
        output_dir=run_dir,
    )
    if combined:
        slots_csv = Path("slots.csv")
        mappings = get_campaign_mappings_for_combined(Path(combined), slots_csv)
        if mappings:
            append_campaign_mappings_to_workbook(Path(combined), mappings)

    fc = run_dir / "financial_detailed_report.csv"
    if not fc.is_file():
        for p in sorted(run_dir.glob("*FINANCIAL*.csv")):
            fc = p
            break
    if not fc.is_file():
        raise SystemExit(
            "Ads auto: no FINANCIAL_DETAILED CSV after analysis; cannot build ads recommendations."
        )

    ads_plan = build_ads_plan(str(fc))
    upload = ralph_ads_upload_rows(ads_plan)
    if not upload:
        raise SystemExit(
            "Ads auto: no sponsored-listing rows (no slots with Ad placement Yes). "
            "Try Manual with an Ads sheet or check financial data coverage."
        )

    ads_csv = run_dir / "ads_auto_upload.csv"
    pd.DataFrame(upload).to_csv(ads_csv, index=False)

    await run_ads_campaigns_from_sheet(
        download_dir=run_dir,
        email=email,
        password=password,
        sheet_path=ads_csv,
    )


asyncio.run(_main())
"""
            subprocess.run(
                [sys.executable, "-c", ads_auto_script],
                cwd=str(reporting_root),
                env=env,
                check=True,
            )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "ads",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        body: dict = {
            "status": "success",
            "run_id": run_id,
            "mode": mode_norm,
            "operator_id": operator_id.strip(),
        }
        if rows_file:
            body["rows_file"] = rows_file
        return JSONResponse(body)
    except HTTPException:
        raise
    except subprocess.CalledProcessError as e:
        raise HTTPException(500, f"Ads browser run failed (exit {e.returncode}). Check API logs / Slack.")
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.get("/api/runs/marketingreco/{run_id}/download/campaigns")
def download_marketingreco_campaigns(run_id: str):
    path = MRK_RUNS_BASE / run_id / "marketingreco_campaigns.xlsx"
    if not path.is_file():
        raise HTTPException(404, "Campaign table not found")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/runs/campaign-review")
async def post_campaign_review(
    operator_id: str = Form(...),
    mode: str = Form("auto"),
    marketing_files: Optional[List[UploadFile]] = File(None),
    data_dir: str = Form(""),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"cr_{run_id[:8]}_"))
    try:
        mode_norm = mode.strip().lower()
        if mode_norm not in ("auto", "manual"):
            raise HTTPException(400, "mode must be 'auto' or 'manual'")

        data_files: list[str] = []
        if mode_norm == "manual":
            if not marketing_files:
                raise HTTPException(
                    400,
                    "Manual mode requires marketing_files (MARKETING_PROMOTION* / MARKETING_SPONSORED_LISTING* csv/zip).",
                )
            for uf in marketing_files:
                if not uf.filename:
                    continue
                raw = await uf.read()
                if not raw:
                    continue
                p = work / Path(uf.filename).name
                p.write_bytes(raw)
                data_files.append(str(p))
            if not data_files:
                raise HTTPException(400, "No non-empty files uploaded for manual campaign review.")

        result = campaign_review_to_json_safe(
            run_campaign_review(
                operator_id=operator_id.strip(),
                mode=mode_norm,  # type: ignore[arg-type]
                data_dir=(data_dir.strip() or None),
                data_files=data_files if mode_norm == "manual" else None,
            )
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        out_dir = CR_RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "run_id": run_id,
            "agent": "campaign_review",
            "operator_id": operator_id.strip(),
            "mode": mode_norm,
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "campaign_reviews": len(result.get("campaign_reviews") or []),
            "datasets_loaded": (result.get("summary_metrics") or {}).get("datasets_loaded", []),
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        (out_dir / "result.json").write_text(
            json.dumps(result, indent=2, allow_nan=False),
            encoding="utf-8",
        )
        _append_index(
            {
                "id": run_id,
                "agent": "campaign_review",
                "operator": operator_id.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse({**result, "run_id": run_id})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/runs/data-run")
async def post_data_run(
    operator_ids: str = Form(..., description="JSON array or comma-separated operator IDs"),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    try:
        raw = (operator_ids or "").strip()
        if not raw:
            raise HTTPException(400, "Select at least one operator.")
        parsed_ids: list[str]
        if raw.startswith("["):
            try:
                as_json = json.loads(raw)
                if not isinstance(as_json, list):
                    raise ValueError
                parsed_ids = [str(v).strip() for v in as_json if str(v).strip()]
            except Exception as exc:
                raise HTTPException(400, "operator_ids JSON must be an array of operator IDs") from exc
        else:
            parsed_ids = [s.strip() for s in raw.split(",") if s.strip()]
        if not parsed_ids:
            raise HTTPException(400, "Select at least one operator.")

        result = run_data_run(
            operator_ids=parsed_ids,
            reporting_root=str(ROOT / "Reporting-browser-use-claude-code"),
        )

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "data_run",
                "operator": f"{len(parsed_ids)} selected",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )
        return JSONResponse(
            {
                "status": "success",
                "run_id": run_id,
                "file_type": "both",
                "selected_operator_count": len(parsed_ids),
                **result,
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/runs/monthly-reporter")
async def post_monthly_reporter(
    pre_range: str = Form(..., description="MM/DD/YYYY-MM/DD/YYYY"),
    post_range: str = Form(...),
    operator_id: str = Form(""),
    operator_name: str = Form(""),
    excluded_dates: str = Form(""),
    dd_store_ids: str = Form(""),
    ue_store_ids: str = Form(""),
    dd_file: Optional[UploadFile] = File(
        None,
        description="Optional DoorDash financial CSV (saved as dd-data.csv when provided).",
    ),
    ue_file: Optional[UploadFile] = File(
        None,
        description="Optional UberEats financial CSV (saved as ue-data.csv when provided).",
    ),
    marketing_files: Optional[List[UploadFile]] = File(
        None,
        description="Optional: multiple MARKETING_*.csv (Streamlit file_upload_screen behavior)",
    ),
):
    run_id = str(uuid.uuid4())
    t0 = datetime.now(timezone.utc)
    work = Path(tempfile.mkdtemp(prefix=f"mr_{run_id[:8]}_"))

    try:
        # Streamlit parity: financial + marketing files are optional; only Pre/Post dates are required.
        if dd_file and dd_file.filename:
            raw = await dd_file.read()
            if raw:
                (work / "dd-data.csv").write_bytes(raw)
        if ue_file and ue_file.filename:
            raw = await ue_file.read()
            if raw:
                (work / "ue-data.csv").write_bytes(raw)

        # Marketing CSVs — same layout as Streamlit `file_upload_screen` (marketing_data/marketing_*).
        mkt_pairs: list[tuple[str, bytes]] = []
        if marketing_files:
            for uf in marketing_files:
                if uf.filename:
                    raw = await uf.read()
                    if raw:
                        mkt_pairs.append((uf.filename, raw))
        if mkt_pairs:
            write_marketing_csvs_to_work_dir(work, mkt_pairs)

        inputs = ReportInputs(
            pre_range=pre_range.strip(),
            post_range=post_range.strip(),
            excluded_dates_text=excluded_dates.strip(),
            operator_name=operator_name.strip(),
            dd_store_ids_text=dd_store_ids.strip(),
            ue_store_ids_text=ue_store_ids.strip(),
        )

        bundle = generate_monthly_report_bundle(inputs, data_root=work)

        out_dir = RUNS_BASE / run_id
        out_dir.mkdir(parents=True, exist_ok=True)

        full_name = bundle["filename"]
        (out_dir / full_name).write_bytes(bundle["excel_bytes"])

        date_name = bundle.get("date_export_filename")
        if bundle.get("date_export_bytes") and date_name:
            (out_dir / date_name).write_bytes(bundle["date_export_bytes"])  # type: ignore[index]

        preview = {"tables": bundle.get("tables") or {}, "summary_text": bundle.get("summary_text")}
        (out_dir / "preview.json").write_text(json.dumps(preview, default=str), encoding="utf-8")

        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()

        meta = {
            "run_id": run_id,
            "agent": "monthly_reporter",
            "operator_id": operator_id.strip() or "—",
            "operator_name": operator_name.strip(),
            "status": "success",
            "started": t0.isoformat(),
            "duration_s": round(duration_s, 2),
            "summary_text": bundle.get("summary_text"),
            "full_report_filename": full_name,
            "date_export_filename": date_name if bundle.get("date_export_bytes") else None,
        }
        (out_dir / "meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

        _append_index(
            {
                "id": run_id,
                "agent": "monthly_reporter",
                "operator": operator_id.strip() or operator_name.strip() or "—",
                "status": "success",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s // 60)}m {int(duration_s % 60):02d}s",
            }
        )

        return JSONResponse(
            {
                "run_id": run_id,
                "summary_text": bundle.get("summary_text"),
                "preview": preview,
                "downloads": {
                    "full": f"/api/runs/{run_id}/download/full",
                    "date": f"/api/runs/{run_id}/download/date"
                    if bundle.get("date_export_bytes")
                    else None,
                },
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        duration_s = (datetime.now(timezone.utc) - t0).total_seconds()
        _append_index(
            {
                "id": run_id,
                "agent": "monthly_reporter",
                "operator": operator_id.strip() or "—",
                "status": "failed",
                "started": t0.isoformat().replace("+00:00", "Z")[:19].replace("T", " "),
                "duration": f"{int(duration_s)}s",
                "error": str(e),
            }
        )
        raise HTTPException(500, str(e)) from e
    finally:
        shutil.rmtree(work, ignore_errors=True)
