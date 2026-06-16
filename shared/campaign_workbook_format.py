"""Shared campaign workbook naming — aligned with reporting_browser_use combined_report_agent."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

logger = logging.getLogger(__name__)

CampaignKind = Literal["offers", "ads"]

# reporting_browser_use convention
COMBINED_ANALYSIS_PREFIX = "combined_analysis_"
COMBINED_ANALYSIS_GLOB = "combined_analysis_*.xlsx"
CAMPAIGN_MAPPINGS_SHEET = "Campaign Mappings"
ADS_CAMPAIGN_MAPPINGS_SHEET = "Ads Campaign Mappings"

CAMPAIGN_MAPPINGS_HEADERS = (
    "Store ID",
    "Store Name",
    "Minimum Subtotal",
    "Slot Tags",
    "Campaign Name",
    "Status",
)
ADS_CAMPAIGN_MAPPINGS_HEADERS = (
    "Store ID",
    "Store Name",
    "Minimum Bid",
    "Weekly Budget",
    "Slot Tags",
    "Campaign Name",
    "Status",
)

# Legacy Strategist sheets (still accepted on read)
_LEGACY_OFFERS_SHEET_NAMES = ("offers campaigns", "offers")
_LEGACY_ADS_SHEET_NAMES = ("ads campaigns", "ads")

def combined_analysis_filename(*, timestamp: str | None = None) -> str:
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{COMBINED_ANALYSIS_PREFIX}{ts}.xlsx"


def is_combined_analysis_path(path: Path | str) -> bool:
    name = Path(path).name.lower()
    return name.startswith(COMBINED_ANALYSIS_PREFIX) and name.endswith(".xlsx")


def find_latest_combined_analysis(*search_dirs: Path) -> Path | None:
    """Return newest combined_analysis_*.xlsx under any of the given directories."""
    candidates: list[Path] = []
    for base in search_dirs:
        if not base or not Path(base).is_dir():
            continue
        candidates.extend(Path(base).glob(COMBINED_ANALYSIS_GLOB))
    if not candidates:
        return None
    return sorted(candidates, key=lambda p: p.name, reverse=True)[0]


def pick_mappings_sheet(sheet_names: list[str], kind: CampaignKind) -> str:
    """Resolve sheet name for offers or ads mappings (reporting names first)."""
    lowered = {s: s.strip().lower() for s in sheet_names}
    primary = CAMPAIGN_MAPPINGS_SHEET if kind == "offers" else ADS_CAMPAIGN_MAPPINGS_SHEET
    if primary.lower() in lowered.values():
        for original, low in lowered.items():
            if low == primary.lower():
                return original
    legacy = _LEGACY_OFFERS_SHEET_NAMES if kind == "offers" else _LEGACY_ADS_SHEET_NAMES
    for target in legacy:
        for original, low in lowered.items():
            if low == target:
                return original
    raise ValueError(
        f'Workbook has no {"Offers" if kind == "offers" else "Ads"} mappings sheet '
        f'(looked for: {primary}, {", ".join(legacy)}). Sheets: {sheet_names}'
    )


def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").strip().lower())


def _parse_slot_tags(raw: Any) -> list[int]:
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        parts = raw
    else:
        parts = re.split(r"[,;\s]+", str(raw).strip())
    out: list[int] = []
    for p in parts:
        s = str(p).strip()
        if not s:
            continue
        try:
            out.append(int(float(s)))
        except ValueError:
            continue
    return sorted(set(out))


def _row_val(row: dict[str, Any], *candidates: str) -> Any:
    if not row:
        return None
    by_norm = {_norm_col(k): v for k, v in row.items()}
    for c in candidates:
        key = _norm_col(c)
        if key in by_norm:
            return by_norm[key]
    return None


def read_offer_combos_from_workbook(workbook: Path) -> list[dict[str, Any]]:
    """Read offer combos from Campaign Mappings (or legacy Offers sheet)."""
    return _read_offer_combos_pandas(Path(workbook))


def _read_offer_combos_pandas(path: Path) -> list[dict[str, Any]]:
    try:
        import pandas as pd
    except ImportError:
        return []
    xl = pd.ExcelFile(path)
    sheet = pick_mappings_sheet(list(xl.sheet_names), "offers")
    df = pd.read_excel(xl, sheet_name=sheet)
    if df is None or df.empty:
        return []
    combos: list[dict[str, Any]] = []
    for rec in df.to_dict(orient="records"):
        cleaned = {k: ("" if pd.isna(v) else v) for k, v in rec.items()}
        store_id = str(_row_val(cleaned, "Store ID", "Merchant store ID") or "").strip()
        if not store_id:
            continue
        slot_tags = _parse_slot_tags(_row_val(cleaned, "Slot Tags", "Slots"))
        if not slot_tags:
            continue
        try:
            min_sub = int(round(float(_row_val(cleaned, "Minimum Subtotal", "Min subtotal") or 10)))
        except (TypeError, ValueError):
            min_sub = 10
        campaign_name = str(
            _row_val(cleaned, "Campaign Name", "Campaign name") or f"TODC-{store_id}-${min_sub}"
        ).strip()
        combos.append(
            {
                "store_id": store_id,
                "store_name": str(_row_val(cleaned, "Store Name", "Store name") or "").strip(),
                "min_subtotal": min_sub,
                "slot_tags": slot_tags,
                "campaign_name": campaign_name,
                "status": str(_row_val(cleaned, "Status") or "Pending"),
            }
        )
    return combos


def read_ads_rows_from_workbook(workbook: Path) -> list[dict[str, Any]]:
    """Read ads rows from Ads Campaign Mappings (or legacy Ads sheet)."""
    path = Path(workbook)
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas is required to read campaign workbooks") from exc

    xl = pd.ExcelFile(path)
    sheet = pick_mappings_sheet(list(xl.sheet_names), "ads")
    df = pd.read_excel(xl, sheet_name=sheet)
    if df is None or df.empty:
        return []
    rows_out: list[dict[str, Any]] = []
    for rec in df.to_dict(orient="records"):
        cleaned = {k: ("" if pd.isna(v) else v) for k, v in rec.items()}
        store_id = str(
            _row_val(cleaned, "Store ID", "Merchant store ID", "Merchant Store ID") or ""
        ).strip()
        if not store_id:
            continue
        slot_tags = _parse_slot_tags(_row_val(cleaned, "Slot Tags", "Slots"))
        if not slot_tags:
            continue
        try:
            bid = float(_row_val(cleaned, "Minimum Bid", "Bid strategy", "Bid Strategy") or 3)
        except (TypeError, ValueError):
            bid = 3.0
        try:
            budget = float(
                _row_val(cleaned, "Weekly Budget", "Budget", "weekly_budget", "Weekly budget") or 0
            )
        except (TypeError, ValueError):
            budget = 0.0
        campaign_name = str(
            _row_val(cleaned, "Campaign Name", "Campaign name") or f"TODC-ADS-{store_id}"
        ).strip()
        rows_out.append(
            {
                "store_id": store_id,
                "store_name": str(_row_val(cleaned, "Store Name", "Store name") or "").strip(),
                "slot_tags": slot_tags,
                "bid_strategy": bid,
                "budget": budget,
                "campaign_name": campaign_name,
                "status": str(_row_val(cleaned, "Status") or "Pending"),
            }
        )
    return rows_out


def write_mappings_sheets(
    workbook_path: Path,
    *,
    offer_rows: list[dict[str, Any]],
    ads_rows: list[dict[str, Any]],
    store_names: dict[str, str],
) -> None:
    """Create or replace Campaign Mappings + Ads Campaign Mappings sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write campaign mappings") from exc

    path = Path(workbook_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.is_file():
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    for sheet_title in (CAMPAIGN_MAPPINGS_SHEET[:31], ADS_CAMPAIGN_MAPPINGS_SHEET[:31]):
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]

    ws_offers = wb.create_sheet(CAMPAIGN_MAPPINGS_SHEET[:31])
    for idx, h in enumerate(CAMPAIGN_MAPPINGS_HEADERS, start=1):
        ws_offers.cell(row=1, column=idx, value=h).font = Font(bold=True)
    r = 2
    for store_id in sorted({row["store_id"] for row in offer_rows}):
        store_offers = [row for row in offer_rows if row["store_id"] == store_id]
        for row in store_offers:
            ws_offers.cell(row=r, column=1, value=store_id)
            ws_offers.cell(row=r, column=2, value=store_names.get(store_id, row.get("store_name", "")))
            ws_offers.cell(row=r, column=3, value=row["min_subtotal"])
            ws_offers.cell(row=r, column=4, value=row["slot_tags"])
            ws_offers.cell(row=r, column=5, value=row["campaign_name"])
            ws_offers.cell(row=r, column=6, value=row.get("status", "Pending"))
            r += 1

    ws_ads = wb.create_sheet(ADS_CAMPAIGN_MAPPINGS_SHEET[:31])
    for idx, h in enumerate(ADS_CAMPAIGN_MAPPINGS_HEADERS, start=1):
        ws_ads.cell(row=1, column=idx, value=h).font = Font(bold=True)
    r = 2
    for row in ads_rows:
        ws_ads.cell(row=r, column=1, value=row["store_id"])
        ws_ads.cell(row=r, column=2, value=store_names.get(row["store_id"], row.get("store_name", "")))
        ws_ads.cell(row=r, column=3, value=row.get("bid_strategy", 3))
        ws_ads.cell(row=r, column=4, value=row.get("budget", 0))
        ws_ads.cell(row=r, column=5, value=row["slot_tags"])
        ws_ads.cell(row=r, column=6, value=row["campaign_name"])
        ws_ads.cell(row=r, column=7, value=row.get("status", "Pending"))
        r += 1

    wb.save(path)


def update_mapping_status(
    workbook_path: Path | str,
    campaign_name: str,
    status: str,
    *,
    kind: CampaignKind = "offers",
) -> bool:
    """Update Status on Campaign Mappings or Ads Campaign Mappings sheet."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed; cannot update mapping status")
        return False

    path = Path(workbook_path)
    if not path.is_file():
        return False

    name = str(campaign_name or "").strip()
    if not name:
        return False

    try:
        wb = openpyxl.load_workbook(path, read_only=False)
    except Exception as exc:
        logger.warning("Could not open %s for status update: %s", path, exc)
        return False

    try:
        sheet_name = pick_mappings_sheet(list(wb.sheetnames), kind)[:31]
        ws = wb[sheet_name]
        headers: dict[str, int] = {}
        for cell in ws[1]:
            if cell.value is not None and str(cell.value).strip():
                headers[_norm_col(str(cell.value))] = cell.column
        name_col = headers.get(_norm_col("Campaign Name")) or headers.get(_norm_col("Campaign name"))
        status_col = headers.get(_norm_col("Status"))
        if not name_col or not status_col:
            logger.warning(
                "Campaign workbook %s missing Campaign Name or Status column (sheet %s)",
                path.name,
                sheet_name,
            )
            return False

        for row_idx in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=name_col).value
            if str(cell_name or "").strip() == name:
                ws.cell(row=row_idx, column=status_col).value = status
                wb.save(path)
                logger.info("Campaign workbook: %s → %s (%s)", name, status, path.name)
                return True
        logger.warning("Campaign '%s' not found in %s for status update", name, path.name)
        return False
    except Exception as exc:
        logger.warning("Status update failed for %s in %s: %s", name, path, exc)
        return False
    finally:
        wb.close()
